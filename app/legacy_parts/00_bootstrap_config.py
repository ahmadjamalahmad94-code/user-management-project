
"""
Hobe Hub Professional+ Edition
- Sidebar UI
- Dashboard stats
- Pagination
- Audit log
- DB-based auth + permissions
- User profile / password change
- Beneficiaries management
- Smart CSV import / export
- Backup SQL (table-level)
"""

from __future__ import annotations

import csv
import hashlib
import hmac
import io
import json
import logging
import math
import os
import re
import secrets
import sqlite3
import threading
from datetime import date, timedelta, datetime, timezone
from zoneinfo import ZoneInfo
from functools import wraps
from uuid import uuid4

from flask import (
    Flask,
    Response,
    flash,
    jsonify,
    redirect,
    render_template,
    render_template_string,
    request,
    session,
    url_for,
)
import psycopg2
from psycopg2 import pool
from psycopg2.extras import Json, RealDictCursor, execute_batch, execute_values
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from app.services.adv_client_api import AdvClientApi, summarize_adv_client_test
from app.services.radius_api import RadiusApiClient, RadiusApiError, default_session_expiry, mask_sensitive_data
from markupsafe import escape
from werkzeug.security import check_password_hash, generate_password_hash

try:
    from dotenv import load_dotenv
except ImportError:  # pragma: no cover - dependency may be absent in old installs
    load_dotenv = None

if load_dotenv:
    load_dotenv()


def env_flag(name: str, default: str = "") -> bool:
    return os.getenv(name, default).strip().lower() in {"1", "true", "yes", "on"}


def is_production_env() -> bool:
    return (
        os.getenv("FLASK_ENV", "").strip().lower() == "production"
        or os.getenv("HOBEHUB_ENV", "").strip().lower() == "production"
        or env_flag("RENDER")
    )


def get_secret_key() -> str:
    secret_key = os.getenv("FLASK_SECRET_KEY", "").strip()
    if secret_key:
        return secret_key
    if is_production_env():
        raise RuntimeError("FLASK_SECRET_KEY is required in production.")
    return "hobehub-local-dev-secret-change-me"


app = Flask(__name__)
app.secret_key = get_secret_key()
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=is_production_env(),
)

DEFAULT_DATABASE_URL = os.getenv("DEFAULT_DATABASE_URL", "").strip()
LOCAL_DEMO_DB_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "instance", "hobehub_local_demo.sqlite3")

DB_POOL_MINCONN = int(os.getenv("DB_POOL_MINCONN", "1"))
DB_POOL_MAXCONN = int(os.getenv("DB_POOL_MAXCONN", "12"))
_connection_pool = None
IMPORT_TASKS = {}
IMPORT_TASKS_LOCK = threading.Lock()
IMPORT_LOG_LIMIT = 300
IMPORT_BATCH_SIZE = int(os.getenv("IMPORT_BATCH_SIZE", "250"))
AUTH_FAILURES = {}
AUTH_FAILURE_LIMIT = int(os.getenv("AUTH_FAILURE_LIMIT", "5"))
AUTH_LOCK_MINUTES = int(os.getenv("AUTH_LOCK_MINUTES", "15"))
MIKROTIK_HOTSPOT_URL = os.getenv("MIKROTIK_HOTSPOT_URL", "").strip().rstrip("/")

APP_TZ = ZoneInfo("Asia/Gaza")

def now_local() -> datetime:
    return datetime.now(APP_TZ)

def today_local() -> date:
    return now_local().date()


def get_database_url() -> str:
    if env_flag("HOBEHUB_LOCAL_DEMO"):
        return f"sqlite:///{os.getenv('HOBEHUB_LOCAL_DB_PATH', LOCAL_DEMO_DB_PATH)}"
    database_url = (os.getenv("DATABASE_URL") or DEFAULT_DATABASE_URL).strip()
    if not database_url:
        if is_production_env() or env_flag("HOBEHUB_STRICT_CONFIG"):
            raise RuntimeError("DATABASE_URL is required outside local demo mode.")
        return f"sqlite:///{os.getenv('HOBEHUB_LOCAL_DB_PATH', LOCAL_DEMO_DB_PATH)}"
    if database_url.startswith("postgresql://") and "sslmode=" not in database_url:
        separator = "&" if "?" in database_url else "?"
        database_url = f"{database_url}{separator}sslmode=require"
    return database_url


def is_sqlite_database_url(database_url: str | None = None) -> bool:
    return (database_url or get_database_url()).startswith("sqlite:///")


def get_sqlite_db_path(database_url: str | None = None) -> str:
    url = database_url or get_database_url()
    if url.startswith("sqlite:///"):
        return url.replace("sqlite:///", "", 1)
    return os.getenv("HOBEHUB_LOCAL_DB_PATH", LOCAL_DEMO_DB_PATH)


def is_local_demo_mode() -> bool:
    return is_sqlite_database_url()
