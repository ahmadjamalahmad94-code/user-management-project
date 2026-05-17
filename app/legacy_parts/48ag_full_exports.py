# 48ag_full_exports.py
# تصدير شامل لكل أنواع البيانات بصيغة XLSX
# كل دالة تصدير تتلقى الجدول والأعمدة + تنسيق موحد ذهبي/أسود

import io
import logging
from flask import Response

_log = logging.getLogger("hobehub.full_exports")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_OK = True
except Exception:
    _OPENPYXL_OK = False


def _xlsx_response(filename, sheet_title, headers_ar, rows, action="export", entity="generic"):
    """يبني ملف XLSX من البيانات. headers_ar = [(field_or_callable, header_label), ...]."""
    if not _OPENPYXL_OK:
        return Response("openpyxl غير مثبت — رفع المكتبة لتفعيل التصدير.", status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or "Export"

    title_fill = PatternFill(fill_type="solid", fgColor="1E1E1E")
    header_fill = PatternFill(fill_type="solid", fgColor="FDF1CF")
    header_font = Font(bold=True, color="1E1E1E", size=11)
    title_font = Font(color="F4BA2A", bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top = Alignment(vertical="top", wrap_text=True)
    thin_side = Side(style="thin", color="E6E3DA")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Title (merged row 1)
    n = len(headers_ar)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    title_cell = ws.cell(row=1, column=1, value=sheet_title + " — Hobe Hub")
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # Headers (row 2)
    for col_idx, (_, label) in enumerate(headers_ar, start=1):
        c = ws.cell(row=2, column=col_idx, value=label)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border
    ws.row_dimensions[2].height = 22

    # Data (row 3+)
    for row_idx, row in enumerate(rows or [], start=3):
        d = dict(row) if not isinstance(row, dict) else row
        for col_idx, (field, _) in enumerate(headers_ar, start=1):
            if callable(field):
                val = field(d)
            elif field is None:
                val = row_idx - 2  # #
            else:
                val = d.get(field, "")
            if val is None:
                val = ""
            c = ws.cell(row=row_idx, column=col_idx, value=str(val))
            c.alignment = top
            c.border = border

    # Freeze + auto filter + widths
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n)}{max(ws.max_row, 2)}"
    for col_idx in range(1, n + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 38)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    log_action(action, entity, None, f"تصدير XLSX: {filename}")
    resp = Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return resp


# ─── 1. حسابات البوابة (مع كلمات المرور) ───
@app.route("/admin/exports/portal-accounts")
@admin_login_required
@permission_required("export")
def export_portal_accounts():
    rows = query_all(
        """
        SELECT pa.id, pa.username, pa.password_plain, pa.is_active, pa.must_set_password,
               pa.last_login_at, pa.activated_at, pa.created_at,
               b.id AS beneficiary_id, b.full_name, b.phone, b.user_type
        FROM beneficiary_portal_accounts pa
        JOIN beneficiaries b ON b.id = pa.beneficiary_id
        ORDER BY pa.id DESC
        """
    ) or []
    headers = [
        (None, "#"),
        ("id", "ID الحساب"),
        ("beneficiary_id", "ID المشترك"),
        ("full_name", "الاسم"),
        ("phone", "الجوال"),
        ("user_type", "النوع"),
        ("username", "اسم المستخدم"),
        ("password_plain", "كلمة المرور"),
        (lambda r: "نشط" if r.get("is_active") else "معطل", "الحالة"),
        (lambda r: "نعم" if r.get("must_set_password") else "—", "مصفّر؟"),
        ("last_login_at", "آخر دخول"),
        ("activated_at", "تاريخ التفعيل"),
        ("created_at", "تاريخ الإنشاء"),
    ]
    return _xlsx_response(
        "portal_accounts.xlsx", "حسابات البوابة", headers, rows,
        action="export_portal_accounts", entity="beneficiary_portal_account",
    )


# ─── 2. فئات البطاقات ───
@app.route("/admin/exports/card-categories")
@admin_login_required
@permission_required("export")
def export_card_categories():
    rows = query_all(
        "SELECT * FROM card_categories ORDER BY display_order ASC, duration_minutes ASC"
    ) or []
    headers = [
        (None, "#"),
        ("id", "ID"),
        ("code", "الكود"),
        ("label_ar", "الاسم بالعربية"),
        ("duration_minutes", "المدة (دق)"),
        ("display_order", "الترتيب"),
        ("icon", "الأيقونة"),
        ("radius_profile_id", "RADIUS profile"),
        (lambda r: "مفعلة" if r.get("is_active") else "معطلة", "الحالة"),
        ("created_at", "تاريخ الإنشاء"),
    ]
    return _xlsx_response(
        "card_categories.xlsx", "فئات البطاقات", headers, rows,
        action="export_card_categories", entity="card_categories",
    )


# ─── 3. سياسات الحصص ───
@app.route("/admin/exports/quota-policies")
@admin_login_required
@permission_required("export")
def export_quota_policies():
    rows = query_all(
        "SELECT * FROM card_quota_policies ORDER BY priority ASC, id DESC"
    ) or []
    scope_map = {"default": "افتراضية", "user": "مشترك", "group": "مجموعة"}
    headers = [
        (None, "#"),
        ("id", "ID"),
        (lambda r: scope_map.get(r.get("scope"), r.get("scope") or ""), "النطاق"),
        ("target_id", "المستهدف ID"),
        ("daily_limit", "الحد اليومي"),
        ("weekly_limit", "الحد الأسبوعي"),
        ("allowed_days", "الأيام المسموحة"),
        ("allowed_category_codes", "الفئات المسموحة"),
        ("priority", "الأولوية"),
        ("valid_from", "صالحة من"),
        ("valid_until", "صالحة حتى"),
        ("valid_time_from", "ساعة بداية الدوام"),
        ("valid_time_until", "ساعة نهاية الدوام"),
        (lambda r: "مفعلة" if r.get("is_active") else "معطلة", "الحالة"),
        ("notes", "ملاحظات"),
        ("created_at", "تاريخ الإنشاء"),
    ]
    return _xlsx_response(
        "quota_policies.xlsx", "سياسات الحصص", headers, rows,
        action="export_quota_policies", entity="card_quota_policies",
    )


# ─── 4. مجموعات السياسات + الأعضاء ───
@app.route("/admin/exports/quota-groups")
@admin_login_required
@permission_required("export")
def export_quota_groups():
    rows = query_all(
        """
        SELECT g.id, g.name, g.description, g.created_at, g.created_by,
               (SELECT COUNT(*) FROM quota_group_members m WHERE m.group_id = g.id) AS members_count
        FROM quota_groups g ORDER BY g.id DESC
        """
    ) or []
    headers = [
        (None, "#"),
        ("id", "ID المجموعة"),
        ("name", "الاسم"),
        ("description", "الوصف"),
        ("members_count", "عدد الأعضاء"),
        ("created_by", "أنشأها"),
        ("created_at", "تاريخ الإنشاء"),
    ]
    return _xlsx_response(
        "quota_groups.xlsx", "مجموعات السياسات", headers, rows,
        action="export_quota_groups", entity="quota_groups",
    )


@app.route("/admin/exports/quota-group-members")
@admin_login_required
@permission_required("export")
def export_quota_group_members():
    rows = query_all(
        """
        SELECT g.id AS group_id, g.name AS group_name,
               m.beneficiary_id, m.added_at,
               b.full_name, b.phone, b.user_type
        FROM quota_group_members m
        JOIN quota_groups g ON g.id = m.group_id
        JOIN beneficiaries b ON b.id = m.beneficiary_id
        ORDER BY g.id, m.added_at DESC
        """
    ) or []
    headers = [
        (None, "#"),
        ("group_id", "ID المجموعة"),
        ("group_name", "اسم المجموعة"),
        ("beneficiary_id", "ID المشترك"),
        ("full_name", "اسم المشترك"),
        ("phone", "الجوال"),
        ("user_type", "النوع"),
        ("added_at", "تاريخ الإضافة"),
    ]
    return _xlsx_response(
        "quota_group_members.xlsx", "أعضاء مجموعات السياسات", headers, rows,
        action="export_quota_group_members", entity="quota_group_members",
    )


# ─── 5. سجل العمليات (audit_logs) ───
@app.route("/admin/exports/audit-logs")
@admin_login_required
@permission_required("export")
def export_audit_logs():
    rows = query_all(
        "SELECT * FROM audit_logs ORDER BY id DESC LIMIT 10000"
    ) or []
    headers = [
        (None, "#"),
        ("id", "ID"),
        ("action_type", "الإجراء"),
        ("target_type", "الكيان"),
        ("target_id", "معرّف الكيان"),
        ("details", "التفاصيل"),
        ("username_snapshot", "بواسطة"),
        ("created_at", "التاريخ"),
    ]
    return _xlsx_response(
        "audit_logs.xlsx", "سجل العمليات", headers, rows,
        action="export_audit_logs", entity="audit_logs",
    )


# ─── 6. سجل البطاقات الصادرة (usage_logs) ───
@app.route("/admin/exports/usage-logs")
@admin_login_required
@permission_required("export")
def export_usage_logs():
    rows = query_all(
        """
        SELECT u.*, b.full_name AS beneficiary_name, b.phone
        FROM beneficiary_usage_logs u
        LEFT JOIN beneficiaries b ON b.id = u.beneficiary_id
        ORDER BY u.id DESC LIMIT 20000
        """
    ) or []
    headers = [
        (None, "#"),
        ("id", "ID"),
        ("beneficiary_id", "ID المشترك"),
        ("beneficiary_name", "اسم المشترك"),
        ("phone", "الجوال"),
        ("usage_reason", "السبب"),
        ("card_type", "الفئة"),
        ("usage_date", "تاريخ الإصدار"),
        ("usage_time", "وقت الإصدار"),
        ("added_by_username", "أصدرها"),
        ("notes", "ملاحظات"),
    ]
    return _xlsx_response(
        "usage_logs.xlsx", "سجل البطاقات الصادرة", headers, rows,
        action="export_usage_logs", entity="beneficiary_usage_logs",
    )


# ─── 7. مخزون البطاقات (manual_access_cards) ───
@app.route("/admin/exports/cards-inventory")
@admin_login_required
@permission_required("export")
def export_cards_inventory():
    rows = query_all(
        "SELECT * FROM manual_access_cards ORDER BY id DESC LIMIT 50000"
    ) or []
    headers = [
        (None, "#"),
        ("id", "ID"),
        ("category_code", "الفئة"),
        ("duration_minutes", "المدة (دق)"),
        ("username", "اسم البطاقة"),
        ("password", "كلمة المرور"),
        (lambda r: "مسلمة" if r.get("is_issued") else "متاحة", "الحالة"),
        ("issued_to_beneficiary_id", "سُلمت لـ ID"),
        ("issued_at", "تاريخ التسليم"),
        ("source_filename", "ملف الاستيراد"),
        ("created_at", "تاريخ الإنشاء"),
    ]
    return _xlsx_response(
        "cards_inventory.xlsx", "مخزون البطاقات", headers, rows,
        action="export_cards_inventory", entity="manual_access_cards",
    )


# ─── 8. حسابات الإدارة ───
@app.route("/admin/exports/admin-accounts")
@admin_login_required
@permission_required("export")
def export_admin_accounts():
    rows = query_all("SELECT * FROM app_accounts ORDER BY id DESC") or []
    headers = [
        (None, "#"),
        ("id", "ID"),
        ("username", "اسم المستخدم"),
        ("full_name", "الاسم الكامل"),
        (lambda r: "نشط" if r.get("is_active") else "معطل", "الحالة"),
        ("last_login_at", "آخر دخول"),
        ("created_at", "تاريخ الإنشاء"),
    ]
    return _xlsx_response(
        "admin_accounts.xlsx", "حسابات الإدارة", headers, rows,
        action="export_admin_accounts", entity="app_accounts",
    )
