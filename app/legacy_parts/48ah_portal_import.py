# 48ah_portal_import.py
# استيراد حسابات البوابة من CSV أو Excel
# الصيغة: beneficiary_id_or_phone, username, password
# (الصف الأول = header يُتجاهَل)
# يدعم: .csv .xlsx .xls (والـ xls فقط إذا توفّر xlrd)

import csv
import hashlib
import io
import logging
import os
from flask import jsonify, request

_log = logging.getLogger("hobehub.portal_import")


def _import_sha256(s):
    return hashlib.sha256((s or "").encode("utf-8")).hexdigest()


def _import_normalize_phone(p):
    p = (p or "").strip()
    # ابقي الأرقام فقط
    return "".join(c for c in p if c.isdigit())


def _import_cell_to_text(v):
    """يحول قيمة خلية إكسل (int/float/str/None) إلى نص نظيف للاستيراد."""
    if v is None:
        return ""
    # أرقام: نخليها بدون .0 لو كانت صحيحة
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return repr(v).strip()
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def _import_parse_xlsx(file_bytes):
    """يقرأ .xlsx من bytes ويرجع قائمة صفوف (كل صف list of strings)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("مكتبة openpyxl غير مثبتة. شغّل: pip install openpyxl")
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([_import_cell_to_text(c) for c in row])
    try:
        wb.close()
    except Exception:
        pass
    return rows


def _import_parse_xls(file_bytes):
    """يقرأ .xls (الصيغة القديمة) من bytes — يتطلب xlrd."""
    try:
        import xlrd
    except ImportError:
        raise RuntimeError("صيغة .xls القديمة تتطلب مكتبة xlrd. الرجاء حفظ الملف كـ .xlsx.")
    book = xlrd.open_workbook(file_contents=file_bytes)
    sheet = book.sheet_by_index(0)
    rows = []
    for r in range(sheet.nrows):
        rows.append([_import_cell_to_text(sheet.cell_value(r, c)) for c in range(sheet.ncols)])
    return rows


def _import_parse_csv(file_bytes):
    """يقرأ CSV من bytes (UTF-8/BOM) ويرجع قائمة صفوف."""
    raw = file_bytes.decode("utf-8-sig", errors="ignore")
    reader = csv.reader(io.StringIO(raw))
    return [list(r) for r in reader]


def _import_parse_upload(upload):
    """يحدد صيغة الملف من الامتداد ويعيد قائمة صفوف (rows)."""
    fname = (getattr(upload, "filename", "") or "").lower()
    file_bytes = upload.read()
    if not file_bytes:
        return []
    ext = os.path.splitext(fname)[1]
    if ext == ".xlsx":
        return _import_parse_xlsx(file_bytes)
    if ext == ".xls":
        # حاول xlrd أولاً، ثم openpyxl لو فشل (بعض الملفات xlsx مغيّر امتدادها)
        try:
            return _import_parse_xls(file_bytes)
        except RuntimeError:
            try:
                return _import_parse_xlsx(file_bytes)
            except Exception:
                raise
    # افتراضي: CSV
    return _import_parse_csv(file_bytes)


@app.route("/admin/import/portal-accounts", methods=["POST"])
@admin_login_required
@permission_required("import")
def admin_import_portal_accounts():
    """يستورد حسابات بوابة من ملف CSV. الصيغة:
       beneficiary_identifier (id أو phone), username, password

       معاملات اختيارية:
         target_method = 'username' أو 'cards' أو فارغ
            → عند تحديده، يضبط طريقة الإنترنت للمستفيد تلقائيًا (جامعي/عمل حر).
    """
    upload = request.files.get("csv_file")
    if not upload or not getattr(upload, "filename", ""):
        return jsonify({"ok": False, "message": "يرجى اختيار ملف."}), 400

    # تحقق من الامتداد المدعوم
    fname_lower = (upload.filename or "").lower()
    allowed_exts = (".csv", ".xlsx", ".xls")
    if not fname_lower.endswith(allowed_exts):
        return jsonify({
            "ok": False,
            "message": f"صيغة غير مدعومة. المسموح: {' / '.join(allowed_exts)}"
        }), 400

    # طريقة الإنترنت المستهدفة لكل صف (تنطبق على beneficiaries.university_internet_method/freelancer_internet_method)
    target_method = (request.form.get("target_method") or "").strip().lower()
    method_value_map = {
        "username": "يوزر إنترنت",
        "cards":    "نظام البطاقات",
    }
    method_label = method_value_map.get(target_method)

    try:
        rows = _import_parse_upload(upload)
    except RuntimeError as e:
        return jsonify({"ok": False, "message": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "message": f"تعذّر قراءة الملف: {e}"}), 400

    if not rows or len(rows) < 2:
        return jsonify({"ok": False, "message": "الملف فارغ أو يحتوي على header فقط."}), 400

    # تخطّى الـ header
    data_rows = rows[1:]
    created = 0
    updated = 0
    skipped = 0
    errors = []

    for idx, r in enumerate(data_rows, start=2):  # idx مع 1 = header
        if not r or len(r) < 3:
            errors.append(f"الصف {idx}: ينقص حقول (يحتاج: المعرف، اليوزر، كلمة المرور)")
            skipped += 1
            continue
        ident = (r[0] or "").strip()
        username = (r[1] or "").strip()
        password = (r[2] or "").strip()
        if not ident or not username or not password:
            errors.append(f"الصف {idx}: قيمة فارغة")
            skipped += 1
            continue

        # حاول كـ id أولاً، ثم كـ phone
        ben = None
        if ident.isdigit() and len(ident) <= 7:
            ben = query_one("SELECT id, phone, full_name, user_type FROM beneficiaries WHERE id=%s", [int(ident)])
        if not ben:
            phone = _import_normalize_phone(ident)
            if phone:
                ben = query_one("SELECT id, phone, full_name, user_type FROM beneficiaries WHERE phone=%s", [phone])
        if not ben:
            errors.append(f"الصف {idx}: لا يوجد مستفيد بمعرّف/جوال '{ident}'")
            skipped += 1
            continue

        # طبّق طريقة الإنترنت لو target_method مُحدّد
        if method_label:
            ut = (ben.get("user_type") or "").strip().lower()
            try:
                if ut == "university":
                    execute_sql(
                        "UPDATE beneficiaries SET university_internet_method=%s WHERE id=%s",
                        [method_label, ben["id"]],
                    )
                elif ut == "freelancer":
                    execute_sql(
                        "UPDATE beneficiaries SET freelancer_internet_method=%s WHERE id=%s",
                        [method_label, ben["id"]],
                    )
            except Exception:
                pass

        # ─ إذا كان الاستيراد لحسابات إنترنت (RADIUS user)
        # نحفظ كلمة المرور في beneficiary_radius_accounts (للـ API) بدلاً من البوابة.
        # كلمة مرور البوابة منفصلة ولا يجب أن تتأثر بهذا الاستيراد.
        if target_method == "username":
            try:
                import hashlib as _hl
                pwd_md5 = _hl.md5(password.encode("utf-8")).hexdigest()
                existing_radius = query_one(
                    "SELECT id FROM beneficiary_radius_accounts WHERE beneficiary_id=%s LIMIT 1",
                    [ben["id"]],
                )
                if existing_radius:
                    execute_sql(
                        """
                        UPDATE beneficiary_radius_accounts SET
                            external_username=%s, plain_password=%s, password_md5=%s,
                            updated_at=CURRENT_TIMESTAMP
                        WHERE beneficiary_id=%s
                        """,
                        [username, password, pwd_md5, ben["id"]],
                    )
                    updated += 1
                else:
                    execute_sql(
                        """
                        INSERT INTO beneficiary_radius_accounts
                            (beneficiary_id, external_username, plain_password, password_md5, status)
                        VALUES (%s, %s, %s, %s, 'pending')
                        """,
                        [ben["id"], username, password, pwd_md5],
                    )
                    created += 1
            except Exception as e:
                errors.append(f"الصف {idx}: فشل حفظ بيانات RADIUS — {e}")
                skipped += 1
            continue  # لا تنشئ/تحدّث portal account لهذا المسار

        # ─ المسار العادي (حسابات البوابة أو بطاقات): نحفظ في portal accounts
        existing = query_one(
            "SELECT id FROM beneficiary_portal_accounts WHERE beneficiary_id=%s",
            [ben["id"]],
        )
        if existing:
            try:
                execute_sql(
                    "UPDATE beneficiary_portal_accounts SET username=%s, password_hash=%s, password_plain=%s, "
                    "must_set_password=0, updated_at=CURRENT_TIMESTAMP WHERE id=%s",
                    [username, _import_sha256(password), password, existing["id"]],
                )
                updated += 1
            except Exception as e:
                errors.append(f"الصف {idx}: فشل التحديث — {e}")
                skipped += 1
        else:
            dup = query_one(
                "SELECT id FROM beneficiary_portal_accounts WHERE username=%s",
                [username],
            )
            if dup:
                errors.append(f"الصف {idx}: اليوزر '{username}' مستخدم لمشترك آخر")
                skipped += 1
                continue
            try:
                execute_sql(
                    """
                    INSERT INTO beneficiary_portal_accounts
                        (beneficiary_id, username, password_hash, password_plain,
                         is_active, must_set_password, activated_at)
                    VALUES (%s, %s, %s, %s, 1, 0, CURRENT_TIMESTAMP)
                    """,
                    [ben["id"], username, _import_sha256(password), password],
                )
                created += 1
            except Exception as e:
                errors.append(f"الصف {idx}: فشل الإنشاء — {e}")
                skipped += 1

    log_action(
        "import_portal_accounts", "beneficiary_portal_account", 0,
        f"created={created} updated={updated} skipped={skipped}",
    )
    return jsonify({
        "ok": True,
        "message": f"تم: أُنشئ {created} • حُدِّث {updated} • تخطّى {skipped}.",
        "stats": {"created": created, "updated": updated, "skipped": skipped},
        "errors": errors[:20],
        "errors_count": len(errors),
    })
