from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

import psycopg2
from openpyxl import load_workbook
from psycopg2.extras import Json, RealDictCursor
from werkzeug.security import generate_password_hash

try:
    from dotenv import load_dotenv
except ImportError:  # pragma: no cover
    load_dotenv = None


ROOT = Path(__file__).resolve().parents[1]
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")
REQUIRED_HEADERS = {
    "username": {"اسم المستخدم", "إسم المستخدم", "user", "username"},
    "password": {"كلمة المرور", "كَلمة المرور", "password"},
    "full_name": {"الاسم", "الإسم", "الإسم الاول", "الاسم الاول", "name", "full_name"},
    "specialization": {"التخصص", "specialization"},
    "classification": {"التصنيف", "classification", "type"},
}


def clean(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return re.sub(r"\s+", " ", text).strip()


def normalize_phone(value) -> str:
    digits = "".join(ch for ch in clean(value) if ch.isdigit())
    if len(digits) == 9 and not digits.startswith("0"):
        digits = "0" + digits
    return digits


def normalize_search_ar(text: str) -> str:
    text = clean(text)
    repl = {"أ": "ا", "إ": "ا", "آ": "ا", "ى": "ي", "ة": "ه", "ؤ": "و", "ئ": "ي"}
    for old, new in repl.items():
        text = text.replace(old, new)
    return re.sub(r"\s+", " ", text).strip()


def split_name(full_name: str) -> tuple[str, str, str, str]:
    parts = clean(full_name).split()
    return (
        parts[0] if len(parts) > 0 else "",
        parts[1] if len(parts) > 1 else "",
        parts[2] if len(parts) > 2 else "",
        " ".join(parts[3:]) if len(parts) > 3 else "",
    )


def classify(raw: str) -> tuple[str, str]:
    value = clean(raw)
    if "عمل" in value or "حر" in value:
        return "freelancer", value
    if "طالب" in value or "طلاب" in value or "جامعة" in value or "جامع" in value:
        return "university", value
    return "", value


def load_env_files() -> None:
    if load_dotenv:
        load_dotenv(ROOT / ".env")
        load_dotenv(ROOT / ".env.render.local", override=False)


def mask_url(url: str) -> str:
    parsed = urlparse(url)
    return f"{parsed.scheme}://***:***@{parsed.hostname or 'unknown'}/{parsed.path.lstrip('/') or 'db'}"


def database_url() -> str:
    explicit = clean(os.getenv("DATABASE_URL"))
    load_env_files()
    url = explicit or clean(os.getenv("DATABASE_URL"))
    if not url:
        raise SystemExit("DATABASE_URL is required.")
    if url.startswith("postgres://"):
        url = "postgresql://" + url[len("postgres://") :]
    if not url.startswith("postgresql://"):
        raise SystemExit("DATABASE_URL must be a PostgreSQL URL.")
    if "sslmode=" not in url:
        url = url + ("&" if "?" in url else "?") + "sslmode=require"
    return url


def find_workbook(path_arg: str) -> Path:
    if path_arg:
        return Path(path_arg).expanduser().resolve()
    candidates = sorted(
        [p for p in (Path.home() / "Desktop").glob("*.xlsx") if not p.name.startswith("~$")],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise SystemExit("No .xlsx files found on Desktop.")
    return candidates[0]


def header_map(headers: list[str]) -> dict[str, int]:
    normalized = [clean(h).lower() for h in headers]
    mapping: dict[str, int] = {}
    for key, aliases in REQUIRED_HEADERS.items():
        aliases_norm = {a.lower() for a in aliases}
        for idx, header in enumerate(normalized):
            if header in aliases_norm:
                mapping[key] = idx
                break
    missing = [key for key in REQUIRED_HEADERS if key not in mapping]
    if missing:
        raise SystemExit(f"Missing required columns: {', '.join(missing)}")
    return mapping


def read_rows(workbook_path: Path) -> tuple[list[dict], dict]:
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb.active
    headers = [clean(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    mapping = header_map(headers)
    rows = []
    errors = []
    seen = Counter()
    categories = Counter()
    for excel_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(clean(v) for v in values):
            continue
        username = normalize_phone(values[mapping["username"]])
        password = clean(values[mapping["password"]])
        full_name = clean(values[mapping["full_name"]])
        specialization = clean(values[mapping["specialization"]])
        classification_raw = clean(values[mapping["classification"]])
        user_type, classification = classify(classification_raw)
        seen[username] += 1
        categories[classification_raw] += 1
        if not username or len(username) != 10 or not username.startswith("0"):
            errors.append({"row": excel_row, "reason": "invalid_username", "username": username})
            continue
        if not password:
            errors.append({"row": excel_row, "reason": "missing_password", "username": username})
            continue
        if not full_name:
            errors.append({"row": excel_row, "reason": "missing_name", "username": username})
            continue
        if user_type not in {"freelancer", "university"}:
            errors.append({"row": excel_row, "reason": "unknown_classification", "username": username, "classification": classification_raw})
            continue
        first, second, third, fourth = split_name(full_name)
        rows.append(
            {
                "excel_row": excel_row,
                "username": username,
                "password": password,
                "full_name": full_name,
                "first_name": first,
                "second_name": second,
                "third_name": third,
                "fourth_name": fourth,
                "specialization": specialization,
                "classification": classification,
                "user_type": user_type,
            }
        )
    duplicates = sorted(username for username, count in seen.items() if username and count > 1)
    if duplicates:
        errors.extend({"row": None, "reason": "duplicate_in_excel", "username": username} for username in duplicates)
    return rows, {"errors": errors, "categories": dict(categories), "sheet": ws.title}


def backup_existing(conn, usernames: list[str]) -> Path:
    out_dir = ROOT / "backups"
    out_dir.mkdir(exist_ok=True)
    out = out_dir / f"internet_users_import_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(
            """
            SELECT * FROM beneficiaries
            WHERE phone = ANY(%s)
            ORDER BY id
            """,
            [usernames],
        )
        beneficiaries = cur.fetchall()
        cur.execute(
            """
            SELECT pa.* FROM beneficiary_portal_accounts pa
            WHERE pa.username = ANY(%s)
            ORDER BY pa.id
            """,
            [usernames],
        )
        portal_accounts = cur.fetchall()
        beneficiary_ids = [row["id"] for row in beneficiaries]
        radius_accounts = []
        if beneficiary_ids:
            cur.execute(
                """
                SELECT * FROM beneficiary_radius_accounts
                WHERE beneficiary_id = ANY(%s)
                ORDER BY id
                """,
                [beneficiary_ids],
            )
            radius_accounts = cur.fetchall()
    data = {
        "created_at": datetime.now().isoformat(),
        "usernames_count": len(usernames),
        "beneficiaries": beneficiaries,
        "beneficiary_portal_accounts": portal_accounts,
        "beneficiary_radius_accounts": radius_accounts,
    }
    out.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return out


def upsert_rows(conn, rows: list[dict], *, dry_run: bool) -> dict:
    summary = Counter()
    skipped: list[dict] = []
    usernames = [row["username"] for row in rows]
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute("SELECT id, phone FROM beneficiaries WHERE phone = ANY(%s) ORDER BY id", [usernames])
        beneficiaries_by_phone: dict[str, list[int]] = {}
        for row in cur.fetchall():
            beneficiaries_by_phone.setdefault(row["phone"], []).append(row["id"])
        cur.execute(
            "SELECT id, username, beneficiary_id FROM beneficiary_portal_accounts WHERE username = ANY(%s) ORDER BY id",
            [usernames],
        )
        portals_by_username: dict[str, list[int]] = {}
        for row in cur.fetchall():
            portals_by_username.setdefault(row["username"], []).append(row["beneficiary_id"])

        for item in rows:
            username = item["username"]
            ben_ids = beneficiaries_by_phone.get(username, [])
            portal_ids = portals_by_username.get(username, [])
            ids = set(ben_ids) | set(portal_ids)
            if len(ids) > 1 or len(ben_ids) > 1 or len(portal_ids) > 1:
                skipped.append({"row": item["excel_row"], "username": username, "reason": "ambiguous_existing_records"})
                summary["skipped"] += 1
                continue

            beneficiary_id = next(iter(ids), None)
            is_new = beneficiary_id is None
            if dry_run:
                summary["would_insert" if is_new else "would_update"] += 1
                continue

            if is_new:
                cur.execute(
                    """
                    INSERT INTO beneficiaries (
                        user_type, first_name, second_name, third_name, fourth_name,
                        full_name, search_name, phone,
                        freelancer_specialization, freelancer_field, freelancer_internet_method,
                        university_name, university_specialization, university_internet_method,
                        weekly_usage_count, weekly_usage_week_start, notes,
                        added_by_username
                    ) VALUES (
                        %(user_type)s, %(first_name)s, %(second_name)s, %(third_name)s, %(fourth_name)s,
                        %(full_name)s, %(search_name)s, %(phone)s,
                        %(freelancer_specialization)s, %(freelancer_field)s, %(freelancer_internet_method)s,
                        %(university_name)s, %(university_specialization)s, %(university_internet_method)s,
                        0, CURRENT_DATE, %(notes)s,
                        'codex_import'
                    )
                    RETURNING id
                    """,
                    {
                        "user_type": item["user_type"],
                        "first_name": item["first_name"],
                        "second_name": item["second_name"],
                        "third_name": item["third_name"],
                        "fourth_name": item["fourth_name"],
                        "full_name": item["full_name"],
                        "search_name": normalize_search_ar(item["full_name"]),
                        "phone": username,
                        "freelancer_specialization": item["specialization"] if item["user_type"] == "freelancer" else "",
                        "freelancer_field": item["specialization"] if item["user_type"] == "freelancer" else "",
                        "freelancer_internet_method": "يوزر إنترنت" if item["user_type"] == "freelancer" else "",
                        "university_name": "",
                        "university_specialization": item["specialization"] if item["user_type"] == "university" else "",
                        "university_internet_method": "يوزر إنترنت" if item["user_type"] == "university" else "",
                        "notes": f"استيراد مستخدمي الإنترنت من Excel. التصنيف: {item['classification']}. التخصص: {item['specialization']}",
                    },
                )
                beneficiary_id = cur.fetchone()["id"]
                summary["inserted"] += 1
            else:
                cur.execute(
                    """
                    UPDATE beneficiaries
                    SET user_type=%(user_type)s,
                        first_name=%(first_name)s,
                        second_name=%(second_name)s,
                        third_name=%(third_name)s,
                        fourth_name=%(fourth_name)s,
                        full_name=%(full_name)s,
                        search_name=%(search_name)s,
                        freelancer_specialization=CASE WHEN %(user_type)s='freelancer' THEN %(specialization)s ELSE freelancer_specialization END,
                        freelancer_field=CASE WHEN %(user_type)s='freelancer' THEN %(specialization)s ELSE freelancer_field END,
                        freelancer_internet_method=CASE WHEN %(user_type)s='freelancer' THEN 'يوزر إنترنت' ELSE freelancer_internet_method END,
                        university_specialization=CASE WHEN %(user_type)s='university' THEN %(specialization)s ELSE university_specialization END,
                        university_internet_method=CASE WHEN %(user_type)s='university' THEN 'يوزر إنترنت' ELSE university_internet_method END,
                        notes=COALESCE(NULLIF(notes, ''), '') || CASE WHEN COALESCE(notes, '') = '' THEN '' ELSE E'\n' END || %(notes)s
                    WHERE id=%(beneficiary_id)s
                    """,
                    {
                        **item,
                        "search_name": normalize_search_ar(item["full_name"]),
                        "beneficiary_id": beneficiary_id,
                        "notes": f"تحديث من استيراد مستخدمي الإنترنت. التصنيف: {item['classification']}. التخصص: {item['specialization']}",
                    },
                )
                summary["updated"] += 1

            password_hash = generate_password_hash(item["password"])
            cur.execute(
                """
                INSERT INTO beneficiary_portal_accounts (
                    beneficiary_id, username, password_hash, password_plain,
                    is_active, must_set_password, activated_at
                ) VALUES (%s,%s,%s,%s,TRUE,FALSE,CURRENT_TIMESTAMP)
                ON CONFLICT (username) DO UPDATE SET
                    beneficiary_id=EXCLUDED.beneficiary_id,
                    password_hash=EXCLUDED.password_hash,
                    password_plain=EXCLUDED.password_plain,
                    is_active=TRUE,
                    must_set_password=FALSE,
                    activated_at=COALESCE(beneficiary_portal_accounts.activated_at, CURRENT_TIMESTAMP),
                    updated_at=CURRENT_TIMESTAMP
                """,
                [beneficiary_id, username, password_hash, item["password"]],
            )
            cur.execute(
                """
                INSERT INTO beneficiary_radius_accounts (
                    beneficiary_id, external_username, plain_password, status, sync_status, notes
                ) VALUES (%s,%s,%s,'pending','pending',%s)
                ON CONFLICT (beneficiary_id) DO UPDATE SET
                    external_username=EXCLUDED.external_username,
                    plain_password=EXCLUDED.plain_password,
                    status='pending',
                    sync_status='pending',
                    notes=EXCLUDED.notes,
                    updated_at=CURRENT_TIMESTAMP
                """,
                [beneficiary_id, username, item["password"], "استيراد مستخدمي الإنترنت من Excel"],
            )
            cur.execute(
                """
                INSERT INTO radius_pending_actions (
                    action_type, target_kind, beneficiary_id, payload_json,
                    requested_by_username, notes, attempted_by_mode
                ) VALUES (%s,%s,%s,%s,%s,%s,'manual')
                """,
                [
                    "create_user",
                    "user",
                    beneficiary_id,
                    Json({"username": username, "password": item["password"], "profile_id": "", "source": "excel_import"}),
                    "codex_import",
                    "استيراد حساب إنترنت من ملف Excel",
                ],
            )
            summary["portal_accounts_upserted"] += 1
            summary["radius_accounts_upserted"] += 1
            summary["pending_actions_created"] += 1
    summary["skipped"] += len(skipped)
    return {"summary": dict(summary), "skipped": skipped}


def main() -> int:
    parser = argparse.ArgumentParser(description="Import internet username subscribers from an Excel workbook.")
    parser.add_argument("--file", default="", help="Path to .xlsx workbook. Defaults to newest .xlsx on Desktop.")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--skip-invalid", action="store_true", help="Import valid rows and report invalid rows.")
    args = parser.parse_args()

    workbook_path = find_workbook(args.file)
    rows, meta = read_rows(workbook_path)
    if meta["errors"] and not args.skip_invalid:
        print(json.dumps({"ok": False, "file": str(workbook_path), "rows_valid": len(rows), **meta}, ensure_ascii=False, indent=2))
        return 1

    url = database_url()
    conn = psycopg2.connect(url, connect_timeout=15)
    try:
        with conn:
            backup_path = None if args.dry_run else backup_existing(conn, [row["username"] for row in rows])
            result = upsert_rows(conn, rows, dry_run=args.dry_run)
        print(
            json.dumps(
                {
                    "ok": True,
                    "dry_run": args.dry_run,
                    "database": mask_url(url),
                    "file": str(workbook_path),
                    "sheet": meta["sheet"],
                    "rows_valid": len(rows),
                    "errors": meta["errors"],
                    "categories": meta["categories"],
                    "backup": str(backup_path) if backup_path else None,
                    **result,
                },
                ensure_ascii=False,
                indent=2,
            )
        )
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
