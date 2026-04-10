
"""
Hobe Hub Professional+ Edition
- Sidebar UI
- Dashboard stats
- Pagination
- Audit log
- DB-based auth + permissions
- User profile / password change
- Beneficiaries management
- Smart CSV import / export
- Backup SQL (table-level)
"""

from __future__ import annotations

import csv
import hashlib
import io
import math
import os
import re
import threading
from datetime import date, timedelta, datetime
from functools import wraps
from uuid import uuid4

from flask import (
    Flask,
    Response,
    flash,
    jsonify,
    redirect,
    render_template_string,
    request,
    session,
    url_for,
)
import psycopg2
from psycopg2 import pool
from psycopg2.extras import RealDictCursor, execute_batch, execute_values
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "change-this-secret-key")

DEFAULT_DATABASE_URL = "postgresql://user_admin:mVDLjTHICnhOoR1eiYZCnObgmVsDSeuy@dpg-d7bm947kijhs73aq6i00-a.oregon-postgres.render.com/user_management_g50l?sslmode=require"

DB_POOL_MINCONN = int(os.getenv("DB_POOL_MINCONN", "1"))
DB_POOL_MAXCONN = int(os.getenv("DB_POOL_MAXCONN", "12"))
_connection_pool = None
IMPORT_TASKS = {}
IMPORT_TASKS_LOCK = threading.Lock()
IMPORT_LOG_LIMIT = 300
IMPORT_BATCH_SIZE = int(os.getenv("IMPORT_BATCH_SIZE", "250"))


def get_database_url() -> str:
    database_url = os.getenv("DATABASE_URL", DEFAULT_DATABASE_URL).strip()
    if database_url.startswith("postgresql://") and "sslmode=" not in database_url:
        separator = "&" if "?" in database_url else "?"
        database_url = f"{database_url}{separator}sslmode=require"
    return database_url

PERMISSIONS = [
    "view",
    "add",
    "edit",
    "delete",
    "import",
    "export",
    "backup",
    "usage_counter",
    "reset_weekly_usage",
    "manage_accounts",
    "view_audit_log",
]

PERMISSION_LABELS = {
    "view": "عرض البيانات",
    "add": "إضافة مستفيد",
    "edit": "تعديل المستفيدين",
    "delete": "حذف المستفيدين",
    "import": "استيراد CSV",
    "export": "تصدير البيانات",
    "backup": "نسخة احتياطية",
    "usage_counter": "زر +1 للبطاقات",
    "reset_weekly_usage": "تجديد كل البطاقات",
    "manage_accounts": "إدارة المستخدمين والصلاحيات",
    "view_audit_log": "عرض سجل العمليات",
}

TAWJIHI_YEARS = ["2006", "2007", "2008", "2009", "2010", "2011"]
TAWJIHI_BRANCHES = ["علمي", "أدبي", "شرعي", "زراعي", "صناعي", "تجاري", "فندقي"]
FREELANCER_SCHEDULE_OPTIONS = ["دوام كامل", "أحد ثلاثاء خميس", "سبت اثنين أربعاء", "بطاقة حسب الحاجة"]
TIME_MODE_OPTIONS = ["عشوائي", "وقت محدد"]
UNIVERSITIES_GAZA = ["الجامعة الإسلامية", "جامعة الأزهر", "جامعة الأقصى", "جامعة فلسطين", "جامعة القدس المفتوحة", "الكلية الجامعية للعلوم التطبيقية", "كلية العلوم والتكنولوجيا", "أخرى"]
UNIVERSITY_DAYS_OPTIONS = ["عشوائي", "سبت اثنين أربعاء", "أحد ثلاثاء خميس"]
INTERNET_ACCESS_METHOD_OPTIONS = ["يمتلك اسم مستخدم", "نظام البطاقات"]

CSV_IMPORT_COLUMNS = [
    "user_type",
    "first_name",
    "second_name",
    "third_name",
    "fourth_name",
    "phone",
    "notes",
    "tawjihi_year",
    "tawjihi_branch",
    "freelancer_specialization",
    "freelancer_company",
    "freelancer_schedule_type",
    "freelancer_internet_method",
    "freelancer_time_mode",
    "freelancer_time_from",
    "freelancer_time_to",
    "university_name",
    "university_college",
    "university_specialization",
    "university_days",
    "university_internet_method",
    "university_time_mode",
    "university_time_from",
    "university_time_to",
]

BASE_TEMPLATE = """
<!DOCTYPE html>
<html lang="ar" dir="rtl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{{ title }}</title>
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
<style>
:root{
  --bg:#f4f8fc; --card:#ffffff; --line:#dce8f3; --text:#142033;
  --muted:#64748b; --primary:#123b6d; --secondary:#35a7e8; --accent:#f7c948;
  --danger:#c0392b; --success:#1f8f51; --soft:#eef5fc; --orange:#f28c28; --purple:#6d4ee8;
}
*{box-sizing:border-box}
html{scroll-behavior:smooth}
body{margin:0;font-family:Arial,sans-serif;background:linear-gradient(135deg,#f6fafd,#eef7fc);color:var(--text)}
a{text-decoration:none}
.layout{display:flex;min-height:100vh;transition:.25s}
.sidebar{
  width:290px;background:linear-gradient(180deg,var(--primary),#0d2f57);color:#fff;padding:20px;position:sticky;top:0;height:100vh;
  overflow-x:hidden;overflow-y:auto;box-shadow:0 8px 24px rgba(0,0,0,.12);transition:width .28s ease,padding .28s ease;white-space:nowrap
}
.layout.sidebar-collapsed .sidebar{width:88px;padding:20px 12px;box-shadow:0 8px 24px rgba(0,0,0,.12)}
.layout.sidebar-collapsed .main{width:100%}
.brand{display:flex;align-items:center;gap:12px;font-size:24px;font-weight:bold;margin-bottom:18px;overflow:hidden}
.brand-badge{width:42px;height:42px;min-width:42px;border-radius:12px;background:var(--accent);color:var(--primary);display:flex;align-items:center;justify-content:center;font-weight:bold}
.brand-text{display:flex;flex-direction:column;transition:opacity .2s ease,transform .2s ease,width .2s ease;overflow:hidden}
.brand small{display:block;font-size:12px;color:#cfe4ff;margin-top:4px}
.layout.sidebar-collapsed .brand{justify-content:center}
.layout.sidebar-collapsed .brand-text{opacity:0;transform:translateX(8px);width:0}
.nav .section{font-size:12px;color:#d8e8ff;margin:16px 0 8px;transition:opacity .2s ease}
.nav a,.nav details summary{display:flex;align-items:center;gap:10px;color:#fff;padding:12px 14px;border-radius:12px;margin-bottom:8px;background:rgba(255,255,255,.08);cursor:pointer;list-style:none;transition:all .22s ease;overflow:hidden}
.nav a:hover,.nav details summary:hover,.nav a.active{background:rgba(255,255,255,.18)}
.nav a i,.nav details summary i{min-width:20px;text-align:center;font-size:16px}
.nav-label{transition:opacity .2s ease,transform .2s ease,width .2s ease;overflow:hidden}
.nav details{margin-bottom:8px}
.nav details summary::-webkit-details-marker{display:none}
.nav details .submenu{padding-right:10px}
.nav details .submenu a{background:rgba(255,255,255,.05);font-size:14px}
.layout.sidebar-collapsed .nav-label{opacity:0;transform:translateX(8px);width:0}
.layout.sidebar-collapsed .nav a,.layout.sidebar-collapsed .nav details summary{justify-content:center;padding:12px 10px}
.layout.sidebar-collapsed .nav .section,.layout.sidebar-collapsed .nav details .submenu{display:none}
.layout.sidebar-collapsed .nav details{margin-bottom:8px}
.layout.sidebar-collapsed .nav details[open] summary{margin-bottom:8px}
.main{flex:1;padding:22px}
.topbar{display:flex;justify-content:space-between;align-items:center;gap:12px;margin-bottom:20px;flex-wrap:wrap}
.topbar-left{display:flex;align-items:center;gap:10px;flex-wrap:wrap}
.sidebar-toggle{width:42px;height:42px;border-radius:12px;border:1px solid var(--line);background:#fff;color:var(--primary);display:inline-flex;align-items:center;justify-content:center;cursor:pointer;font-size:18px}
.topbar .userbox{display:flex;gap:8px;align-items:center;flex-wrap:wrap}
.badge{display:inline-block;padding:6px 10px;border-radius:999px;background:var(--soft);color:var(--primary);font-size:12px;font-weight:bold}
.badge-green{background:#ecfdf3;color:#166534}
.badge-orange{background:#fff7ed;color:#c2410c}
.badge-purple{background:#f5f3ff;color:#6d28d9}
.card{background:var(--card);border:1px solid var(--line);border-radius:20px;padding:20px;box-shadow:0 8px 24px rgba(15,23,42,.05)}
.hero{background:linear-gradient(135deg,var(--primary),var(--secondary));color:#fff;border-radius:24px;padding:28px;margin-bottom:18px;box-shadow:0 10px 28px rgba(18,59,109,.16)}
.hero h1{margin:0 0 8px}
.hero p{margin:0;color:#eaf5ff;line-height:1.8}
.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:14px}
.grid-2{display:grid;grid-template-columns:repeat(auto-fit,minmax(320px,1fr));gap:16px}
.grid-3{display:grid;grid-template-columns:repeat(auto-fit,minmax(260px,1fr));gap:16px}
.stat{padding:18px;border-radius:18px;background:#fff;border:1px solid var(--line)}
.stat h3{margin:0;color:var(--muted);font-size:14px}
.stat .num{font-size:30px;font-weight:bold;margin-top:8px;color:var(--primary)}
.mini-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:16px}
.menu-card{display:block;color:var(--text);background:#fff;border:1px solid var(--line);border-radius:18px;padding:22px;box-shadow:0 6px 18px rgba(0,0,0,.05);transition:.2s}
.menu-card:hover{transform:translateY(-3px);box-shadow:0 10px 24px rgba(0,0,0,.08);border-color:var(--secondary)}
.menu-icon{width:54px;height:54px;border-radius:16px;background:linear-gradient(135deg,var(--accent),var(--orange));color:var(--primary);display:flex;align-items:center;justify-content:center;font-size:24px;margin-bottom:14px}
.menu-card h3{margin:0 0 8px 0;color:var(--primary);font-size:18px}
.menu-card p{margin:0;color:#666;font-size:14px;line-height:1.7}
.table-wrap{overflow:auto;border:1px solid var(--line);border-radius:16px;background:#fff}
table{width:100%;border-collapse:collapse;background:#fff;min-width:900px}
th,td{padding:11px;border-bottom:1px solid var(--line);text-align:center;vertical-align:middle;white-space:nowrap}
th{background:#eef4fb}
th a{color:inherit}
.cell-wrap{white-space:normal;max-width:260px;line-height:1.7}
input,select,textarea{width:100%;padding:11px;border:1px solid var(--line);border-radius:12px;background:#fff}
textarea{min-height:100px}
.row{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:12px}
.actions{display:flex;gap:8px;flex-wrap:wrap}
.btn{display:inline-flex;align-items:center;gap:8px;text-decoration:none;border:none;cursor:pointer;padding:11px 16px;border-radius:12px;font-size:14px}
.btn-icon{width:34px;height:34px;padding:0;border-radius:10px;justify-content:center;font-size:13px}
.btn-icon i{margin:0}
.btn-primary{background:var(--primary);color:#fff}
.btn-secondary{background:var(--secondary);color:#fff}
.btn-accent{background:var(--accent);color:#213547}
.btn-danger{background:var(--danger);color:#fff}
.btn-outline{background:#fff7ef;border:1px solid #ffd7af;color:#d97706}
.btn-soft{background:#f8fbff;border:1px solid var(--line);color:var(--primary)}
.flash{padding:12px 15px;border-radius:12px;margin-bottom:10px}
.flash.success{background:#ecfdf3;color:#14532d;border:1px solid #bbf7d0}
.flash.error{background:#fef2f2;color:#991b1b;border:1px solid #fecaca}
.small{font-size:12px;color:var(--muted)}
.pagination{display:flex;gap:8px;flex-wrap:wrap;justify-content:center;margin-top:14px}
.pagination a,.pagination span{padding:8px 12px;border-radius:10px;border:1px solid var(--line);background:#fff;text-decoration:none;color:var(--text)}
.pagination .active{background:var(--primary);color:#fff}
.login-wrap{max-width:480px;margin:70px auto}
.inline-form{display:inline}
.permissions-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:8px}
.info-note{padding:14px;border:1px solid var(--line);background:#f8fbff;border-radius:14px}
.tabs{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:14px}
.tab-pill{padding:10px 14px;border-radius:999px;border:1px solid var(--line);background:#fff;color:var(--text);font-weight:bold}
.tab-pill.active{background:var(--primary);color:#fff;border-color:var(--primary)}
.filter-box{background:#f9fcff;border:1px solid var(--line);border-radius:16px;padding:14px}
.advanced-filters{margin-top:12px;border-top:1px dashed var(--line);padding-top:12px}
.advanced-filters summary{cursor:pointer;font-weight:bold;color:var(--primary);margin-bottom:12px}
.kpi-strip{display:grid;grid-template-columns:repeat(auto-fit,minmax(170px,1fr));gap:10px;margin-top:14px}
.kpi{background:#fff;border:1px solid var(--line);border-radius:16px;padding:14px}
.kpi .label{font-size:12px;color:var(--muted)}
.kpi .value{font-size:24px;color:var(--primary);font-weight:bold;margin-top:6px}
.chart-card{background:#fff;border:1px solid var(--line);border-radius:18px;padding:18px}
.chart-card h3{margin:0 0 14px 0}
.bar-list{display:flex;flex-direction:column;gap:12px}
.bar-row{display:grid;grid-template-columns:130px 1fr 56px;gap:10px;align-items:center}
.bar-label{font-size:13px;color:var(--text);overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.bar-track{background:#eef4fb;border-radius:999px;height:12px;overflow:hidden}
.bar-fill{height:100%;background:linear-gradient(90deg,var(--secondary),var(--purple));border-radius:999px}
.metric-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:10px}
.metric-box{border:1px solid var(--line);border-radius:16px;background:#fff;padding:14px}
.metric-box h4{margin:0;color:var(--muted);font-size:13px}
.metric-box .num{margin-top:8px;color:var(--primary);font-size:24px;font-weight:bold}
.modal{position:fixed;inset:0;background:rgba(15,23,42,.6);display:none;align-items:center;justify-content:center;padding:18px;z-index:999}
.modal:target{display:flex}
.modal-card{background:#fff;width:min(1100px,100%);max-height:90vh;overflow:auto;border-radius:20px;padding:20px;position:relative}
.modal-close{position:absolute;left:16px;top:16px;width:38px;height:38px;border-radius:50%;display:flex;align-items:center;justify-content:center;background:#f3f6fb;color:var(--text);font-size:20px}
.form-section{display:none}
.form-section.active{display:block}
.section-title{margin:18px 0 10px 0;color:var(--primary);font-size:18px}
.empty-state{padding:28px;text-align:center;color:var(--muted)}
.type-badge{display:inline-flex;align-items:center;justify-content:center;padding:6px 12px;border-radius:999px;font-size:12px;font-weight:bold;border:1px solid transparent}
.type-badge.type-green{background:#ecfdf3;color:#166534;border-color:#bbf7d0}
.type-badge.type-blue{background:#eff6ff;color:#1d4ed8;border-color:#bfdbfe}
.type-badge.type-purple{background:#f5f3ff;color:#7c3aed;border-color:#ddd6fe}
.type-badge.type-default{background:#f8fafc;color:#334155;border-color:#cbd5e1}
.type-badge{display:inline-flex;align-items:center;justify-content:center;padding:6px 12px;border-radius:999px;font-size:12px;font-weight:700}
.type-tawjihi{background:#ecfdf3;color:#166534}
.type-university{background:#eef2ff;color:#4338ca}
.type-freelancer{background:#eff6ff;color:#1d4ed8}
.row-type-tawjihi td{background:#f7fff9}
.row-type-university td{background:#f8faff}
.row-type-freelancer td{background:#f8fcff}
.row-complete td{background:#fff1f2 !important}
.notes-box{min-height:90px}
.ajax-saving{opacity:.65;pointer-events:none}
.bulk-toolbar{display:flex;gap:10px;align-items:center;justify-content:flex-end;flex-wrap:wrap;margin-top:12px}.checkbox-cell{width:44px}.row-select,.select-all{width:18px;height:18px;cursor:pointer}.selected-count{font-size:13px;color:var(--muted)}
.timer-alert{display:none;position:sticky;top:0;z-index:1200;margin-bottom:12px;padding:14px 16px;border-radius:16px;background:linear-gradient(135deg,#fff7ed,#ffedd5);border:1px solid #fdba74;color:#9a3412;box-shadow:0 8px 24px rgba(0,0,0,.08)}
.timer-alert.show{display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap;animation:timerAlertIn .35s ease}
.timer-mini{display:inline-flex;align-items:center;gap:8px;padding:8px 12px;border-radius:999px;background:#eef4fb;color:var(--primary);font-weight:700;box-shadow:inset 0 0 0 1px rgba(18,59,109,.08)}
.timer-status-running{background:#ecfdf3;color:#166534}.timer-status-paused{background:#fff7ed;color:#c2410c}.timer-status-stopped{background:#f8fafc;color:#334155}.timer-status-alarm{background:#fef2f2;color:#b91c1c}
.timer-big{font-size:48px;font-weight:800;letter-spacing:2px;color:var(--primary);text-align:center;margin:10px 0 18px}
.timer-state-badge{display:inline-flex;align-items:center;gap:8px;padding:8px 14px;border-radius:999px;font-weight:700}
.timer-page-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:14px}
.topbar-clock{display:inline-flex;align-items:center;gap:8px;padding:10px 14px;border-radius:14px;background:#fff;border:1px solid var(--line);color:var(--primary);font-weight:800;box-shadow:0 8px 20px rgba(15,23,42,.05)}
.topbar-clock small{display:block;color:var(--muted);font-weight:600;font-size:11px}
.userbox .logout-btn{display:inline-flex;align-items:center;gap:8px;padding:10px 14px;border-radius:14px;background:linear-gradient(135deg,#0f2f56,#1c4f8f);color:#fff;border:none;box-shadow:0 10px 24px rgba(18,59,109,.22)}
.timer-visual-card{position:relative;overflow:hidden;background:radial-gradient(circle at top right,#eff6ff,transparent 35%),linear-gradient(180deg,#ffffff,#f8fbff);border:1px solid var(--line)}
.timer-visual-card:before{content:'';position:absolute;inset:-80px auto auto -80px;width:220px;height:220px;border-radius:50%;background:rgba(53,167,232,.10);filter:blur(10px)}
.timer-ring-wrap{display:flex;align-items:center;justify-content:center;padding:8px 0 2px}
.timer-ring{position:relative;width:280px;height:280px;display:grid;place-items:center}
.timer-ring svg{width:100%;height:100%;transform:rotate(-90deg)}
.timer-ring .bg{fill:none;stroke:#e6eef8;stroke-width:13}
.timer-ring .progress{fill:none;stroke:url(#timerRingGradient);stroke-width:13;stroke-linecap:round;stroke-dasharray:816;stroke-dashoffset:816;filter:drop-shadow(0 8px 18px rgba(53,167,232,.28));transition:stroke-dashoffset .8s linear}
.timer-ring-center{position:absolute;inset:32px;border-radius:50%;background:linear-gradient(180deg,#ffffff,#f7fbff);box-shadow:inset 0 0 0 1px rgba(18,59,109,.06),0 10px 30px rgba(15,23,42,.06);display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center;padding:18px}
.timer-ring-label{font-size:13px;color:var(--muted);margin-bottom:6px}
.timer-big.timer-big-ring{margin:0;font-size:56px;line-height:1}
.timer-ring-phase{margin-top:10px;font-size:13px;color:var(--muted)}
.timer-progress-strip{margin-top:14px;background:#edf4fb;border-radius:999px;height:14px;overflow:hidden;box-shadow:inset 0 1px 2px rgba(15,23,42,.06)}
.timer-progress-fill{height:100%;width:0%;background:linear-gradient(90deg,var(--secondary),var(--purple));border-radius:999px;transition:width .8s linear}
.timer-meta{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:12px;margin-top:16px}
.timer-meta-card{padding:14px;border-radius:18px;background:#fff;border:1px solid var(--line);box-shadow:0 8px 20px rgba(15,23,42,.04)}
.timer-meta-card h4{margin:0;color:var(--muted);font-size:13px}
.timer-meta-card .v{margin-top:8px;color:var(--primary);font-size:28px;font-weight:800}
.dashboard-timer-card{position:relative;overflow:hidden;background:linear-gradient(135deg,#f8fbff,#ffffff)}
.dashboard-timer-card:after{content:'';position:absolute;inset:auto -30px -40px auto;width:180px;height:180px;border-radius:50%;background:rgba(109,78,232,.07)}
.timer-pulse{position:relative}
.timer-pulse:after{content:'';position:absolute;inset:-6px;border-radius:inherit;border:2px solid currentColor;opacity:0;animation:pulseRing 1.8s infinite}
@keyframes pulseRing{0%{transform:scale(.92);opacity:.35}70%{transform:scale(1.08);opacity:0}100%{opacity:0}}
@keyframes timerAlertIn{from{opacity:0;transform:translateY(-8px)}to{opacity:1;transform:translateY(0)}}
@media (max-width:900px){.layout{display:block}.sidebar{width:100%;height:auto;position:relative}.layout.sidebar-collapsed .sidebar{width:72px;height:auto;padding:16px 10px}.main{padding:16px}.modal-card{padding:16px}.bar-row{grid-template-columns:100px 1fr 44px}.timer-big{font-size:36px}.timer-ring{width:230px;height:230px}.timer-big.timer-big-ring{font-size:44px}.topbar{align-items:flex-start}.userbox{width:100%;justify-content:flex-start}}
</style>
<script>
function toggleBeneficiarySections(selectEl, scopeId){
  var container = document.getElementById(scopeId);
  if(!container) return;
  var value = (selectEl && selectEl.value) || '';
  var sections = container.querySelectorAll('.form-section');
  sections.forEach(function(sec){sec.classList.remove('active');});
  var target = container.querySelector('.section-' + value);
  if(target){target.classList.add('active');}
}
function syncTypeTabs(scopeId, value){
  var container = document.getElementById(scopeId);
  if(!container) return;
  container.querySelectorAll('.type-tab').forEach(function(btn){ btn.classList.toggle('active', btn.dataset.value === value); });
}
function setBeneficiaryType(scopeId, value){
  var container = document.getElementById(scopeId);
  if(!container) return false;
  var input = container.querySelector('select[name="user_type"], input[name="user_type"]');
  if(input){ input.value = value; toggleBeneficiarySections(input, scopeId); }
  syncTypeTabs(scopeId, value);
  return false;
}
function initBeneficiaryForms(){
  document.querySelectorAll('[data-beneficiary-scope]').forEach(function(scope){
    var input = scope.querySelector('select[name="user_type"], input[name="user_type"]');
    if(input){toggleBeneficiarySections(input, scope.id); syncTypeTabs(scope.id, input.value || 'tawjihi');}
  });
}
function applySidebarState(collapsed){
  var layout = document.getElementById('app-layout');
  if(!layout) return;
  layout.classList.toggle('sidebar-collapsed', !!collapsed);
  try{ localStorage.setItem('sidebar-collapsed', collapsed ? '1' : '0'); }catch(e){}
}
function toggleSidebar(){
  var layout = document.getElementById('app-layout');
  if(!layout) return false;
  applySidebarState(!layout.classList.contains('sidebar-collapsed'));
  return false;
}
function showLiveFlash(message, category){
  var area = document.getElementById('live-flash-area');
  if(!area) return;
  var box = document.createElement('div');
  box.className = 'flash ' + (category || 'success');
  box.textContent = message;
  area.prepend(box);
  setTimeout(function(){ box.remove(); }, 4000);
}
function guardSingleSubmit(form){
  if(form.dataset.submitting === '1') return false;
  form.dataset.submitting = '1';
  form.classList.add('ajax-saving');
  var btn = form.querySelector('button[type="submit"]');
  if(btn){ btn.disabled = true; }
  return true;
}
async function ajaxPost(url, body){
  const options = {method:'POST', headers:{'X-Requested-With':'XMLHttpRequest'}};
  if(body instanceof FormData){ options.body = body; }
  const response = await fetch(url, options);
  const data = await response.json();
  if(!response.ok || !data.ok){ throw new Error((data && data.message) || 'حدث خطأ'); }
  return data;
}
function replaceRowAndModal(data, rowId, modalId){
  if(data.remove_row){
    const row = document.getElementById('beneficiary-row-' + rowId);
    if(row) row.remove();
  } else if(data.row_html){
    const row = document.getElementById('beneficiary-row-' + rowId);
    if(row) row.outerHTML = data.row_html;
  }
  if(modalId && data.modal_html){
    const oldModal = document.getElementById(modalId);
    if(oldModal){ oldModal.outerHTML = data.modal_html; }
  }
  if(data.message){ showLiveFlash(data.message, data.category || 'success'); }
}
async function incrementUsageAjax(url, rowId, modalId){
  try{
    const data = await ajaxPost(url);
    replaceRowAndModal(data, rowId, modalId);
  }catch(err){ showLiveFlash(err.message || 'تعذر إضافة البطاقة', 'error'); }
  return false;
}
async function submitBeneficiaryEdit(form, rowId, modalId){
  form.classList.add('ajax-saving');
  try{
    const data = await ajaxPost(form.action, new FormData(form));
    replaceRowAndModal(data, rowId, modalId);
    window.location.hash = '#!';
  }catch(err){
    showLiveFlash(err.message || 'تعذر حفظ التعديل', 'error');
  }finally{
    form.classList.remove('ajax-saving');
  }
  return false;
}

function updateBulkSelectedCount(){
  const checks = Array.from(document.querySelectorAll('.row-select'));
  const count = checks.filter(cb => cb.checked).length;
  const box = document.getElementById('selected-count');
  if(box){ box.textContent = String(count); }
  const master = document.getElementById('select-all');
  if(master){
    master.checked = checks.length > 0 && count === checks.length;
    master.indeterminate = count > 0 && count < checks.length;
  }
}
function toggleSelectAll(master){
  document.querySelectorAll('.row-select').forEach(function(cb){ cb.checked = !!master.checked; });
  updateBulkSelectedCount();
  return true;
}
function getSelectedBeneficiaryIds(){
  return Array.from(document.querySelectorAll('.row-select:checked')).map(cb => cb.value);
}
function submitBulkDelete(){
  const ids = getSelectedBeneficiaryIds();
  if(!ids.length){ showLiveFlash('حدد مستفيدًا واحدًا على الأقل.', 'error'); return false; }
  if(!confirm('هل تريد حذف المستفيدين المحددين؟')) return false;
  const form = document.getElementById('bulk-delete-form');
  form.querySelector('input[name="ids"]').value = ids.join(',');
  form.submit();
  return false;
}
function submitBulkExport(){
  const ids = getSelectedBeneficiaryIds();
  if(!ids.length){ showLiveFlash('حدد مستفيدًا واحدًا على الأقل.', 'error'); return false; }
  const form = document.getElementById('bulk-export-form');
  form.querySelector('input[name="ids"]').value = ids.join(',');
  form.submit();
  return false;
}

let powerTimerPollHandle = null;
let powerTimerBeepHandle = null;
let lastPowerTimerAlertKey = null;
function formatTimerSeconds(total){
  total = Math.max(0, parseInt(total || 0, 10));
  const h = Math.floor(total / 3600);
  const m = Math.floor((total % 3600) / 60);
  const s = total % 60;
  if(h > 0){ return String(h).padStart(2,'0') + ':' + String(m).padStart(2,'0') + ':' + String(s).padStart(2,'0'); }
  return String(m).padStart(2,'0') + ':' + String(s).padStart(2,'0');
}
function timerStateLabel(state){
  return ({running:'يعمل الآن', paused:'متوقف مؤقتًا', stopped:'متوقف', alarm:'انتهى الوقت'}[state] || 'غير معروف');
}
function timerStateClass(state){
  return 'timer-status-' + (state || 'stopped');
}
function stopPowerTimerSound(){
  if(powerTimerBeepHandle){ clearInterval(powerTimerBeepHandle); powerTimerBeepHandle = null; }
  const audio = document.getElementById('power-timer-audio');
  if(audio){ try{ audio.pause(); audio.currentTime = 0; }catch(e){} }
}
function startPowerTimerSound(){
  const audio = document.getElementById('power-timer-audio');
  if(!audio) return;
  if(powerTimerBeepHandle) return;
  const playNow = function(){ audio.play().catch(function(){}); };
  playNow();
  powerTimerBeepHandle = setInterval(playNow, 2500);
}
function showPowerTimerAlert(message, sticky){
  const box = document.getElementById('power-timer-alert');
  const text = document.getElementById('power-timer-alert-text');
  if(!box || !text) return;
  text.textContent = message;
  box.classList.add('show');
  if(!sticky){
    clearTimeout(box._hideTimeout);
    box._hideTimeout = setTimeout(function(){ box.classList.remove('show'); }, 7000);
  }
}
function dismissPowerTimerAlert(){
  const box = document.getElementById('power-timer-alert');
  if(box) box.classList.remove('show');
  stopPowerTimerSound();
  return false;
}
async function postPowerTimer(url, bodyObj){
  const fd = new URLSearchParams();
  Object.keys(bodyObj || {}).forEach(function(k){ fd.append(k, bodyObj[k]); });
  const res = await fetch(url, {method:'POST', headers:{'X-Requested-With':'XMLHttpRequest'}, body:fd});
  const data = await res.json();
  if(!res.ok || !data.ok){ throw new Error(data.message || 'حدث خطأ'); }
  return data;
}
function requestPowerTimerNotificationPermission(){
  if(typeof Notification === 'undefined') return;
  if(Notification.permission === 'default'){ Notification.requestPermission().catch(function(){}); }
}
function notifyPowerTimerDone(key){
  if(lastPowerTimerAlertKey === key) return;
  lastPowerTimerAlertKey = key;
  showPowerTimerAlert('انتهى وقت دورة الكهرباء. سيبدأ العد من جديد بعد 10 ثوانٍ.', true);
  startPowerTimerSound();
  if(typeof Notification !== 'undefined' && Notification.permission === 'granted'){
    try{ new Notification('مؤقت الكهرباء', {body:'انتهى الوقت. سيبدأ عد جديد تلقائيًا بعد 10 ثوانٍ.'}); }catch(e){}
  }
}
function updateLiveClock(){
  const now = new Date();
  const timeText = now.toLocaleTimeString('en-GB', {hour:'2-digit', minute:'2-digit', second:'2-digit'});
  const dateText = now.toLocaleDateString('en-GB');
  const ids = ['global-live-clock-time','timer-page-live-time'];
  ids.forEach(function(id){ const el = document.getElementById(id); if(el) el.textContent = timeText; });
  const dates = ['global-live-clock-date','timer-page-live-date'];
  dates.forEach(function(id){ const el = document.getElementById(id); if(el) el.textContent = dateText; });
}
function updateTimerProgressVisual(data){
  const duration = Math.max(60, parseInt(data.duration_seconds || ((data.duration_minutes || 30) * 60), 10));
  let progress = 0;
  if(data.state === 'running'){ progress = ((duration - (data.display_remaining_seconds || 0)) / duration) * 100; }
  else if(data.state === 'alarm'){ progress = 100; }
  else if(data.state === 'paused'){ progress = ((duration - (data.display_remaining_seconds || 0)) / duration) * 100; }
  progress = Math.max(0, Math.min(100, progress));
  const circle = document.getElementById('timer-ring-progress');
  if(circle){
    const radius = 130;
    const circumference = 2 * Math.PI * radius;
    circle.style.strokeDasharray = String(circumference);
    circle.style.strokeDashoffset = String(circumference - (progress / 100) * circumference);
  }
  const strip = document.getElementById('timer-progress-fill');
  if(strip) strip.style.width = progress.toFixed(1) + '%';
  const mini = document.getElementById('dashboard-timer-progress-fill');
  if(mini) mini.style.width = progress.toFixed(1) + '%';
}
function updatePowerTimerWidgets(data){
  const badge = document.getElementById('timer-global-badge');
  const badgeState = document.getElementById('timer-global-badge-state');
  const badgeRemain = document.getElementById('timer-global-badge-remaining');
  if(badge && badgeState && badgeRemain){
    badge.className = 'timer-mini ' + timerStateClass(data.state) + (data.state === 'running' ? ' timer-pulse' : '');
    badgeState.textContent = timerStateLabel(data.state);
    badgeRemain.textContent = data.state === 'stopped' ? '--:--' : formatTimerSeconds(data.display_remaining_seconds || 0);
  }
  const ids = ['dashboard-timer-remaining','timer-page-remaining'];
  ids.forEach(function(id){ const el = document.getElementById(id); if(el) el.textContent = data.state === 'stopped' ? '--:--' : formatTimerSeconds(data.display_remaining_seconds || 0); });
  const states = ['dashboard-timer-state','timer-page-state'];
  states.forEach(function(id){ const el = document.getElementById(id); if(el){ el.textContent = timerStateLabel(data.state); el.className = 'timer-state-badge ' + timerStateClass(data.state) + (data.state === 'running' ? ' timer-pulse' : ''); } });
  const cycle = document.getElementById('timer-page-cycle-minutes');
  if(cycle) cycle.textContent = String(data.duration_minutes || 30);
  const phase = document.getElementById('timer-page-phase');
  if(phase) phase.textContent = data.phase_label || '-';
  const restart = document.getElementById('timer-page-restart');
  if(restart) restart.textContent = String(data.auto_restart_delay_seconds || 10);
  const minutesInput = document.getElementById('timer-minutes-input');
  if(minutesInput && document.activeElement !== minutesInput){ minutesInput.value = String(data.duration_minutes || 30); }
  const pauseBtn = document.getElementById('timer-pause-btn');
  const resumeBtn = document.getElementById('timer-resume-btn');
  const stopBtn = document.getElementById('timer-stop-btn');
  if(pauseBtn) pauseBtn.style.display = data.state === 'running' ? 'inline-flex' : 'none';
  if(resumeBtn) resumeBtn.style.display = data.state === 'paused' ? 'inline-flex' : 'none';
  if(stopBtn) stopBtn.style.display = data.state === 'stopped' ? 'none' : 'inline-flex';
  updateTimerProgressVisual(data);
  if(data.state === 'alarm'){
    notifyPowerTimerDone(data.alert_key || 'alarm');
  } else {
    if(lastPowerTimerAlertKey && data.state === 'running'){ dismissPowerTimerAlert(); }
  }
}
async function refreshPowerTimerStatus(){
  try{
    const res = await fetch('/api/power-timer/status', {headers:{'X-Requested-With':'XMLHttpRequest'}});
    const data = await res.json();
    if(!res.ok || !data.ok){ return; }
    updatePowerTimerWidgets(data);
  }catch(e){}
}
async function startPowerTimer(){
  const input = document.getElementById('timer-minutes-input');
  const minutes = input ? parseInt(input.value || '30', 10) : 30;
  try{ const data = await postPowerTimer('/api/power-timer/start', {minutes: minutes}); showLiveFlash(data.message || 'تم تشغيل المؤقت.', 'success'); refreshPowerTimerStatus(); }
  catch(err){ showLiveFlash(err.message || 'تعذر تشغيل المؤقت', 'error'); }
  return false;
}
async function pausePowerTimer(){ try{ const data = await postPowerTimer('/api/power-timer/pause', {}); showLiveFlash(data.message || 'تم الإيقاف المؤقت.', 'success'); refreshPowerTimerStatus(); }catch(err){ showLiveFlash(err.message || 'تعذر الإيقاف المؤقت', 'error'); } return false; }
async function resumePowerTimer(){ try{ const data = await postPowerTimer('/api/power-timer/resume', {}); showLiveFlash(data.message || 'تم الاستئناف.', 'success'); refreshPowerTimerStatus(); }catch(err){ showLiveFlash(err.message || 'تعذر الاستئناف', 'error'); } return false; }
async function stopPowerTimer(){ if(!confirm('هل تريد إيقاف المؤقت نهائيًا لهذا اليوم؟')) return false; try{ const data = await postPowerTimer('/api/power-timer/stop', {}); showLiveFlash(data.message || 'تم الإيقاف النهائي.', 'success'); dismissPowerTimerAlert(); refreshPowerTimerStatus(); }catch(err){ showLiveFlash(err.message || 'تعذر الإيقاف النهائي', 'error'); } return false; }

document.addEventListener('DOMContentLoaded', function(){
  initBeneficiaryForms();
  try{
    const saved = localStorage.getItem('sidebar-collapsed');
    applySidebarState(saved === null ? true : saved === '1');
  }catch(e){
    applySidebarState(true);
  }
  updateLiveClock();
  setInterval(updateLiveClock, 1000);
  requestPowerTimerNotificationPermission();
  refreshPowerTimerStatus();
  powerTimerPollHandle = setInterval(refreshPowerTimerStatus, 1000);
});
</script>
</head>
<body>
{% if session.get('account_id') %}
<div class="layout" id="app-layout">
  <aside class="sidebar">
    <div class="brand">
      <div class="brand-badge">H</div>
      <div class="brand-text">Hobe Hub<small>Professional+ Edition</small></div>
    </div>
    <div class="nav">
      <a href="{{ url_for('dashboard') }}"><i class="fa-solid fa-gauge"></i><span class="nav-label">لوحة التحكم</span></a>
      <a href="{{ url_for('beneficiaries_page') }}"><i class="fa-solid fa-users"></i><span class="nav-label">المستفيدون</span></a>
      <a href="{{ url_for('power_timer_page') }}"><i class="fa-solid fa-bolt"></i><span class="nav-label">مؤقت الكهرباء</span></a>
      {% if has_permission('add') %}
      <details open>
        <summary><i class="fa-solid fa-user-plus"></i><span class="nav-label">إضافة مستفيد</span></summary>
        <div class="submenu">
          <a href="{{ url_for('add_beneficiary_page') }}?user_type=tawjihi"><i class="fa-solid fa-user-graduate"></i><span class="nav-label">طالب توجيهي</span></a>
          <a href="{{ url_for('add_beneficiary_page') }}?user_type=university"><i class="fa-solid fa-building-columns"></i><span class="nav-label">طالب جامعي</span></a>
          <a href="{{ url_for('add_beneficiary_page') }}?user_type=freelancer"><i class="fa-solid fa-laptop-code"></i><span class="nav-label">فري لانسر</span></a>
        </div>
      </details>
      {% endif %}
      {% if has_permission('import') %}
      <a href="{{ url_for('import_page') }}"><i class="fa-solid fa-file-arrow-up"></i><span class="nav-label">استيراد CSV</span></a>
      {% endif %}
      {% if has_permission('export') %}
      <a href="{{ url_for('export_center') }}"><i class="fa-solid fa-file-arrow-down"></i><span class="nav-label">مركز التصدير</span></a>
      {% endif %}
      {% if has_permission('backup') %}
      <a href="{{ url_for('backup_sql') }}"><i class="fa-solid fa-database"></i><span class="nav-label">Backup SQL</span></a>
      {% endif %}
      <div class="section">الحساب</div>
      <a href="{{ url_for('profile_page') }}"><i class="fa-solid fa-id-badge"></i><span class="nav-label">صفحتي الشخصية</span></a>
      {% if has_permission('manage_accounts') %}
      <a href="{{ url_for('accounts_page') }}"><i class="fa-solid fa-user-shield"></i><span class="nav-label">إدارة المستخدمين</span></a>
      {% endif %}
      {% if has_permission('view_audit_log') %}
      <a href="{{ url_for('audit_log_page') }}"><i class="fa-solid fa-clock-rotate-left"></i><span class="nav-label">سجل العمليات</span></a>
      {% endif %}
    </div>
  </aside>
  <main class="main">
    <div class="topbar">
      <div class="topbar-left"><button class="sidebar-toggle" type="button" onclick="return toggleSidebar()" title="إظهار/إخفاء القائمة الجانبية"><i class="fa-solid fa-bars"></i></button><strong>{{ title }}</strong></div>
      <div class="userbox">
        <div class="topbar-clock"><i class="fa-regular fa-clock"></i><div><div id="global-live-clock-time">--:--:--</div><small id="global-live-clock-date">--/--/----</small></div></div>
        <a href="{{ url_for('power_timer_page') }}" id="timer-global-badge" class="timer-mini timer-status-stopped" title="مؤقت الكهرباء"><i class="fa-solid fa-bolt"></i> <span id="timer-global-badge-state">متوقف</span> <span id="timer-global-badge-remaining">--:--</span></a>
        <span class="badge">{{ session.get('username','') }}</span>
        {% if session.get('full_name') %}<span class="badge">{{ session.get('full_name') }}</span>{% endif %}
        <a class="logout-btn" href="{{ url_for('logout') }}"><i class="fa-solid fa-right-from-bracket"></i> تسجيل الخروج</a>
      </div>
    </div>
    <div id="live-flash-area"></div>
    <div id="power-timer-alert" class="timer-alert"><div id="power-timer-alert-text">انتهى الوقت.</div><div class="actions"><button class="btn btn-danger btn-icon" type="button" onclick="return dismissPowerTimerAlert()" title="إيقاف الصوت"><i class="fa-solid fa-volume-xmark"></i></button></div></div>
    <audio id="power-timer-audio" preload="auto"><source src="https://actions.google.com/sounds/v1/alarms/alarm_clock.ogg" type="audio/ogg"></audio>
    {% with messages = get_flashed_messages(with_categories=true) %}
      {% if messages %}
        {% for cat, msg in messages %}
          <div class="flash {{ cat }}">{{ msg }}</div>
        {% endfor %}
      {% endif %}
    {% endwith %}
    {{ content|safe }}
  </main>
</div>
{% else %}
<div class="main">
  {% with messages = get_flashed_messages(with_categories=true) %}
    {% if messages %}
      {% for cat, msg in messages %}
        <div class="flash {{ cat }}">{{ msg }}</div>
      {% endfor %}
    {% endif %}
  {% endwith %}
  {{ content|safe }}
</div>
{% endif %}
</body>
</html>
"""


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()



def init_connection_pool():
    global _connection_pool
    if _connection_pool is None:
        _connection_pool = pool.ThreadedConnectionPool(
            DB_POOL_MINCONN,
            DB_POOL_MAXCONN,
            get_database_url(),
            connect_timeout=10,
        )
    return _connection_pool


def get_connection():
    return init_connection_pool().getconn()


def release_connection(conn, close=False):
    if conn is None:
        return
    try:
        init_connection_pool().putconn(conn, close=close)
    except Exception:
        try:
            conn.close()
        except Exception:
            pass




def _now_text():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def create_import_task(username: str, account_id: int | None, filename: str) -> str:
    task_id = uuid4().hex
    task = {
        "id": task_id,
        "filename": filename,
        "status": "queued",
        "message": "تم إنشاء مهمة الاستيراد وبانتظار بدء المعالجة.",
        "created_at": _now_text(),
        "started_at": "",
        "finished_at": "",
        "total": 0,
        "processed": 0,
        "inserted": 0,
        "updated": 0,
        "skipped": 0,
        "error_count": 0,
        "percent": 0,
        "current_step": "بانتظار البدء",
        "logs": [],
        "errors": [],
        "username": username or "",
        "account_id": account_id,
    }
    with IMPORT_TASKS_LOCK:
        IMPORT_TASKS[task_id] = task
    return task_id


def get_import_task(task_id: str):
    with IMPORT_TASKS_LOCK:
        task = IMPORT_TASKS.get(task_id)
        if not task:
            return None
        return {**task, "logs": list(task.get("logs", [])), "errors": list(task.get("errors", []))}


def update_import_task(task_id: str, **kwargs):
    with IMPORT_TASKS_LOCK:
        task = IMPORT_TASKS.get(task_id)
        if not task:
            return
        task.update(kwargs)
        total = int(task.get("total") or 0)
        processed = int(task.get("processed") or 0)
        task["percent"] = min(100, round((processed / total) * 100)) if total else 0


def append_import_log(task_id: str, message: str, is_error: bool = False):
    timestamped = f"[{_now_text()}] {message}"
    with IMPORT_TASKS_LOCK:
        task = IMPORT_TASKS.get(task_id)
        if not task:
            return
        task.setdefault("logs", []).append(timestamped)
        if len(task["logs"]) > IMPORT_LOG_LIMIT:
            task["logs"] = task["logs"][-IMPORT_LOG_LIMIT:]
        if is_error:
            task.setdefault("errors", []).append(timestamped)
            task["error_count"] = len(task["errors"])


def finalize_import_task(task_id: str, status: str, message: str):
    update_import_task(task_id, status=status, message=message, finished_at=_now_text(), current_step=message)


def _infer_user_type(data: dict) -> str:
    if data.get("user_type"):
        return data["user_type"]
    if data.get("tawjihi_year") or data.get("tawjihi_branch"):
        return "tawjihi"
    if data.get("university_name") or data.get("university_specialization"):
        return "university"
    return "freelancer"


def _normalize_import_row(row: dict) -> dict:
    data = {c: clean_csv_value(row.get(c, "")) for c in CSV_IMPORT_COLUMNS}
    data["phone"] = normalize_phone(data["phone"])
    data["full_name"] = full_name_from_parts(data["first_name"], data["second_name"], data["third_name"], data["fourth_name"])
    data["search_name"] = normalize_search_ar(data["full_name"])
    data["user_type"] = _infer_user_type(data)
    data["weekly_usage_week_start"] = get_week_start()
    return data


def _build_existing_lookup():
    rows = query_all("SELECT id, phone, full_name, user_type FROM beneficiaries")
    phone_map = {}
    name_type_map = {}
    for r in rows:
        phone = clean_csv_value(r.get("phone"))
        if phone and phone not in phone_map:
            phone_map[phone] = r["id"]
        key = (clean_csv_value(r.get("full_name")), clean_csv_value(r.get("user_type")))
        if key[0] and key not in name_type_map:
            name_type_map[key] = r["id"]
    return phone_map, name_type_map


def _bulk_insert_beneficiaries(cur, records: list[dict]):
    if not records:
        return
    sql = """
        INSERT INTO beneficiaries (
            user_type, first_name, second_name, third_name, fourth_name,
            full_name, search_name, phone, tawjihi_year, tawjihi_branch,
            freelancer_specialization, freelancer_company, freelancer_schedule_type,
            freelancer_internet_method, freelancer_time_mode, freelancer_time_from,
            freelancer_time_to, university_name, university_college, university_specialization,
            university_days, university_internet_method, university_time_mode,
            university_time_from, university_time_to, weekly_usage_count, weekly_usage_week_start,
            notes, added_by_account_id, added_by_username
        ) VALUES %s
    """
    values = [
        (
            d["user_type"], d["first_name"], d["second_name"], d["third_name"], d["fourth_name"],
            d["full_name"], d["search_name"], d["phone"], d["tawjihi_year"], d["tawjihi_branch"],
            d["freelancer_specialization"], d["freelancer_company"], d["freelancer_schedule_type"],
            d["freelancer_internet_method"], d["freelancer_time_mode"], d["freelancer_time_from"],
            d["freelancer_time_to"], d["university_name"], d["university_college"], d["university_specialization"],
            d["university_days"], d["university_internet_method"], d["university_time_mode"],
            d["university_time_from"], d["university_time_to"], 0, d["weekly_usage_week_start"],
            d.get("notes", ""), d.get("added_by_account_id"), d.get("added_by_username", "")
        )
        for d in records
    ]
    execute_values(cur, sql, values, page_size=min(IMPORT_BATCH_SIZE, 500))


def _bulk_update_beneficiaries(cur, records: list[dict]):
    if not records:
        return
    sql = """
        UPDATE beneficiaries SET
            user_type=%s, first_name=%s, second_name=%s, third_name=%s, fourth_name=%s,
            full_name=%s, search_name=%s, phone=%s, tawjihi_year=%s, tawjihi_branch=%s,
            freelancer_specialization=%s, freelancer_company=%s, freelancer_schedule_type=%s,
            freelancer_internet_method=%s, freelancer_time_mode=%s, freelancer_time_from=%s,
            freelancer_time_to=%s, university_name=%s, university_college=%s,
            university_specialization=%s, university_days=%s, university_internet_method=%s,
            university_time_mode=%s, university_time_from=%s, university_time_to=%s,
            notes=%s
        WHERE id=%s
    """
    params = [
        (
            d["user_type"], d["first_name"], d["second_name"], d["third_name"], d["fourth_name"],
            d["full_name"], d["search_name"], d["phone"], d["tawjihi_year"], d["tawjihi_branch"],
            d["freelancer_specialization"], d["freelancer_company"], d["freelancer_schedule_type"],
            d["freelancer_internet_method"], d["freelancer_time_mode"], d["freelancer_time_from"],
            d["freelancer_time_to"], d["university_name"], d["university_college"],
            d["university_specialization"], d["university_days"], d["university_internet_method"],
            d["university_time_mode"], d["university_time_from"], d["university_time_to"],
            d.get("notes", ""), d["id"]
        )
        for d in records
    ]
    execute_batch(cur, sql, params, page_size=min(IMPORT_BATCH_SIZE, 500))


def _process_rows_fallback(cur, task_id: str, op_name: str, batch: list[tuple[int, dict]], is_update: bool):
    ok = 0
    for row_num, data in batch:
        try:
            if is_update:
                _bulk_update_beneficiaries(cur, [data])
            else:
                _bulk_insert_beneficiaries(cur, [data])
            ok += 1
        except Exception as exc:
            cur.connection.rollback()
            append_import_log(task_id, f"خطأ في السطر {row_num} أثناء {op_name}: {exc}", is_error=True)
    return ok


def run_import_task(task_id: str, content: str):
    update_import_task(task_id, status="running", started_at=_now_text(), current_step="قراءة الملف")
    append_import_log(task_id, "بدأت مهمة الاستيراد في الخلفية.")
    conn = None
    cur = None
    try:
        reader = csv.DictReader(io.StringIO(content))
        if not reader.fieldnames:
            finalize_import_task(task_id, "failed", "الملف فارغ أو غير صالح.")
            append_import_log(task_id, "تعذر العثور على رؤوس الأعمدة داخل الملف.", is_error=True)
            return

        normalized_rows = []
        seen_file_keys = set()
        task_meta = get_import_task(task_id) or {}
        import_username = clean_csv_value(task_meta.get("username")) or "استيراد CSV"
        import_account_id = task_meta.get("account_id")
        for idx, row in enumerate(reader, start=2):
            data = _normalize_import_row(row)
            data["added_by_username"] = import_username
            data["added_by_account_id"] = import_account_id
            dedupe_key = (data.get("phone") or "", data.get("full_name") or "", data.get("user_type") or "")
            if dedupe_key in seen_file_keys:
                append_import_log(task_id, f"تم تجاوز السطر {idx} لأنه مكرر داخل نفس الملف.", is_error=True)
                update_import_task(task_id, skipped=(get_import_task(task_id)["skipped"] + 1))
                continue
            seen_file_keys.add(dedupe_key)
            normalized_rows.append((idx, data))

        total = len(normalized_rows)
        update_import_task(task_id, total=total, current_step="تحليل البيانات")
        append_import_log(task_id, f"تمت قراءة الملف بنجاح. عدد السجلات القابلة للمعالجة: {total}.")
        if total == 0:
            finalize_import_task(task_id, "completed", "لا توجد سجلات صالحة للاستيراد.")
            return

        phone_map, name_type_map = _build_existing_lookup()
        inserts = []
        updates = []
        for row_num, data in normalized_rows:
            existing_id = None
            if data.get("phone"):
                existing_id = phone_map.get(data["phone"])
            if not existing_id:
                existing_id = name_type_map.get((data["full_name"], data["user_type"]))
            if existing_id:
                data["id"] = existing_id
                updates.append((row_num, data))
            else:
                inserts.append((row_num, data))
                if data.get("phone"):
                    phone_map[data["phone"]] = -1
                name_type_map[(data["full_name"], data["user_type"])] = -1

        append_import_log(task_id, f"تجهيز العملية: {len(inserts)} سجل جديد، {len(updates)} سجل للتحديث.")
        update_import_task(task_id, current_step="الاتصال بقاعدة البيانات")

        conn = get_connection()
        cur = conn.cursor()

        def process_batches(items, is_update: bool):
            inserted_count = 0
            updated_count = 0
            label = "تحديث" if is_update else "إضافة"
            for start in range(0, len(items), IMPORT_BATCH_SIZE):
                batch = items[start:start + IMPORT_BATCH_SIZE]
                batch_records = [data for _, data in batch]
                ok_count = 0
                try:
                    if is_update:
                        _bulk_update_beneficiaries(cur, batch_records)
                    else:
                        _bulk_insert_beneficiaries(cur, batch_records)
                    conn.commit()
                    ok_count = len(batch)
                except Exception as exc:
                    conn.rollback()
                    append_import_log(task_id, f"فشل {label} دفعة من {len(batch)} سجلات، سيتم التحويل إلى معالجة تفصيلية: {exc}", is_error=True)
                    ok_count = _process_rows_fallback(cur, task_id, label, batch, is_update)
                    conn.commit()
                if is_update:
                    updated_count += ok_count
                else:
                    inserted_count += ok_count
                task_snapshot = get_import_task(task_id)
                update_import_task(
                    task_id,
                    processed=(task_snapshot["processed"] + len(batch)),
                    inserted=(task_snapshot["inserted"] + (0 if is_update else ok_count)),
                    updated=(task_snapshot["updated"] + (ok_count if is_update else 0)),
                    current_step=f"{label} البيانات... {min(start + len(batch), len(items))}/{len(items)}"
                )
                append_import_log(task_id, f"{label} دفعة: {min(start + len(batch), len(items))}/{len(items)}. نجاح {ok_count} من {len(batch)}.")
            return inserted_count, updated_count

        total_inserted = 0
        total_updated = 0
        if inserts:
            ins, _ = process_batches(inserts, is_update=False)
            total_inserted += ins
            update_import_task(task_id, inserted=total_inserted)
        if updates:
            _, upd = process_batches(updates, is_update=True)
            total_updated += upd
            update_import_task(task_id, updated=total_updated)

        task_snapshot = get_import_task(task_id)
        msg = f"اكتملت المعالجة. تمت إضافة {task_snapshot['inserted']} وتحديث {task_snapshot['updated']} مع {task_snapshot['error_count']} خطأ."
        finalize_import_task(task_id, "completed", msg)
        append_import_log(task_id, msg)

        if task_snapshot.get("account_id"):
            conn2 = get_connection()
            cur2 = conn2.cursor()
            try:
                cur2.execute(
                    """
                    INSERT INTO audit_logs (account_id, username_snapshot, action_type, target_type, target_id, details)
                    VALUES (%s,%s,%s,%s,%s,%s)
                    """,
                    [task_snapshot.get("account_id"), task_snapshot.get("username"), "import", "beneficiary", None, msg]
                )
                conn2.commit()
            finally:
                cur2.close()
                release_connection(conn2)
    except Exception as exc:
        if conn is not None:
            try:
                conn.rollback()
            except Exception:
                pass
        append_import_log(task_id, f"توقفت المهمة بسبب خطأ عام: {exc}", is_error=True)
        finalize_import_task(task_id, "failed", f"فشلت المعالجة: {exc}")
    finally:
        if cur is not None:
            cur.close()
        if conn is not None:
            release_connection(conn)


def launch_import_task(task_id: str, content: str):
    thread = threading.Thread(target=run_import_task, args=(task_id, content), daemon=True)
    thread.start()

def query_all(sql, params=None):
    conn = get_connection()
    cur = None
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(sql, params or [])
        return cur.fetchall()
    finally:
        if cur is not None:
            cur.close()
        release_connection(conn)


def query_one(sql, params=None):
    conn = get_connection()
    cur = None
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(sql, params or [])
        return cur.fetchone()
    finally:
        if cur is not None:
            cur.close()
        release_connection(conn)


def execute_sql(sql, params=None, fetchone=False):
    conn = get_connection()
    cur = None
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(sql, params or [])
        row = cur.fetchone() if fetchone else None
        conn.commit()
        return row
    except Exception:
        conn.rollback()
        raise
    finally:
        if cur is not None:
            cur.close()
        release_connection(conn)


def safe(v):
    return "" if v is None else str(v)


def format_dt_short(value):
    if not value:
        return ""
    try:
        return value.strftime('%Y-%m-%d %H:%M')
    except AttributeError:
        text = str(value)
        return text[:16] if len(text) >= 16 else text


def format_dt_compact(value):
    if not value:
        return ""
    try:
        return value.strftime('%Y-%m-%d %H:%M')
    except AttributeError:
        text = str(value).replace('T', ' ')
        return text[:16] if len(text) >= 16 else text


def action_type_label(action_type=None):
    return {
        'login': 'تسجيل دخول',
        'logout': 'تسجيل خروج',
        'add': 'إضافة',
        'edit': 'تعديل',
        'delete': 'حذف',
        'bulk_delete': 'حذف جماعي',
        'import': 'استيراد',
        'usage_counter': 'إضافة بطاقة',
        'reset_weekly_usage': 'تجديد البطاقات',
        'power_timer_start': 'تشغيل المؤقت',
        'power_timer_pause': 'إيقاف مؤقت للمؤقت',
        'power_timer_resume': 'استئناف المؤقت',
        'power_timer_stop': 'إيقاف المؤقت',
        'export': 'تصدير',
        'backup': 'نسخة احتياطية',
    }.get(clean_csv_value(action_type), safe(action_type or ''))


def target_type_label(target_type=None):
    return {
        'beneficiary': 'مستفيد',
        'account': 'حساب',
        'power_timer': 'مؤقت الكهرباء',
    }.get(clean_csv_value(target_type), safe(target_type or ''))


def permission_label(permission_name=None):
    key = clean_csv_value(permission_name)
    return PERMISSION_LABELS.get(key, safe(permission_name or ''))


def clean_csv_value(v):
    if v is None:
        return ""
    t = str(v).strip()
    if t.lower() in ("nan", "none", "null"):
        return ""
    return t


def normalize_phone(phone):
    text = clean_csv_value(phone)
    if text.endswith(".0"):
        text = text[:-2]
    return "".join(ch for ch in text if ch.isdigit())


def normalize_search_ar(text):
    text = clean_csv_value(text)
    repl = {
        "أ": "ا", "إ": "ا", "آ": "ا", "ى": "ي", "ة": "ه", "ؤ": "و", "ئ": "ي",
    }
    for old, new in repl.items():
        text = text.replace(old, new)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def split_full_name(name):
    parts = clean_csv_value(name).split()
    parts += [""] * (4 - len(parts))
    first = parts[0] if len(parts) > 0 else ""
    second = parts[1] if len(parts) > 1 else ""
    third = parts[2] if len(parts) > 2 else ""
    fourth = " ".join(parts[3:]).strip() if len(parts) > 3 else ""
    return first, second, third, fourth


def full_name_from_parts(first, second, third, fourth):
    return " ".join([x.strip() for x in [first, second, third, fourth] if clean_csv_value(x)]).strip()


def get_week_start(today=None):
    today = today or date.today()
    delta = (today.weekday() - 5) % 7
    return today - timedelta(days=delta)


def log_action(action_type, target_type="", target_id=None, details=""):
    if not session.get("account_id"):
        return
    execute_sql("""
        INSERT INTO audit_logs (account_id, username_snapshot, action_type, target_type, target_id, details)
        VALUES (%s,%s,%s,%s,%s,%s)
    """, [
        session.get("account_id"),
        session.get("username"),
        action_type,
        target_type,
        target_id,
        details,
    ])


def get_account_permissions(account_id):
    rows = query_all("""
        SELECT p.name
        FROM account_permissions ap
        JOIN permissions p ON p.id = ap.permission_id
        WHERE ap.account_id = %s
    """, [account_id])
    return [r["name"] for r in rows]


def refresh_session_permissions(account_id):
    session["permissions"] = get_account_permissions(account_id)


def has_permission(permission_name):
    aid = session.get("account_id")
    if not aid:
        return False
    refresh_session_permissions(aid)
    return permission_name in session.get("permissions", [])


@app.context_processor
def inject_helpers():
    return {"has_permission": has_permission, "session": session}


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("account_id"):
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapped


def permission_required(permission_name):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            if not session.get("account_id"):
                return redirect(url_for("login"))
            if not has_permission(permission_name):
                flash("غير مصرح لك بهذه العملية.", "error")
                return redirect(url_for("dashboard"))
            return view(*args, **kwargs)
        return wrapped
    return decorator


def render_page(title, content):
    return render_template_string(BASE_TEMPLATE, title=title, content=content)


def setup_database():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS beneficiaries (
        id SERIAL PRIMARY KEY,
        user_type TEXT NOT NULL,
        first_name TEXT,
        second_name TEXT,
        third_name TEXT,
        fourth_name TEXT,
        full_name TEXT,
        search_name TEXT,
        phone TEXT,
        tawjihi_year TEXT,
        tawjihi_branch TEXT,
        freelancer_specialization TEXT,
        freelancer_company TEXT,
        freelancer_schedule_type TEXT,
        freelancer_internet_method TEXT,
        freelancer_time_mode TEXT,
        freelancer_time_from TEXT,
        freelancer_time_to TEXT,
        university_name TEXT,
        university_college TEXT,
        university_specialization TEXT,
        university_days TEXT,
        university_internet_method TEXT,
        university_time_mode TEXT,
        university_time_from TEXT,
        university_time_to TEXT,
        weekly_usage_count INTEGER DEFAULT 0,
        weekly_usage_week_start DATE,
        notes TEXT DEFAULT '',
        added_by_account_id INTEGER,
        added_by_username TEXT DEFAULT '',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)

    cur.execute("ALTER TABLE beneficiaries ADD COLUMN IF NOT EXISTS notes TEXT DEFAULT ''")
    cur.execute("ALTER TABLE beneficiaries ADD COLUMN IF NOT EXISTS added_by_account_id INTEGER")
    cur.execute("ALTER TABLE beneficiaries ADD COLUMN IF NOT EXISTS added_by_username TEXT DEFAULT ''")
    cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS beneficiaries_phone_unique_idx ON beneficiaries (phone) WHERE phone IS NOT NULL AND btrim(phone) <> ''")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS app_accounts (
        id SERIAL PRIMARY KEY,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT,
        full_name TEXT DEFAULT '',
        is_active BOOLEAN DEFAULT TRUE,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)

    # توافق مع قواعد بيانات قديمة كانت تحتوي أعمدة مختلفة أو ناقصة.
    cur.execute("ALTER TABLE app_accounts ADD COLUMN IF NOT EXISTS password_hash TEXT")
    cur.execute("ALTER TABLE app_accounts ADD COLUMN IF NOT EXISTS full_name TEXT DEFAULT ''")
    cur.execute("ALTER TABLE app_accounts ADD COLUMN IF NOT EXISTS is_active BOOLEAN DEFAULT TRUE")
    cur.execute("ALTER TABLE app_accounts ADD COLUMN IF NOT EXISTS created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP")
    cur.execute("CREATE EXTENSION IF NOT EXISTS pgcrypto")

    cur.execute("""
        SELECT EXISTS (
            SELECT 1
            FROM information_schema.columns
            WHERE table_name = 'app_accounts' AND column_name = 'password'
        )
    """)
    has_old_password_col = cur.fetchone()[0]
    if has_old_password_col:
        cur.execute("""
            UPDATE app_accounts
            SET password_hash = CASE
                WHEN password_hash IS NOT NULL AND btrim(password_hash) <> '' THEN password_hash
                WHEN password IS NULL OR btrim(password) = '' THEN NULL
                WHEN password ~ '^[a-f0-9]{64}$' THEN password
                ELSE encode(digest(password, 'sha256'), 'hex')
            END
            WHERE password_hash IS NULL OR btrim(password_hash) = ''
        """)

    cur.execute("""
        UPDATE app_accounts
        SET password_hash = %s
        WHERE username = 'admin' AND (password_hash IS NULL OR btrim(password_hash) = '')
    """, [sha256_text("123456")])

    cur.execute("""
    CREATE TABLE IF NOT EXISTS permissions (
        id SERIAL PRIMARY KEY,
        name TEXT UNIQUE NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS account_permissions (
        account_id INTEGER REFERENCES app_accounts(id) ON DELETE CASCADE,
        permission_id INTEGER REFERENCES permissions(id) ON DELETE CASCADE,
        PRIMARY KEY (account_id, permission_id)
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS audit_logs (
        id SERIAL PRIMARY KEY,
        account_id INTEGER,
        username_snapshot TEXT,
        action_type TEXT,
        target_type TEXT,
        target_id INTEGER,
        details TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)


    cur.execute("""
    CREATE TABLE IF NOT EXISTS power_timer (
        id INTEGER PRIMARY KEY,
        duration_minutes INTEGER NOT NULL DEFAULT 30,
        cycle_started_at TIMESTAMP NULL,
        paused_remaining_seconds INTEGER NULL,
        auto_restart_delay_seconds INTEGER NOT NULL DEFAULT 10,
        state TEXT NOT NULL DEFAULT 'stopped',
        updated_by_username TEXT DEFAULT '',
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)
    cur.execute("""
        INSERT INTO power_timer (id, duration_minutes, auto_restart_delay_seconds, state, updated_by_username)
        VALUES (1, 30, 10, 'stopped', '')
        ON CONFLICT (id) DO NOTHING
    """)

    for perm in PERMISSIONS:
        cur.execute("INSERT INTO permissions (name) VALUES (%s) ON CONFLICT (name) DO NOTHING", [perm])

    cur.execute("SELECT id FROM app_accounts WHERE username='admin' LIMIT 1")
    admin_row = cur.fetchone()
    if not admin_row:
        cur.execute("""
            INSERT INTO app_accounts (username, password_hash, full_name, is_active)
            VALUES (%s,%s,%s,TRUE)
            RETURNING id
        """, ["admin", sha256_text("123456"), "System Administrator"])
        admin_id = cur.fetchone()[0]
    else:
        admin_id = admin_row[0]

    cur.execute("""
        INSERT INTO account_permissions (account_id, permission_id)
        SELECT %s, id FROM permissions
        ON CONFLICT DO NOTHING
    """, [admin_id])

    conn.commit()
    cur.close()
    release_connection(conn)


try:
    setup_database()
except psycopg2.Error as exc:
    print(f"[DB INIT ERROR] {exc}")


def normalize_beneficiary_usage(user_id):
    row = query_one("SELECT weekly_usage_week_start FROM beneficiaries WHERE id=%s", [user_id])
    if not row:
        return
    current_start = get_week_start()
    saved = row["weekly_usage_week_start"]
    if saved != current_start:
        execute_sql("""
            UPDATE beneficiaries
            SET weekly_usage_count = 0, weekly_usage_week_start = %s
            WHERE id = %s
        """, [current_start, user_id])


def normalize_all_usage():
    execute_sql("""
        UPDATE beneficiaries
        SET weekly_usage_count = 0, weekly_usage_week_start = %s
        WHERE weekly_usage_week_start IS DISTINCT FROM %s
    """, [get_week_start(), get_week_start()])


@app.route("/")
def root():
    if session.get("account_id"):
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("account_id"):
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        username = clean_csv_value(request.form.get("username"))
        password = clean_csv_value(request.form.get("password"))
        row = query_one("""
            SELECT * FROM app_accounts
            WHERE username=%s AND password_hash=%s AND is_active=TRUE
            LIMIT 1
        """, [username, sha256_text(password)])
        if row:
            session["account_id"] = row["id"]
            session["username"] = row["username"]
            session["full_name"] = row["full_name"]
            refresh_session_permissions(row["id"])
            log_action("login", "account", row["id"], "تسجيل دخول")
            return redirect(url_for("dashboard"))
        flash("اسم المستخدم أو كلمة المرور غير صحيحة.", "error")
    content = """
    <div class="login-wrap">
      <div class="hero">
        <h1>تسجيل الدخول</h1>
      </div>
      <div class="card">
        <form method="POST">
          <div class="row">
            <div><label>اسم المستخدم</label><input name="username" required></div>
            <div><label>كلمة المرور</label><input type="password" name="password" required></div>
          </div>
          <div class="actions" style="margin-top:14px">
            <button class="btn btn-primary" type="submit">دخول</button>
          </div>
        </form>
      </div>
    </div>
    """
    return render_page("تسجيل الدخول", content)


@app.route("/logout")
@login_required
def logout():
    log_action("logout", "account", session.get("account_id"), "تسجيل خروج")
    session.clear()
    return redirect(url_for("login"))


def get_type_label(user_type=None):
    return {
        "tawjihi": "توجيهي",
        "university": "جامعة",
        "freelancer": "فري لانسر",
    }.get(user_type, safe(user_type or ""))


def get_type_css(user_type=None):
    return {
        "tawjihi": "type-green",
        "university": "type-purple",
        "freelancer": "type-blue",
    }.get(user_type, "type-default")


def get_usage_label(row):
    count = row.get("weekly_usage_count") or 0
    limited = (
        row.get("user_type") == "tawjihi"
        or (row.get("user_type") == "freelancer" and safe(row.get("freelancer_internet_method")) == "نظام البطاقات")
        or (row.get("user_type") == "university" and safe(row.get("university_internet_method")) == "نظام البطاقات")
    )
    return (f"{count} / 3" if limited else "غير محدود"), limited, count


def distinct_values(column, user_type=None):
    sql = f"SELECT DISTINCT {column} AS value FROM beneficiaries WHERE COALESCE({column}, '') <> ''"
    params = []
    if user_type:
        sql += " AND user_type = %s"
        params.append(user_type)
    sql += f" ORDER BY {column}"
    return [r["value"] for r in query_all(sql, params)]


def build_chart_rows(items, label_key="label", value_key="value", empty_text="لا توجد بيانات كافية."):
    if not items:
        return f"<div class='empty-state'>{empty_text}</div>"
    max_value = max(int(item[value_key]) for item in items) or 1
    html = "<div class='bar-list'>"
    for item in items:
        label = safe(item[label_key])
        value = int(item[value_key])
        width = max(8, round((value / max_value) * 100))
        html += f"""
        <div class="bar-row">
          <div class="bar-label" title="{label}">{label}</div>
          <div class="bar-track"><div class="bar-fill" style="width:{width}%"></div></div>
          <div><strong>{value}</strong></div>
        </div>
        """
    html += "</div>"
    return html




def get_power_timer_row():
    row = query_one("SELECT * FROM power_timer WHERE id=1")
    if row:
        return row
    execute_sql("""
        INSERT INTO power_timer (id, duration_minutes, auto_restart_delay_seconds, state, updated_by_username)
        VALUES (1, 30, 10, 'stopped', '')
        ON CONFLICT (id) DO NOTHING
    """)
    return query_one("SELECT * FROM power_timer WHERE id=1")


def build_power_timer_status(row=None):
    row = row or get_power_timer_row()
    duration_minutes = int(row.get('duration_minutes') or 30)
    duration_seconds = max(60, duration_minutes * 60)
    restart_delay = int(row.get('auto_restart_delay_seconds') or 10)
    state = (row.get('state') or 'stopped')
    paused_remaining = row.get('paused_remaining_seconds')
    start_at = row.get('cycle_started_at')
    now = datetime.now()
    payload = {
        'ok': True,
        'state': state,
        'duration_minutes': duration_minutes,
        'duration_seconds': duration_seconds,
        'auto_restart_delay_seconds': restart_delay,
        'display_remaining_seconds': duration_seconds if state == 'stopped' else 0,
        'phase_label': 'متوقف',
        'alert_key': '',
        'updated_by_username': safe(row.get('updated_by_username')),
    }
    if state == 'paused':
        remaining = max(0, int(paused_remaining or duration_seconds))
        payload.update({'display_remaining_seconds': remaining, 'phase_label': 'متوقف مؤقتًا'})
        return payload
    if state != 'running' or not start_at:
        return payload
    if isinstance(start_at, str):
        try:
            start_at = datetime.fromisoformat(start_at)
        except Exception:
            return payload
    elapsed = max(0, int((now - start_at).total_seconds()))
    cycle_total = duration_seconds + restart_delay
    pos = elapsed % cycle_total
    cycle_index = elapsed // cycle_total
    if pos < duration_seconds:
        remaining = duration_seconds - pos
        payload.update({'state': 'running', 'display_remaining_seconds': remaining, 'phase_label': 'العد التنازلي يعمل'})
    else:
        remaining = cycle_total - pos
        payload.update({'state': 'alarm', 'display_remaining_seconds': remaining, 'phase_label': 'انتهى الوقت - إعادة التشغيل بعد قليل', 'alert_key': f"{start_at.isoformat()}::{cycle_index}"})
    return payload


@app.route("/dashboard")
@login_required
def dashboard():
    normalize_all_usage()
    stats = {
        "all_count": query_one("SELECT COUNT(*) AS c FROM beneficiaries")["c"],
        "tawjihi_count": query_one("SELECT COUNT(*) AS c FROM beneficiaries WHERE user_type='tawjihi'")["c"],
        "university_count": query_one("SELECT COUNT(*) AS c FROM beneficiaries WHERE user_type='university'")["c"],
        "freelancer_count": query_one("SELECT COUNT(*) AS c FROM beneficiaries WHERE user_type='freelancer'")["c"],
        "accounts_count": query_one("SELECT COUNT(*) AS c FROM app_accounts")["c"],
        "active_accounts": query_one("SELECT COUNT(*) AS c FROM app_accounts WHERE is_active=TRUE")["c"],
        "card_based_count": query_one("""
            SELECT COUNT(*) AS c FROM beneficiaries
            WHERE user_type='tawjihi'
               OR (user_type='freelancer' AND freelancer_internet_method='نظام البطاقات')
               OR (user_type='university' AND university_internet_method='نظام البطاقات')
        """)["c"],
        "week_usage_total": query_one("SELECT COALESCE(SUM(weekly_usage_count),0) AS c FROM beneficiaries")["c"],
    }

    type_distribution = query_all("""
        SELECT
          CASE user_type
            WHEN 'tawjihi' THEN 'توجيهي'
            WHEN 'university' THEN 'جامعة'
            WHEN 'freelancer' THEN 'فري لانسر'
            ELSE user_type
          END AS label,
          COUNT(*) AS value
        FROM beneficiaries
        GROUP BY user_type
        ORDER BY value DESC
    """)

    tawjihi_by_year = query_all("""
        SELECT COALESCE(tawjihi_year, 'غير محدد') AS label, COUNT(*) AS value
        FROM beneficiaries
        WHERE user_type='tawjihi'
        GROUP BY COALESCE(tawjihi_year, 'غير محدد')
        ORDER BY value DESC, label
    """)

    universities_top = query_all("""
        SELECT COALESCE(university_name, 'غير محدد') AS label, COUNT(*) AS value
        FROM beneficiaries
        WHERE user_type='university'
        GROUP BY COALESCE(university_name, 'غير محدد')
        ORDER BY value DESC, label
        LIMIT 6
    """)

    recent_logs = query_all("""
        SELECT username_snapshot, action_type, target_type, details, created_at
        FROM audit_logs
        ORDER BY created_at DESC
        LIMIT 8
    """)

    logs_html = ""
    if recent_logs:
        logs_html = "<div class='table-wrap'><table><thead><tr><th>المستخدم</th><th>العملية</th><th>الهدف</th><th>التفاصيل</th><th>الوقت</th></tr></thead><tbody>"
        for r in recent_logs:
            logs_html += f"<tr><td>{safe(r['username_snapshot'])}</td><td>{action_type_label(r['action_type'])}</td><td>{target_type_label(r['target_type'])}</td><td class='cell-wrap'>{safe(r['details'])}</td><td>{format_dt_compact(r['created_at'])}</td></tr>"
        logs_html += "</tbody></table></div>"
    else:
        logs_html = "<div class='empty-state'>لا توجد عمليات بعد.</div>"

    quick_cards = []
    if has_permission("add"):
        quick_cards.extend([
            f"<a class='menu-card' href='{url_for('add_beneficiary_page')}?user_type=tawjihi'><div class='menu-icon'><i class='fa-solid fa-user-graduate'></i></div><h3>إضافة طالب توجيهي</h3><p>إدخال طالب جديد مع السنة والفرع والجوال.</p></a>",
            f"<a class='menu-card' href='{url_for('add_beneficiary_page')}?user_type=freelancer'><div class='menu-icon'><i class='fa-solid fa-laptop-code'></i></div><h3>إضافة فري لانسر</h3><p>تخصص، شركة، نظام دوام وطريقة الإنترنت.</p></a>",
            f"<a class='menu-card' href='{url_for('add_beneficiary_page')}?user_type=university'><div class='menu-icon'><i class='fa-solid fa-building-columns'></i></div><h3>إضافة طالب جامعي</h3><p>جامعة، كلية، تخصص وأيام الحضور.</p></a>",
        ])
    quick_cards.append(f"<a class='menu-card' href='{url_for('beneficiaries_page')}'><div class='menu-icon'><i class='fa-solid fa-users-viewfinder'></i></div><h3>المستفيدون</h3><p>فلترة متقدمة، تبويبات، ترتيب، وتعديل منبثق.</p></a>")
    if has_permission("manage_accounts"):
        quick_cards.append(f"<a class='menu-card' href='{url_for('accounts_page')}'><div class='menu-icon'><i class='fa-solid fa-user-shield'></i></div><h3>إدارة المستخدمين</h3><p>الحسابات والصلاحيات وحالة كل مستخدم.</p></a>")

    content = f"""
    <div class="hero">
      <h1>لوحة التحكم الاحترافية</h1>
      <p>إحصائيات مباشرة ورسوم مرئية ووصول سريع لكل أجزاء النظام.</p>
    </div>

    <div class="grid">
      <div class="stat"><h3>إجمالي المستفيدين</h3><div class="num">{stats['all_count']}</div></div>
      <div class="stat"><h3>طلاب التوجيهي</h3><div class="num">{stats['tawjihi_count']}</div></div>
      <div class="stat"><h3>الطلاب الجامعيون</h3><div class="num">{stats['university_count']}</div></div>
      <div class="stat"><h3>الفري لانسر</h3><div class="num">{stats['freelancer_count']}</div></div>
      <div class="stat"><h3>حسابات النظام</h3><div class="num">{stats['accounts_count']}</div></div>
      <div class="stat"><h3>الحسابات الفعالة</h3><div class="num">{stats['active_accounts']}</div></div>
    </div>

    <div class="kpi-strip">
      <div class="kpi"><div class="label">إجمالي الاستفادات هذا الأسبوع</div><div class="value">{stats['week_usage_total']}</div></div>
      <div class="kpi"><div class="label">الخاضعون لنظام البطاقات / الحد الأسبوعي</div><div class="value">{stats['card_based_count']}</div></div>
      <div class="kpi"><div class="label">نسبة الجامعات من الإجمالي</div><div class="value">{round((stats['university_count'] / stats['all_count']) * 100) if stats['all_count'] else 0}%</div></div>
      <div class="kpi"><div class="label">نسبة التوجيهي من الإجمالي</div><div class="value">{round((stats['tawjihi_count'] / stats['all_count']) * 100) if stats['all_count'] else 0}%</div></div>
    </div>

    <div class="grid-3" style="margin-top:16px">
      <div class="chart-card">
        <h3>التوزيع حسب النوع</h3>
        {build_chart_rows(type_distribution)}
      </div>
      <div class="chart-card">
        <h3>طلاب التوجيهي حسب السنة</h3>
        {build_chart_rows(tawjihi_by_year)}
      </div>
      <div class="chart-card">
        <h3>أكثر الجامعات حضورًا</h3>
        {build_chart_rows(universities_top)}
      </div>
    </div>

    <div class="card" style="margin-top:16px">
      <h3 style="margin-top:0">وصول سريع</h3>
      <div class="mini-grid">{''.join(quick_cards)}</div>
    </div>

    <div class="grid-2" style="margin-top:16px">
      <div class="card dashboard-timer-card">
        <h3 style="margin-top:0">مؤقت الكهرباء</h3>
        <div class="actions" style="justify-content:space-between;align-items:center">
          <div id="dashboard-timer-state" class="timer-state-badge timer-status-stopped">متوقف</div>
          <a class="btn btn-secondary" href="{url_for('power_timer_page')}"><i class="fa-solid fa-bolt"></i> فتح صفحة المؤقت</a>
        </div>
        <div id="dashboard-timer-remaining" class="timer-big">--:--</div>
        <div class="timer-progress-strip"><div id="dashboard-timer-progress-fill" class="timer-progress-fill"></div></div>
        <div class="small" style="margin-top:10px">يُحدَّث تلقائيًا من جميع صفحات النظام مع تنبيه وصوت عند انتهاء الوقت.</div>
      </div>
      <div class="card">
        <h3 style="margin-top:0">آخر العمليات</h3>
        {logs_html}
      </div>
    </div>
    """
    return render_page("لوحة التحكم", content)


def build_query_string(args_dict):
    parts = []
    for k, v in args_dict.items():
        if v not in (None, ""):
            parts.append(f"{k}={v}")
    return "&".join(parts)


def beneficiary_sort_link(args_dict, column):
    current_sort = clean_csv_value(args_dict.get("sort_by", "id"))
    current_order = clean_csv_value(args_dict.get("sort_order", "desc"))
    next_order = "asc" if current_sort != column or current_order == "desc" else "desc"
    query = build_query_string({**args_dict, "sort_by": column, "sort_order": next_order, "page": 1})
    return f"?{query}" if query else ""


def format_modal_fields(data=None, action="", scope_id="beneficiary-form", submit_label="حفظ", show_type_selector=True, fixed_user_type=None):
    data = data or {}
    selected_type = clean_csv_value(fixed_user_type or data.get("user_type", "tawjihi")) or "tawjihi"
    years_html = "".join([f'<option value="{y}" {"selected" if safe(data.get("tawjihi_year", "")) == y else ""}>{y}</option>' for y in TAWJIHI_YEARS])
    branches_html = "".join([f'<option value="{b}" {"selected" if safe(data.get("tawjihi_branch", "")) == b else ""}>{b}</option>' for b in TAWJIHI_BRANCHES])
    freelancer_schedule_html = "".join([f'<option value="{x}" {"selected" if safe(data.get("freelancer_schedule_type", "")) == x else ""}>{x}</option>' for x in FREELANCER_SCHEDULE_OPTIONS])
    time_mode_f_html = "".join([f'<option value="{x}" {"selected" if safe(data.get("freelancer_time_mode", "")) == x else ""}>{x}</option>' for x in TIME_MODE_OPTIONS])
    internet_f_html = "".join([f'<option value="{x}" {"selected" if safe(data.get("freelancer_internet_method", "")) == x else ""}>{x}</option>' for x in INTERNET_ACCESS_METHOD_OPTIONS])
    universities_html = "".join([f'<option value="{x}" {"selected" if safe(data.get("university_name", "")) == x else ""}>{x}</option>' for x in UNIVERSITIES_GAZA])
    days_html = "".join([f'<option value="{x}" {"selected" if safe(data.get("university_days", "")) == x else ""}>{x}</option>' for x in UNIVERSITY_DAYS_OPTIONS])
    time_mode_u_html = "".join([f'<option value="{x}" {"selected" if safe(data.get("university_time_mode", "")) == x else ""}>{x}</option>' for x in TIME_MODE_OPTIONS])
    internet_u_html = "".join([f'<option value="{x}" {"selected" if safe(data.get("university_internet_method", "")) == x else ""}>{x}</option>' for x in INTERNET_ACCESS_METHOD_OPTIONS])

    if show_type_selector:
        type_selector_html = f"""
        <div>
          <label>النوع</label>
          <select name="user_type" required onchange="toggleBeneficiarySections(this, '{scope_id}')">
            <option value="tawjihi" {"selected" if selected_type == "tawjihi" else ""}>توجيهي</option>
            <option value="university" {"selected" if selected_type == "university" else ""}>جامعة</option>
            <option value="freelancer" {"selected" if selected_type == "freelancer" else ""}>فري لانسر</option>
          </select>
        </div>
        """
    else:
        type_selector_html = f"""
        <div>
          <label>النوع</label>
          <input value="{get_type_label(selected_type)}" disabled>
          <input type="hidden" name="user_type" value="{selected_type}">
        </div>
        """

    return f"""
    <form method="POST" action="{action}">
      <div id="{scope_id}" data-beneficiary-scope="1">
        <div class="row">
          {type_selector_html}
          <div><label>الاسم الأول</label><input name="first_name" value="{safe(data.get('first_name', ''))}" required></div>
          <div><label>الاسم الثاني</label><input name="second_name" value="{safe(data.get('second_name', ''))}"></div>
          <div><label>الاسم الثالث</label><input name="third_name" value="{safe(data.get('third_name', ''))}"></div>
          <div><label>الاسم الرابع</label><input name="fourth_name" value="{safe(data.get('fourth_name', ''))}"></div>
          <div><label>الجوال</label><input name="phone" value="{safe(data.get('phone', ''))}"></div>
        </div>

        <div class="form-section section-tawjihi {'active' if selected_type == 'tawjihi' else ''}">
          <div class="section-title">بيانات التوجيهي</div>
          <div class="row">
            <div><label>سنة التوجيهي</label><select name="tawjihi_year"><option value="">اختر السنة</option>{years_html}</select></div>
            <div><label>فرع التوجيهي</label><select name="tawjihi_branch"><option value="">اختر الفرع</option>{branches_html}</select></div>
          </div>
        </div>

        <div class="form-section section-freelancer {'active' if selected_type == 'freelancer' else ''}">
          <div class="section-title">بيانات الفري لانسر</div>
          <div class="row">
            <div><label>التخصص</label><input name="freelancer_specialization" value="{safe(data.get('freelancer_specialization', ''))}"></div>
            <div><label>الشركة</label><input name="freelancer_company" value="{safe(data.get('freelancer_company', ''))}"></div>
            <div><label>نوع الدوام</label><select name="freelancer_schedule_type"><option value="">اختر نوع الدوام</option>{freelancer_schedule_html}</select></div>
            <div><label>طريقة الاتصال</label><select name="freelancer_internet_method"><option value="">اختر الطريقة</option>{internet_f_html}</select></div>
            <div><label>وضع الوقت</label><select name="freelancer_time_mode"><option value="">اختر وضع الوقت</option>{time_mode_f_html}</select></div>
            <div><label>من ساعة</label><input type="time" name="freelancer_time_from" value="{safe(data.get('freelancer_time_from', ''))}"></div>
            <div><label>إلى ساعة</label><input type="time" name="freelancer_time_to" value="{safe(data.get('freelancer_time_to', ''))}"></div>
          </div>
        </div>

        <div class="form-section section-university {'active' if selected_type == 'university' else ''}">
          <div class="section-title">بيانات الطالب الجامعي</div>
          <div class="row">
            <div><label>الجامعة</label><select name="university_name"><option value="">اختر الجامعة</option>{universities_html}</select></div>
            <div><label>الكلية</label><input name="university_college" value="{safe(data.get('university_college', ''))}"></div>
            <div><label>التخصص</label><input name="university_specialization" value="{safe(data.get('university_specialization', ''))}"></div>
            <div><label>أيام الجامعة</label><select name="university_days"><option value="">اختر الأيام</option>{days_html}</select></div>
            <div><label>طريقة الاتصال</label><select name="university_internet_method"><option value="">اختر الطريقة</option>{internet_u_html}</select></div>
            <div><label>وضع الوقت</label><select name="university_time_mode"><option value="">اختر وضع الوقت</option>{time_mode_u_html}</select></div>
            <div><label>من ساعة</label><input type="time" name="university_time_from" value="{safe(data.get('university_time_from', ''))}"></div>
            <div><label>إلى ساعة</label><input type="time" name="university_time_to" value="{safe(data.get('university_time_to', ''))}"></div>
          </div>
        </div>

        <div class="section-title">ملاحظات</div>
        <div><textarea class="notes-box" name="notes" placeholder="اكتب أي ملاحظة تخص المستفيد">{safe(data.get('notes', ''))}</textarea></div>
      </div>

      <div class="actions" style="margin-top:14px">
        <button class="btn btn-primary" type="submit"><i class="fa-solid fa-floppy-disk"></i> {submit_label}</button>
        <a class="btn btn-soft" href="{url_for('beneficiaries_page')}">رجوع</a>
      </div>
    </form>
    """
def build_request_args_dict():
    return {
        "q": clean_csv_value(request.args.get("q", "")).strip(),
        "user_type": clean_csv_value(request.args.get("user_type", "")).strip(),
        "tawjihi_year": clean_csv_value(request.args.get("tawjihi_year", "")).strip(),
        "tawjihi_branch": clean_csv_value(request.args.get("tawjihi_branch", "")).strip(),
        "university_name": clean_csv_value(request.args.get("university_name", "")).strip(),
        "university_college": clean_csv_value(request.args.get("university_college", "")).strip(),
        "university_specialization": clean_csv_value(request.args.get("university_specialization", "")).strip(),
        "freelancer_specialization": clean_csv_value(request.args.get("freelancer_specialization", "")).strip(),
        "freelancer_company": clean_csv_value(request.args.get("freelancer_company", "")).strip(),
        "internet_method": clean_csv_value(request.args.get("internet_method", "")).strip(),
        "sort_by": clean_csv_value(request.args.get("sort_by", "id")).strip() or "id",
        "sort_order": clean_csv_value(request.args.get("sort_order", "desc")).strip() or "desc",
    }



def build_beneficiary_filters(args_dict):
    filters = ["1=1"]
    params = []
    q = args_dict["q"]
    if q:
        normalized_q = normalize_search_ar(q)
        filters.append("""
        (
            search_name ILIKE %s OR phone ILIKE %s OR
            COALESCE(tawjihi_year,'') ILIKE %s OR COALESCE(tawjihi_branch,'') ILIKE %s OR
            COALESCE(freelancer_specialization,'') ILIKE %s OR COALESCE(freelancer_company,'') ILIKE %s OR
            COALESCE(university_name,'') ILIKE %s OR COALESCE(university_college,'') ILIKE %s OR
            COALESCE(university_specialization,'') ILIKE %s
        )
        """)
        like_q = f"%{normalized_q}%"
        raw_like = f"%{q}%"
        params.extend([like_q, raw_like, raw_like, raw_like, raw_like, raw_like, raw_like, raw_like, raw_like])

    if args_dict["user_type"]:
        filters.append("user_type = %s")
        params.append(args_dict["user_type"])
    if args_dict["tawjihi_year"]:
        filters.append("tawjihi_year = %s")
        params.append(args_dict["tawjihi_year"])
    if args_dict["tawjihi_branch"]:
        filters.append("tawjihi_branch = %s")
        params.append(args_dict["tawjihi_branch"])
    if args_dict["university_name"]:
        filters.append("university_name = %s")
        params.append(args_dict["university_name"])
    if args_dict["university_college"]:
        filters.append("university_college ILIKE %s")
        params.append(f"%{args_dict['university_college']}%")
    if args_dict["university_specialization"]:
        filters.append("university_specialization ILIKE %s")
        params.append(f"%{args_dict['university_specialization']}%")
    if args_dict["freelancer_specialization"]:
        filters.append("freelancer_specialization ILIKE %s")
        params.append(f"%{args_dict['freelancer_specialization']}%")
    if args_dict["freelancer_company"]:
        filters.append("freelancer_company ILIKE %s")
        params.append(f"%{args_dict['freelancer_company']}%")

    if args_dict["internet_method"] == "cards":
        filters.append("""
        (
          user_type='tawjihi'
          OR (user_type='freelancer' AND freelancer_internet_method='نظام البطاقات')
          OR (user_type='university' AND university_internet_method='نظام البطاقات')
        )
        """)
    elif args_dict["internet_method"] == "username":
        filters.append("""
        (
          (user_type='freelancer' AND freelancer_internet_method='يمتلك اسم مستخدم')
          OR (user_type='university' AND university_internet_method='يمتلك اسم مستخدم')
        )
        """)

    return filters, params


def find_duplicate_phone(phone, exclude_id=None):
    phone = normalize_phone(phone)
    if not phone:
        return None
    sql = "SELECT id, full_name FROM beneficiaries WHERE phone=%s"
    params = [phone]
    if exclude_id is not None:
        sql += " AND id<>%s"
        params.append(exclude_id)
    sql += " LIMIT 1"
    return query_one(sql, params)


def build_beneficiary_row_html(r, selected_type, args_dict, page=1, display_index=None):
    usage_label, limited, count = get_usage_label(r)
    current_qs = build_query_string({**args_dict, "page": page})
    row_class = f"row-type-{safe(r.get('user_type'))}"
    if limited and count >= 3:
        row_class += " row-complete"
    modal_id = f"edit-{r['id']}"
    actions = []
    modal_html = ""
    if has_permission("edit"):
        edit_action = f"{url_for('edit_beneficiary_page', beneficiary_id=r['id'])}?current_user_type={selected_type}"
        modal_body = format_modal_fields(
            r,
            action=edit_action,
            scope_id=f"scope-{r['id']}",
            submit_label="حفظ التعديلات",
            show_type_selector=True,
        )
        edit_onsubmit = f"<form method=\"POST\" onsubmit=\"return submitBeneficiaryEdit(this, {r['id']}, '{modal_id}')\" action=\""
        modal_body = modal_body.replace('<form method="POST" action="', edit_onsubmit, 1)
        modal_html = f"""
        <div id="{modal_id}" class="modal">
          <div class="modal-card">
            <a href="#!" class="modal-close">×</a>
            <div class="hero" style="margin-bottom:14px"><h1>تعديل المستفيد #{r['id']}</h1><p>{safe(r.get('full_name'))}</p></div>
            {modal_body}
          </div>
        </div>
        """
        actions.append(f"<a class='btn btn-secondary btn-icon' href='#{modal_id}' title='تعديل'><i class='fa-solid fa-pen'></i></a>")
    if has_permission("delete"):
        delete_url = url_for('delete_beneficiary', beneficiary_id=r['id'])
        actions.append(
            f"<form class='inline-form' method='POST' action='{delete_url}' onsubmit=\"return confirm('هل أنت متأكد من الحذف؟')\"><button class='btn btn-danger btn-icon' type='submit' title='حذف'><i class='fa-solid fa-trash'></i></button></form>"
        )
    if has_permission("usage_counter") and limited and count < 3:
        usage_url = f"{url_for('add_usage', beneficiary_id=r['id'])}?{current_qs}"
        actions.append(
            f"<button class='btn btn-accent btn-icon' type='button' onclick=\"return incrementUsageAjax('{usage_url}', {r['id']}, '{modal_id}')\" title='+1 بطاقة'><i class='fa-solid fa-plus'></i></button>"
        )

    type_html = f"<span class='type-badge {get_type_css(r.get('user_type'))}'>{get_type_label(r.get('user_type'))}</span>"
    added_by = safe(r.get('added_by_username')) or '-'
    created_at = format_dt_short(r.get('created_at'))
    notes = safe(r.get('notes')) or '-'
    if selected_type == "tawjihi":
        row_cells = [safe(r['id']), safe(r['full_name']), safe(r['phone']), safe(r['tawjihi_year']), safe(r['tawjihi_branch']), usage_label, added_by, created_at, notes, " ".join(actions) if actions else "-"]
    elif selected_type == "university":
        row_cells = [safe(r['id']), safe(r['full_name']), safe(r['phone']), safe(r['university_name']), safe(r['university_college']), safe(r['university_specialization']), usage_label, added_by, created_at, notes, " ".join(actions) if actions else "-"]
    elif selected_type == "freelancer":
        row_cells = [safe(r['id']), safe(r['full_name']), safe(r['phone']), safe(r['freelancer_specialization']), safe(r['freelancer_company']), usage_label, added_by, created_at, notes, " ".join(actions) if actions else "-"]
    else:
        row_cells = [safe(r['id']), safe(r['full_name']), safe(r['phone']), type_html, usage_label, added_by, created_at, notes, " ".join(actions) if actions else "-"]
    row_html = f"<tr id='beneficiary-row-{r['id']}' class='{row_class}'>" + ''.join([f"<td class='cell-wrap'>{cell}</td>" for cell in row_cells]) + "</tr>"
    return row_html, modal_html


@app.route("/beneficiaries")
@login_required
@permission_required("view")
def beneficiaries_page():
    normalize_all_usage()

    args_dict = build_request_args_dict()

    try:
        page = max(1, int(request.args.get("page", "1") or "1"))
    except ValueError:
        page = 1

    per_page = 25

    filters, params = build_beneficiary_filters(args_dict)
    where = " AND ".join(filters)

    allowed_sort = {
        "id": "id",
        "full_name": "full_name",
        "phone": "phone",
        "tawjihi_year": "tawjihi_year",
        "tawjihi_branch": "tawjihi_branch",
        "university_name": "university_name",
        "university_specialization": "university_specialization",
        "freelancer_specialization": "freelancer_specialization",
        "freelancer_company": "freelancer_company",
        "weekly_usage_count": "weekly_usage_count",
        "created_at": "created_at",
    }
    sort_by = allowed_sort.get(args_dict["sort_by"], "id")
    sort_order = "ASC" if args_dict["sort_order"] == "asc" else "DESC"

    total = query_one(f"SELECT COUNT(*) AS c FROM beneficiaries WHERE {where}", params)["c"]
    pages = max(1, math.ceil(total / per_page))
    page = min(page, pages)
    offset = (page - 1) * per_page

    rows = query_all(
        f"""
        SELECT * FROM beneficiaries
        WHERE {where}
        ORDER BY {sort_by} {sort_order}, id DESC
        LIMIT %s OFFSET %s
        """,
        params + [per_page, offset],
    )

    tawjihi_years = distinct_values("tawjihi_year", "tawjihi")
    tawjihi_branches = distinct_values("tawjihi_branch", "tawjihi")
    university_names = distinct_values("university_name", "university")
    universities_colleges = distinct_values("university_college", "university")
    freelancer_companies = distinct_values("freelancer_company", "freelancer")
    freelancer_specs = distinct_values("freelancer_specialization", "freelancer")

    selected_type = args_dict["user_type"]
    tabs = [
        ("", "الكل", total if not selected_type else query_one("SELECT COUNT(*) AS c FROM beneficiaries", [])["c"]),
        ("tawjihi", "توجيهي", query_one("SELECT COUNT(*) AS c FROM beneficiaries WHERE user_type='tawjihi'")["c"]),
        ("university", "جامعة", query_one("SELECT COUNT(*) AS c FROM beneficiaries WHERE user_type='university'")["c"]),
        ("freelancer", "فري لانسر", query_one("SELECT COUNT(*) AS c FROM beneficiaries WHERE user_type='freelancer'")["c"]),
    ]
    tabs_html = ""
    for tab_value, label, count in tabs:
        tab_query = build_query_string({**args_dict, "user_type": tab_value, "page": 1})
        cls = "tab-pill active" if selected_type == tab_value else "tab-pill"
        tabs_html += f"<a class='{cls}' href='?{tab_query}'>{label} <span class='badge'>{count}</span></a>"

    metrics = {
        "النتائج الحالية": total,
        "المعروض في الصفحة": len(rows),
        "الخاضعون للحد الأسبوعي": sum(1 for r in rows if get_usage_label(r)[1]),
        "مكتملو الحد": sum(1 for r in rows if get_usage_label(r)[1] and get_usage_label(r)[2] >= 3),
    }
    metric_html = "".join([f"<div class='metric-box'><h4>{k}</h4><div class='num'>{v}</div></div>" for k, v in metrics.items()])

    if selected_type == "tawjihi":
        headers = [
            (None, "#"),
            ("full_name", "الاسم الكامل"),
            ("phone", "الجوال"),
            ("tawjihi_year", "السنة"),
            ("tawjihi_branch", "الفرع"),
            ("weekly_usage_count", "الاستخدام"),
            (None, "أضيف بواسطة"),
            ("created_at", "التاريخ"),
            (None, "ملاحظات"),
            (None, "إجراءات"),
        ]
    elif selected_type == "university":
        headers = [
            (None, "#"),
            ("full_name", "الاسم الكامل"),
            ("phone", "الجوال"),
            ("university_name", "الجامعة"),
            ("university_college", "الكلية"),
            ("university_specialization", "التخصص"),
            ("weekly_usage_count", "الاستخدام"),
            (None, "أضيف بواسطة"),
            ("created_at", "التاريخ"),
            (None, "ملاحظات"),
            (None, "إجراءات"),
        ]
    elif selected_type == "freelancer":
        headers = [
            (None, "#"),
            ("full_name", "الاسم الكامل"),
            ("phone", "الجوال"),
            ("freelancer_specialization", "التخصص"),
            ("freelancer_company", "الشركة"),
            ("weekly_usage_count", "الاستخدام"),
            (None, "أضيف بواسطة"),
            ("created_at", "التاريخ"),
            (None, "ملاحظات"),
            (None, "إجراءات"),
        ]
    else:
        headers = [
            (None, "#"),
            ("full_name", "الاسم الكامل"),
            ("phone", "الجوال"),
            (None, "النوع"),
            ("weekly_usage_count", "الاستخدام"),
            (None, "أضيف بواسطة"),
            ("created_at", "التاريخ"),
            (None, "ملاحظات"),
            (None, "إجراءات"),
        ]

    thead = "<tr>"
    for col, label in headers:
        if col:
            thead += f"<th><a href='{beneficiary_sort_link(args_dict, col)}'>{label}</a></th>"
        else:
            thead += f"<th>{label}</th>"
    can_bulk_select = has_permission('delete') or has_permission('export')
    if can_bulk_select:
        thead += "<th class='checkbox-cell'><input id='select-all' class='select-all' type='checkbox' onchange='toggleSelectAll(this)'></th>"
    thead += "</tr>"

    rows_html = ""
    modals_html = ""
    start_index = offset + 1
    for idx, r in enumerate(rows, start=start_index):
        row_html, modal_html = build_beneficiary_row_html(r, selected_type, args_dict, page=page, display_index=idx)
        rows_html += row_html
        modals_html += modal_html

    if not rows_html:
        rows_html = f"<tr><td colspan='{len(headers) + (1 if can_bulk_select else 0)}' class='empty-state'>لا توجد نتائج مطابقة لخيارات البحث الحالية.</td></tr>"

    pag_html = ""
    if pages > 1:
        pag_html = "<div class='pagination'>"
        for p in range(1, pages + 1):
            cls = "active" if p == page else ""
            page_query = build_query_string({**args_dict, "page": p})
            pag_html += f"<a class='{cls}' href='?{page_query}'>{p}</a>"
        pag_html += "</div>"

    years_options = "".join([f"<option value='{safe(y)}' {'selected' if args_dict['tawjihi_year']==y else ''}>{safe(y)}</option>" for y in tawjihi_years])
    branches_options = "".join([f"<option value='{safe(x)}' {'selected' if args_dict['tawjihi_branch']==x else ''}>{safe(x)}</option>" for x in tawjihi_branches])
    uni_options = "".join([f"<option value='{safe(x)}' {'selected' if args_dict['university_name']==x else ''}>{safe(x)}</option>" for x in university_names])
    college_options = "".join([f"<option value='{safe(x)}'>{safe(x)}</option>" for x in universities_colleges])
    free_company_options = "".join([f"<option value='{safe(x)}'>{safe(x)}</option>" for x in freelancer_companies])
    free_spec_options = "".join([f"<option value='{safe(x)}'>{safe(x)}</option>" for x in freelancer_specs])
    add_button_html = ""
    if has_permission('add'):
        add_button_html = "<a class='btn btn-secondary' href='#add-beneficiary-modal'><i class='fa-solid fa-user-plus'></i> إضافة مستفيد</a>"
    reset_cards_button_html = ""
    if has_permission('reset_weekly_usage'):
        reset_url = url_for('reset_weekly_usage')
        reset_cards_button_html = f"<button class='btn btn-outline' type='button' onclick=\"return resetWeeklyUsageAjax('{reset_url}')\"><i class='fa-solid fa-rotate'></i> تجديد كل البطاقات</button>"

    content = f"""
    <div class="hero">
      <h1>إدارة المستفيدين المتقدمة</h1>
      <p>تبويبات حسب النوع، بحث شامل، فلاتر متخصصة، ترتيب أعمدة، وتعديل منبثق بدون الخروج من الصفحة.</p>
    </div>

    <div class="tabs">{tabs_html}</div>

    <div class="card">
      <div class="filter-box">
        <form method="GET">
          <div class="row">
            <div><label>بحث عام</label><input name="q" value="{safe(args_dict['q'])}" placeholder="اسم، جوال، جامعة، تخصص، شركة، سنة..."></div>
            <div>
              <label>النوع</label>
              <select name="user_type">
                <option value="">الكل</option>
                <option value="tawjihi" {"selected" if selected_type=="tawjihi" else ""}>توجيهي</option>
                <option value="university" {"selected" if selected_type=="university" else ""}>جامعة</option>
                <option value="freelancer" {"selected" if selected_type=="freelancer" else ""}>فري لانسر</option>
              </select>
            </div>
            <div>
              <label>الترتيب</label>
              <select name="sort_by">
                <option value="id" {"selected" if args_dict['sort_by']=="id" else ""}>ID</option>
                <option value="full_name" {"selected" if args_dict['sort_by']=="full_name" else ""}>الاسم</option>
                <option value="phone" {"selected" if args_dict['sort_by']=="phone" else ""}>الجوال</option>
                <option value="weekly_usage_count" {"selected" if args_dict['sort_by']=="weekly_usage_count" else ""}>الاستخدام</option>
                <option value="created_at" {"selected" if args_dict['sort_by']=="created_at" else ""}>تاريخ الإضافة</option>
              </select>
            </div>
            <div>
              <label>اتجاه الترتيب</label>
              <select name="sort_order">
                <option value="desc" {"selected" if args_dict['sort_order']=="desc" else ""}>تنازلي</option>
                <option value="asc" {"selected" if args_dict['sort_order']=="asc" else ""}>تصاعدي</option>
              </select>
            </div>
          </div>

          <details class="advanced-filters" {"open" if any([args_dict['tawjihi_year'], args_dict['tawjihi_branch'], args_dict['university_name'], args_dict['university_college'], args_dict['university_specialization'], args_dict['freelancer_specialization'], args_dict['freelancer_company'], args_dict['internet_method']]) else ""}>
            <summary><i class="fa-solid fa-sliders"></i> فلاتر متقدمة</summary>
            <div class="row">
              <div>
                <label>سنة التوجيهي</label>
                <select name="tawjihi_year"><option value="">الكل</option>{years_options}</select>
              </div>
              <div>
                <label>فرع التوجيهي</label>
                <select name="tawjihi_branch"><option value="">الكل</option>{branches_options}</select>
              </div>
              <div>
                <label>الجامعة</label>
                <select name="university_name"><option value="">الكل</option>{uni_options}</select>
              </div>
              <div><label>الكلية</label><input name="university_college" list="colleges-list" value="{safe(args_dict['university_college'])}" placeholder="ابحث حسب الكلية"></div>
              <div><label>التخصص الجامعي</label><input name="university_specialization" value="{safe(args_dict['university_specialization'])}" placeholder="مثال: هندسة برمجيات"></div>
              <div><label>تخصص الفري لانسر</label><input name="freelancer_specialization" list="freelancer-specs" value="{safe(args_dict['freelancer_specialization'])}" placeholder="مثال: تصميم جرافيك"></div>
              <div><label>شركة الفري لانسر</label><input name="freelancer_company" list="freelancer-companies" value="{safe(args_dict['freelancer_company'])}" placeholder="مثال: Upwork"></div>
              <div>
                <label>طريقة الإنترنت</label>
                <select name="internet_method">
                  <option value="">الكل</option>
                  <option value="cards" {"selected" if args_dict['internet_method']=="cards" else ""}>نظام البطاقات / محدود</option>
                  <option value="username" {"selected" if args_dict['internet_method']=="username" else ""}>يمتلك اسم مستخدم</option>
                </select>
              </div>
            </div>
          </details>

          <datalist id="colleges-list">{college_options}</datalist>
          <datalist id="freelancer-companies">{free_company_options}</datalist>
          <datalist id="freelancer-specs">{free_spec_options}</datalist>

          <div class="actions" style="margin-top:14px">
            <button class="btn btn-primary" type="submit"><i class="fa-solid fa-magnifying-glass"></i> تطبيق البحث</button>
            <a class="btn btn-soft" href="{url_for('beneficiaries_page')}"><i class="fa-solid fa-rotate-left"></i> إعادة ضبط</a>
            {add_button_html}
            {reset_cards_button_html}
          </div>
        </form>
      </div>

      <div class="metric-grid" style="margin-top:16px">{metric_html}</div>
    </div>

    <div class="card" style="margin-top:14px">
      <div class='bulk-toolbar'>
        <span class='selected-count'>المحدد: <strong id='selected-count'>0</strong></span>
        {"<button class='btn btn-soft btn-icon' type='button' onclick='return submitBulkExport()' title='تصدير المحدد'><i class='fa-solid fa-file-export'></i></button>" if has_permission('export') else ""}
        {"<button class='btn btn-danger btn-icon' type='button' onclick='return submitBulkDelete()' title='حذف المحدد'><i class='fa-solid fa-trash'></i></button>" if has_permission('delete') else ""}
      </div>
      <form id='bulk-delete-form' method='POST' action='{url_for('bulk_delete_beneficiaries')}' style='display:none'><input type='hidden' name='ids' value=''></form>
      <form id='bulk-export-form' method='POST' action='{url_for('export_selected_beneficiaries')}' style='display:none'><input type='hidden' name='ids' value=''></form>
      <div class="table-wrap">
        <table>
          <thead>{thead}</thead>
          <tbody>{rows_html}</tbody>
        </table>
      </div>
      {pag_html}
    </div>

    {modals_html}
    {build_add_beneficiary_modal(selected_type or 'tawjihi')}
    """
    return render_page("المستفيدون", content)


def beneficiary_form_html(data=None, action="", title=""):
    selected_type = clean_csv_value((data or {}).get("user_type") or request.args.get("user_type", "tawjihi")) or "tawjihi"
    links = "".join([
        f"<a class='menu-card' href='{url_for('add_beneficiary_page')}?user_type=tawjihi'><div class='menu-icon'><i class='fa-solid fa-user-graduate'></i></div><h3>توجيهي</h3><p>سنة التوجيهي والفرع.</p></a>",
        f"<a class='menu-card' href='{url_for('add_beneficiary_page')}?user_type=university'><div class='menu-icon'><i class='fa-solid fa-building-columns'></i></div><h3>جامعة</h3><p>جامعة، كلية، تخصص.</p></a>",
        f"<a class='menu-card' href='{url_for('add_beneficiary_page')}?user_type=freelancer'><div class='menu-icon'><i class='fa-solid fa-laptop-code'></i></div><h3>فري لانسر</h3><p>شركة، تخصص، إنترنت.</p></a>",
    ])
    form = format_modal_fields(
        {**(data or {}), "user_type": selected_type},
        action=action,
        scope_id="standalone-beneficiary-form",
        submit_label="حفظ",
        show_type_selector=False,
        fixed_user_type=selected_type,
    )
    return f"""
    <div class="hero"><h1>{title}</h1><p>أدخل البيانات الخاصة بالمستفيد في هذه الصفحة المخصصة.</p></div>
    <div class="card">{form}</div>
    """


def collect_beneficiary_form():
    data = {}
    for col in CSV_IMPORT_COLUMNS:
        val = clean_csv_value(request.form.get(col, ""))
        if col == "phone":
            val = normalize_phone(val)
        data[col] = val
    data["full_name"] = full_name_from_parts(data["first_name"], data["second_name"], data["third_name"], data["fourth_name"])
    data["search_name"] = normalize_search_ar(data["full_name"])
    data["weekly_usage_week_start"] = get_week_start()
    return data


@app.route("/beneficiaries/add", methods=["GET", "POST"])
@login_required
@permission_required("add")
def add_beneficiary_page():
    if request.method == "POST":
        data = collect_beneficiary_form()
        duplicate = find_duplicate_phone(data.get("phone"))
        if duplicate:
            flash(f"رقم الجوال مستخدم مسبقًا للمستفيد: {safe(duplicate.get('full_name'))}.", "error")
            return render_page("إضافة مستفيد", beneficiary_form_html(data, action=url_for("add_beneficiary_page"), title=page_title_map.get(data.get('user_type') or clean_csv_value(request.args.get('user_type', 'tawjihi')) or 'tawjihi', "إضافة مستفيد")))
        data["added_by_account_id"] = session.get("account_id")
        data["added_by_username"] = session.get("username")
        try:
            row = execute_sql("""
                INSERT INTO beneficiaries (
                    user_type, first_name, second_name, third_name, fourth_name, full_name, search_name, phone,
                    tawjihi_year, tawjihi_branch, freelancer_specialization, freelancer_company,
                    freelancer_schedule_type, freelancer_internet_method, freelancer_time_mode,
                    freelancer_time_from, freelancer_time_to, university_name, university_college,
                    university_specialization, university_days, university_internet_method,
                    university_time_mode, university_time_from, university_time_to,
                    weekly_usage_count, weekly_usage_week_start, notes, added_by_account_id, added_by_username
                ) VALUES (
                    %(user_type)s, %(first_name)s, %(second_name)s, %(third_name)s, %(fourth_name)s, %(full_name)s, %(search_name)s, %(phone)s,
                    %(tawjihi_year)s, %(tawjihi_branch)s, %(freelancer_specialization)s, %(freelancer_company)s,
                    %(freelancer_schedule_type)s, %(freelancer_internet_method)s, %(freelancer_time_mode)s,
                    %(freelancer_time_from)s, %(freelancer_time_to)s, %(university_name)s, %(university_college)s,
                    %(university_specialization)s, %(university_days)s, %(university_internet_method)s,
                    %(university_time_mode)s, %(university_time_from)s, %(university_time_to)s,
                    0, %(weekly_usage_week_start)s, %(notes)s, %(added_by_account_id)s, %(added_by_username)s
                ) RETURNING id
            """, data, fetchone=True)
        except psycopg2.Error:
            flash("رقم الجوال مستخدم مسبقًا للمستفيد.", "error")
            return redirect(url_for("beneficiaries_page", user_type=data["user_type"]))
        new_id = row['id'] if row else None
        log_action("add", "beneficiary", new_id, f"إضافة مستفيد: {data['full_name']}")
        flash("تمت إضافة المستفيد.", "success")
        return redirect(url_for("beneficiaries_page", user_type=data["user_type"]))
    page_user_type = clean_csv_value(request.args.get("user_type", "tawjihi")) or "tawjihi"
    page_title_map = {
        "tawjihi": "إضافة طالب توجيهي",
        "university": "إضافة طالب جامعي",
        "freelancer": "إضافة فري لانسر",
    }
    return render_page("إضافة مستفيد", beneficiary_form_html(action=url_for("add_beneficiary_page"), title=page_title_map.get(page_user_type, "إضافة مستفيد")))


@app.route("/beneficiaries/edit/<int:beneficiary_id>", methods=["GET", "POST"])
@login_required
@permission_required("edit")
def edit_beneficiary_page(beneficiary_id):
    row = query_one("SELECT * FROM beneficiaries WHERE id=%s", [beneficiary_id])
    if not row:
        flash("المستفيد غير موجود.", "error")
        return redirect(url_for("beneficiaries_page"))
    if request.method == "POST":
        data = collect_beneficiary_form()
        data["id"] = beneficiary_id
        duplicate = find_duplicate_phone(data.get("phone"), exclude_id=beneficiary_id)
        if duplicate:
            message = f"رقم الجوال مستخدم مسبقًا للمستفيد: {safe(duplicate.get('full_name'))}."
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return jsonify({"ok": False, "message": message, "category": "error"}), 400
            flash(message, "error")
            return redirect(url_for("beneficiaries_page", user_type=row.get("user_type")))
        execute_sql("""
            UPDATE beneficiaries SET
                user_type=%(user_type)s,
                first_name=%(first_name)s,
                second_name=%(second_name)s,
                third_name=%(third_name)s,
                fourth_name=%(fourth_name)s,
                full_name=%(full_name)s,
                search_name=%(search_name)s,
                phone=%(phone)s,
                tawjihi_year=%(tawjihi_year)s,
                tawjihi_branch=%(tawjihi_branch)s,
                freelancer_specialization=%(freelancer_specialization)s,
                freelancer_company=%(freelancer_company)s,
                freelancer_schedule_type=%(freelancer_schedule_type)s,
                freelancer_internet_method=%(freelancer_internet_method)s,
                freelancer_time_mode=%(freelancer_time_mode)s,
                freelancer_time_from=%(freelancer_time_from)s,
                freelancer_time_to=%(freelancer_time_to)s,
                university_name=%(university_name)s,
                university_college=%(university_college)s,
                university_specialization=%(university_specialization)s,
                university_days=%(university_days)s,
                university_internet_method=%(university_internet_method)s,
                university_time_mode=%(university_time_mode)s,
                university_time_from=%(university_time_from)s,
                university_time_to=%(university_time_to)s,
                notes=%(notes)s
            WHERE id=%(id)s
        """, data)
        log_action("edit", "beneficiary", beneficiary_id, f"تعديل مستفيد: {data['full_name']}")
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            updated_row = query_one("SELECT * FROM beneficiaries WHERE id=%s", [beneficiary_id])
            current_type = clean_csv_value(request.args.get("current_user_type", ""))
            row_html, modal_html = build_beneficiary_row_html(updated_row, current_type or updated_row.get("user_type") if updated_row else current_type, build_request_args_dict(), page=1)
            return jsonify({"ok": True, "row_html": row_html, "modal_html": modal_html, "message": "تم حفظ التعديلات.", "category": "success"})
        flash("تم حفظ التعديلات.", "success")
        return redirect(url_for("beneficiaries_page", user_type=data["user_type"]))
    return render_page("تعديل مستفيد", beneficiary_form_html(row, url_for("edit_beneficiary_page", beneficiary_id=beneficiary_id), "تعديل مستفيد"))


@app.route("/beneficiaries/delete/<int:beneficiary_id>", methods=["POST"])
@login_required
@permission_required("delete")
def delete_beneficiary(beneficiary_id):
    row = query_one("SELECT full_name, user_type FROM beneficiaries WHERE id=%s", [beneficiary_id])
    execute_sql("DELETE FROM beneficiaries WHERE id=%s", [beneficiary_id])
    log_action("delete", "beneficiary", beneficiary_id, f"حذف مستفيد: {safe(row['full_name']) if row else ''}")
    flash("تم حذف المستفيد.", "success")
    redirect_type = row["user_type"] if row else ""
    return redirect(url_for("beneficiaries_page", user_type=redirect_type))


@app.route("/beneficiaries/add_usage/<int:beneficiary_id>", methods=["POST"])
@login_required
@permission_required("usage_counter")
def add_usage(beneficiary_id):
    normalize_beneficiary_usage(beneficiary_id)
    row = query_one("SELECT * FROM beneficiaries WHERE id=%s", [beneficiary_id])
    if not row:
        flash("المستفيد غير موجود.", "error")
        return redirect(url_for("beneficiaries_page"))
    usage_label, limited, count = get_usage_label(row)
    message = ""
    category = "success"
    if limited:
        if count >= 3:
            message = "اكتمل الحد الأسبوعي."
            category = "error"
            flash(message, category)
        else:
            execute_sql("UPDATE beneficiaries SET weekly_usage_count=COALESCE(weekly_usage_count,0)+1 WHERE id=%s", [beneficiary_id])
            log_action("usage_counter", "beneficiary", beneficiary_id, f"+1 بطاقة لـ {safe(row['full_name'])}")
            message = "تمت إضافة استخدام."
            flash(message, "success")
    else:
        message = "هذا المستفيد غير خاضع لعداد 3 مرات."
        category = "error"
        flash(message, category)
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        updated_row = query_one("SELECT * FROM beneficiaries WHERE id=%s", [beneficiary_id])
        args_dict = build_request_args_dict()
        row_html, modal_html = build_beneficiary_row_html(updated_row, args_dict.get("user_type", ""), args_dict, page=max(1, int(request.args.get("page", "1") or "1")))
        return jsonify({"ok": True, "row_html": row_html, "modal_html": modal_html, "message": message, "category": category})
    return redirect(request.referrer or url_for("beneficiaries_page"))


@app.route("/beneficiaries/reset-weekly-usage", methods=["POST"])
@login_required
@permission_required("reset_weekly_usage")
def reset_weekly_usage():
    reset_start = get_week_start(date.today() + timedelta(days=1))
    execute_sql("""
        UPDATE beneficiaries
        SET weekly_usage_count = 0, weekly_usage_week_start = %s
        WHERE weekly_usage_count IS DISTINCT FROM 0 OR weekly_usage_week_start IS DISTINCT FROM %s
    """, [reset_start, reset_start])
    log_action("reset_weekly_usage", "beneficiary", None, f"تصفير يدوي لكل البطاقات - week_start={reset_start}")
    flash("تم تجديد جميع البطاقات يدويًا بنجاح.", "success")
    return redirect(request.referrer or url_for("beneficiaries_page"))


def find_existing_beneficiary(cur, data):
    if data.get("phone"):
        cur.execute("SELECT id FROM beneficiaries WHERE phone=%s LIMIT 1", [data["phone"]])
        row = cur.fetchone()
        if row:
            return row[0]
    cur.execute("SELECT id FROM beneficiaries WHERE full_name=%s AND user_type=%s LIMIT 1", [data["full_name"], data["user_type"]])
    row = cur.fetchone()
    return row[0] if row else None


@app.route("/import")
@login_required
@permission_required("import")
def import_page():
    last_task_id = session.get("last_import_task_id", "")
    last_task = get_import_task(last_task_id) if last_task_id else None
    status_box = ""
    if last_task:
        status_box = f"""
        <div class="card" style="margin-top:14px">
          <h3 style="margin-top:0">آخر مهمة استيراد</h3>
          <div class="info-note">
            <div><strong>الملف:</strong> {safe(last_task.get('filename'))}</div>
            <div><strong>الحالة:</strong> {safe(last_task.get('status'))}</div>
            <div><strong>الرسالة:</strong> {safe(last_task.get('message'))}</div>
          </div>
          <div class="actions" style="margin-top:12px">
            <a class="btn btn-secondary" href="{url_for('import_status_page', task_id=last_task['id'])}"><i class="fa-solid fa-chart-line"></i> فتح شاشة المتابعة</a>
          </div>
        </div>
        """

    content = f"""
    <div class="hero"><h1>استيراد CSV احترافي</h1><p>رفع الملف وتشغيل المعالجة في الخلفية مع شاشة تقدم حية وسجل خطوات مباشر.</p></div>
    <div class="card">
      <form method="POST" action="{url_for('import_csv')}" enctype="multipart/form-data">
        <div class="row">
          <div>
            <label>ملف CSV</label>
            <input type="file" name="csv_file" accept=".csv" required>
            <div class="small" style="margin-top:8px">بعد الضغط على البدء سيتم إنشاء مهمة استيراد مستقلة، ويمكنك متابعة التقدم وسير العمليات مباشرة.</div>
          </div>
        </div>
        <div class="actions" style="margin-top:14px">
          <button class="btn btn-primary" type="submit"><i class="fa-solid fa-play"></i> بدء الاستيراد الاحترافي</button>
          <a class="btn btn-outline" href="{url_for('download_template')}">تنزيل القالب</a>
        </div>
      </form>
    </div>

    <div class="card" style="margin-top:14px">
      <h3 style="margin-top:0">كيف تعمل المعالجة؟</h3>
      <div class="info-note">
        <div>1) يتم قراءة الملف وتحليله أولًا.</div>
        <div>2) يتم تحديد السجلات الجديدة والسجلات التي تحتاج تحديثًا.</div>
        <div>3) تتم المعالجة في الخلفية على دفعات سريعة بدل سجل بسجل.</div>
        <div>4) يمكنك مشاهدة النسبة، عدد المعالجات، وعدد الأخطاء وسجل التنفيذ لحظة بلحظة.</div>
      </div>
    </div>

    {status_box}
    """
    return render_page("استيراد CSV", content)


@app.route("/import_status/<task_id>")
@login_required
@permission_required("import")
def import_status_page(task_id):
    task = get_import_task(task_id)
    if not task:
        flash("مهمة الاستيراد غير موجودة.", "error")
        return redirect(url_for("import_page"))
    content = f"""
    <div class="hero"><h1>متابعة الاستيراد المباشرة</h1><p>شاشة حية تعرض التقدم الفعلي وسجل العمليات أثناء المعالجة.</p></div>

    <div class="card">
      <div class="metric-grid">
        <div class="metric-box"><h4>الحالة</h4><div class="num" id="task-status">{safe(task.get('status'))}</div></div>
        <div class="metric-box"><h4>إجمالي السجلات</h4><div class="num" id="task-total">{task.get('total',0)}</div></div>
        <div class="metric-box"><h4>تمت معالجته</h4><div class="num" id="task-processed">{task.get('processed',0)}</div></div>
        <div class="metric-box"><h4>تمت إضافته</h4><div class="num" id="task-inserted">{task.get('inserted',0)}</div></div>
        <div class="metric-box"><h4>تم تحديثه</h4><div class="num" id="task-updated">{task.get('updated',0)}</div></div>
        <div class="metric-box"><h4>الأخطاء</h4><div class="num" id="task-errors">{task.get('error_count',0)}</div></div>
      </div>

      <div style="margin-top:18px">
        <div class="small" style="margin-bottom:6px">نسبة التقدم: <strong id="task-percent">{task.get('percent',0)}%</strong></div>
        <div class="bar-track" style="height:18px"><div id="task-bar" class="bar-fill" style="width:{task.get('percent',0)}%"></div></div>
      </div>

      <div class="info-note" style="margin-top:14px">
        <div><strong>الملف:</strong> {safe(task.get('filename'))}</div>
        <div><strong>الخطوة الحالية:</strong> <span id="task-step">{safe(task.get('current_step'))}</span></div>
        <div><strong>الرسالة:</strong> <span id="task-message">{safe(task.get('message'))}</span></div>
        <div><strong>بدأت:</strong> <span id="task-started">{safe(task.get('started_at'))}</span></div>
        <div><strong>انتهت:</strong> <span id="task-finished">{safe(task.get('finished_at'))}</span></div>
      </div>

      <div class="actions" style="margin-top:14px">
        <a class="btn btn-soft" href="{url_for('import_page')}">الرجوع لصفحة الاستيراد</a>
        <a class="btn btn-secondary" href="{url_for('beneficiaries_page')}">فتح المستفيدين</a>
      </div>
    </div>

    <div class="card" style="margin-top:14px">
      <h3 style="margin-top:0">سير العمليات</h3>
      <div id="task-logs" style="background:#0f172a;color:#e2e8f0;border-radius:16px;padding:14px;max-height:420px;overflow:auto;font-family:Consolas,monospace;font-size:13px;line-height:1.8">
        {''.join(f'<div>{safe(log)}</div>' for log in task.get('logs', [])) or '<div>بانتظار بدء المعالجة...</div>'}
      </div>
    </div>

    <script>
    const taskId = {task_id!r};
    let pollTimer = null;

    function setText(id, value) {{
      const el = document.getElementById(id);
      if (el) el.textContent = value ?? '';
    }}

    function renderLogs(logs) {{
      const box = document.getElementById('task-logs');
      if (!box) return;
      box.innerHTML = (logs && logs.length) ? logs.map(x => `<div>${{x.replace(/</g,'&lt;').replace(/>/g,'&gt;')}}</div>`).join('') : '<div>لا توجد رسائل حتى الآن.</div>';
      box.scrollTop = box.scrollHeight;
    }}

    function refreshTask() {{
      fetch(`/import_progress/${{taskId}}`)
        .then(r => r.json())
        .then(data => {{
          setText('task-status', data.status || '');
          setText('task-total', data.total || 0);
          setText('task-processed', data.processed || 0);
          setText('task-inserted', data.inserted || 0);
          setText('task-updated', data.updated || 0);
          setText('task-errors', data.error_count || 0);
          setText('task-percent', `${{data.percent || 0}}%`);
          setText('task-step', data.current_step || '');
          setText('task-message', data.message || '');
          setText('task-started', data.started_at || '');
          setText('task-finished', data.finished_at || '');
          const bar = document.getElementById('task-bar');
          if (bar) bar.style.width = `${{data.percent || 0}}%`;
          renderLogs(data.logs || []);
          if (['completed', 'failed'].includes(data.status)) {{
            clearInterval(pollTimer);
          }}
        }})
        .catch(() => {{}});
    }}

    refreshTask();
    pollTimer = setInterval(refreshTask, 1200);
    </script>
    """
    return render_page("متابعة الاستيراد", content)


@app.route("/import_progress/<task_id>")
@login_required
@permission_required("import")
def import_progress(task_id):
    task = get_import_task(task_id)
    if not task:
        return jsonify({"error": "not_found"}), 404
    return jsonify(task)


@app.route("/download_template")
@login_required
@permission_required("import")
def download_template():
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=CSV_IMPORT_COLUMNS)
    writer.writeheader()
    resp = Response(output.getvalue(), mimetype="text/csv")
    resp.headers["Content-Disposition"] = "attachment; filename=users_import_template.csv"
    return resp


@app.route("/import_csv", methods=["POST"])
@login_required
@permission_required("import")
def import_csv():
    file = request.files.get("csv_file")
    if not file or not file.filename:
        flash("اختر ملف CSV أولًا.", "error")
        return redirect(url_for("import_page"))
    try:
        content = file.read().decode("utf-8-sig")
    except UnicodeDecodeError:
        flash("الملف ليس بترميز UTF-8.", "error")
        return redirect(url_for("import_page"))

    reader = csv.DictReader(io.StringIO(content))
    if not reader.fieldnames:
        flash("الملف فارغ أو غير صالح.", "error")
        return redirect(url_for("import_page"))

    task_id = create_import_task(session.get("username", ""), session.get("account_id"), file.filename)
    session["last_import_task_id"] = task_id
    append_import_log(task_id, f"تم استلام الملف: {file.filename}")
    launch_import_task(task_id, content)
    flash("تم بدء الاستيراد في الخلفية. يمكنك متابعة التقدم المباشر الآن.", "success")
    return redirect(url_for("import_status_page", task_id=task_id))


@app.route("/exports")
@login_required
@permission_required("export")
def export_center():
    universities = distinct_values("university_name", "university")
    uni_options = "".join([f"<option value='{safe(x)}'>{safe(x)}</option>" for x in universities])
    content = f"""
    <div class="hero"><h1>مركز التصدير الاحترافي</h1><p>اختر ما تريد تصديره بدقة بدل تصدير كل المستفيدين مرة واحدة.</p></div>
    <div class="grid-3">
      <a class="menu-card" href="{url_for('export_csv')}?user_type=tawjihi">
        <div class="menu-icon"><i class="fa-solid fa-user-graduate"></i></div><h3>تصدير التوجيهي</h3><p>كل طلاب التوجيهي فقط.</p>
      </a>
      <a class="menu-card" href="{url_for('export_csv')}?user_type=freelancer">
        <div class="menu-icon"><i class="fa-solid fa-laptop-code"></i></div><h3>تصدير الفري لانسر</h3><p>كل الفري لانسر فقط.</p>
      </a>
      <a class="menu-card" href="{url_for('export_csv')}?user_type=university">
        <div class="menu-icon"><i class="fa-solid fa-building-columns"></i></div><h3>تصدير الجامعات</h3><p>كل الطلاب الجامعيين فقط.</p>
      </a>
    </div>
    <div class="card" style="margin-top:16px">
      <h3 style="margin-top:0">تصدير مخصص حسب الجامعة</h3>
      <form method="GET" action="{url_for('export_csv')}">
        <div class="row">
          <div>
            <label>الجامعة</label>
            <select name="university_name" required>
              <option value="">اختر الجامعة</option>{uni_options}
            </select>
          </div>
          <div>
            <label>النوع</label>
            <input value="جامعة" disabled>
            <input type="hidden" name="user_type" value="university">
          </div>
        </div>
        <div class="actions" style="margin-top:14px">
          <button class="btn btn-primary" type="submit"><i class="fa-solid fa-file-excel"></i> تصدير الجامعة المحددة</button>
        </div>
      </form>
    </div>
    <div class="card" style="margin-top:16px">
      <h3 style="margin-top:0">تصدير ذكي حسب الفلاتر الحالية</h3>
      <p class="small">من صفحة المستفيدين، أي بحث أو فلترة تطبقها ستنتقل تلقائيًا إلى ملف Excel عند الضغط على زر التصدير.</p>
      <div class="actions">
        <a class="btn btn-secondary" href="{url_for('beneficiaries_page')}"><i class="fa-solid fa-filter"></i> افتح صفحة المستفيدين وحدد الفلاتر</a>
      </div>
    </div>
    """
    return render_page("مركز التصدير", content)


@app.route("/export_csv")
@login_required
@permission_required("export")
def export_csv():
    args_dict = build_request_args_dict()
    filters, params = build_beneficiary_filters(args_dict)
    where = " AND ".join(filters)
    rows = query_all(f"SELECT * FROM beneficiaries WHERE {where} ORDER BY id DESC", params)

    export_columns = [
        (None, "#"),
        ("user_type", "نوع المستفيد"),
        ("first_name", "الاسم الأول"),
        ("second_name", "الاسم الثاني"),
        ("third_name", "الاسم الثالث"),
        ("fourth_name", "الاسم الرابع"),
        ("full_name", "الاسم الكامل"),
        ("search_name", "اسم البحث"),
        ("phone", "رقم الجوال"),
        ("tawjihi_year", "سنة التوجيهي"),
        ("tawjihi_branch", "فرع التوجيهي"),
        ("freelancer_specialization", "تخصص الفري لانسر"),
        ("freelancer_company", "شركة الفري لانسر"),
        ("freelancer_schedule_type", "نوع دوام الفري لانسر"),
        ("freelancer_internet_method", "طريقة إنترنت الفري لانسر"),
        ("freelancer_time_mode", "وضع وقت الفري لانسر"),
        ("freelancer_time_from", "وقت الفري لانسر من"),
        ("freelancer_time_to", "وقت الفري لانسر إلى"),
        ("university_name", "الجامعة"),
        ("university_college", "الكلية"),
        ("university_specialization", "التخصص الجامعي"),
        ("university_days", "أيام الجامعة"),
        ("university_internet_method", "طريقة إنترنت الجامعة"),
        ("university_time_mode", "وضع وقت الجامعة"),
        ("university_time_from", "وقت الجامعة من"),
        ("university_time_to", "وقت الجامعة إلى"),
        ("weekly_usage_count", "عدد الاستخدام الأسبوعي"),
        ("weekly_usage_week_start", "بداية الأسبوع"),
        ("created_at", "تاريخ الإنشاء"),
    ]

    def excel_value(value):
        if value is None:
            return ""
        return str(value)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "المستفيدون"

    title_fill = PatternFill(fill_type="solid", fgColor="123B6D")
    title_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="EAF3FB")
    header_font = Font(bold=True, color="123B6D")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top_alignment = Alignment(vertical="top", wrap_text=True)
    thin_side = Side(style="thin", color="DCE8F3")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # report title
    report_title = "تقرير المستفيدين المفلتر - Hobe Hub" if any(v for k, v in args_dict.items() if k not in {"sort_by", "sort_order"}) else "تقرير المستفيدين - Hobe Hub"
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(export_columns))
    title_cell = sheet.cell(row=1, column=1, value=report_title)
    title_cell.fill = title_fill
    title_cell.font = Font(color="FFFFFF", bold=True, size=14)
    title_cell.alignment = center_alignment
    sheet.row_dimensions[1].height = 24

    # headers
    for col_idx, (_, header_label) in enumerate(export_columns, start=1):
        cell = sheet.cell(row=2, column=col_idx, value=header_label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border

    # data rows
    for row_idx, row in enumerate(rows, start=3):
        row_dict = dict(row)
        for col_idx, (field_name, _) in enumerate(export_columns, start=1):
            value = row_dict.get(field_name, "")
            cell = sheet.cell(row=row_idx, column=col_idx, value=excel_value(value))
            cell.border = thin_border
            if field_name in {"id", "weekly_usage_count"}:
                cell.alignment = center_alignment
            else:
                cell.alignment = top_alignment

    # freeze, filter, widths
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:{get_column_letter(len(export_columns))}{max(sheet.max_row, 2)}"

    width_overrides = {
        "A": 10, "B": 16, "C": 18, "D": 18, "E": 18, "F": 18, "G": 28, "H": 24, "I": 18,
        "J": 16, "K": 16, "L": 24, "M": 24, "N": 22, "O": 22, "P": 20, "Q": 16, "R": 16,
        "S": 24, "T": 24, "U": 24, "V": 18, "W": 22, "X": 20, "Y": 16, "Z": 16, "AA": 16,
        "AB": 18, "AC": 22
    }
    for col_idx in range(1, len(export_columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row_idx in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_length:
                    max_length = cell_len
            except Exception:
                pass
        sheet.column_dimensions[col_letter].width = width_overrides.get(col_letter, min(max(max_length + 2, 12), 35))

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    log_action("export", "beneficiary", None, "تصدير Excel XLSX")
    resp = Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp.headers["Content-Disposition"] = "attachment; filename=beneficiaries_export.xlsx"
    return resp


@app.route("/backup_sql")
@login_required
@permission_required("backup")
def backup_sql():
    rows = query_all("SELECT * FROM beneficiaries ORDER BY id")
    lines = []
    for r in rows:
        cols = []
        vals = []
        for k, v in r.items():
            cols.append(k)
            if v is None:
                vals.append("NULL")
            else:
                sval = str(v).replace("'", "''")
                vals.append(f"'{sval}'")
        lines.append(f"INSERT INTO beneficiaries ({', '.join(cols)}) VALUES ({', '.join(vals)});")
    data = "\n".join(lines)
    log_action("backup", "beneficiary", None, "Backup SQL")
    resp = Response(data, mimetype="application/sql")
    resp.headers["Content-Disposition"] = "attachment; filename=beneficiaries_backup.sql"
    return resp


@app.route("/accounts")
@login_required
@permission_required("manage_accounts")
def accounts_page():
    rows = query_all("""
        SELECT a.*,
               COALESCE(string_agg(p.name, ', ' ORDER BY p.name), '') AS perms
        FROM app_accounts a
        LEFT JOIN account_permissions ap ON ap.account_id = a.id
        LEFT JOIN permissions p ON p.id = ap.permission_id
        GROUP BY a.id
        ORDER BY a.id DESC
    """)
    html = """
    <div class="hero"><h1>إدارة المستخدمين</h1><p>إضافة حسابات، تفعيل وتعطيل، وتخصيص الصلاحيات.</p></div>
    <div class="actions" style="margin-bottom:14px"><a class="btn btn-primary" href="/accounts/add">إضافة مستخدم</a></div>
    <div class="card"><table><thead><tr><th>ID</th><th>اسم المستخدم</th><th>الاسم الكامل</th><th>الحالة</th><th>الصلاحيات</th><th>إجراءات</th></tr></thead><tbody>
    """
    for r in rows:
        perms_text = "، ".join(permission_label(p.strip()) for p in safe(r['perms']).split(',') if p.strip()) or "-"
        html += f"""
        <tr>
          <td>{r['id']}</td>
          <td>{safe(r['username'])}</td>
          <td>{safe(r['full_name'])}</td>
          <td>{"مفعل" if r['is_active'] else "معطل"}</td>
          <td style="max-width:300px;white-space:normal">{safe(perms_text)}</td>
          <td>
            <a class="btn btn-secondary" href="/accounts/edit/{r['id']}">تعديل</a>
            <form class="inline-form" method="POST" action="/accounts/toggle/{r['id']}"><button class="btn btn-outline" type="submit">تفعيل/تعطيل</button></form>
          </td>
        </tr>
        """
    html += "</tbody></table></div>"
    return render_page("إدارة المستخدمين", html)


def permissions_checkboxes(selected=None):
    selected = set(selected or [])
    html = "<div class='permissions-grid'>"
    for p in PERMISSIONS:
        checked = "checked" if p in selected else ""
        html += f"<label class='info-note' style='display:flex;align-items:center;gap:8px'><input type='checkbox' name='permissions' value='{p}' {checked}> <span>{permission_label(p)}</span></label>"
    html += "</div>"
    return html


@app.route("/accounts/add", methods=["GET", "POST"])
@login_required
@permission_required("manage_accounts")
def add_account():
    if request.method == "POST":
        username = clean_csv_value(request.form.get("username"))
        password = clean_csv_value(request.form.get("password"))
        full_name = clean_csv_value(request.form.get("full_name"))
        perms = request.form.getlist("permissions")
        if not username or not password:
            flash("اسم المستخدم وكلمة المرور مطلوبان.", "error")
            return redirect(url_for("add_account"))
        try:
            row = execute_sql("""
                INSERT INTO app_accounts (username, password_hash, full_name, is_active)
                VALUES (%s,%s,%s,TRUE)
                RETURNING id
            """, [username, sha256_text(password), full_name], fetchone=True)
        except psycopg2.Error:
            flash("اسم المستخدم مستخدم مسبقًا أو يوجد خطأ في قاعدة البيانات.", "error")
            return redirect(url_for("add_account"))
        aid = row["id"]
        for p in perms:
            execute_sql("""
                INSERT INTO account_permissions (account_id, permission_id)
                SELECT %s, id FROM permissions WHERE name=%s
                ON CONFLICT DO NOTHING
            """, [aid, p])
        log_action("add_account", "account", aid, f"إنشاء حساب {username}")
        flash("تم إنشاء الحساب.", "success")
        return redirect(url_for("accounts_page"))
    content = f"""
    <div class="hero"><h1>إضافة مستخدم</h1><p>إنشاء مستخدم جديد وتحديد صلاحياته.</p></div>
    <div class="card">
      <form method="POST">
        <div class="row">
          <div><label>اسم المستخدم</label><input name="username" required></div>
          <div><label>كلمة المرور</label><input type="password" name="password" required></div>
          <div><label>الاسم الكامل</label><input name="full_name"></div>
        </div>
        <div style="margin-top:14px">
          <label>الصلاحيات</label>
          {permissions_checkboxes()}
        </div>
        <div class="actions" style="margin-top:14px">
          <button class="btn btn-primary" type="submit">حفظ</button>
          <a class="btn btn-outline" href="{url_for('accounts_page')}">إلغاء</a>
        </div>
      </form>
    </div>
    """
    return render_page("إضافة مستخدم", content)


@app.route("/accounts/edit/<int:account_id>", methods=["GET", "POST"])
@login_required
@permission_required("manage_accounts")
def edit_account(account_id):
    row = query_one("SELECT * FROM app_accounts WHERE id=%s", [account_id])
    if not row:
        flash("الحساب غير موجود.", "error")
        return redirect(url_for("accounts_page"))
    assigned = query_all("""
        SELECT p.name
        FROM account_permissions ap
        JOIN permissions p ON p.id = ap.permission_id
        WHERE ap.account_id=%s
    """, [account_id])
    assigned_names = [x["name"] for x in assigned]
    if request.method == "POST":
        username = clean_csv_value(request.form.get("username"))
        full_name = clean_csv_value(request.form.get("full_name"))
        password = clean_csv_value(request.form.get("password"))
        perms = request.form.getlist("permissions")
        try:
            execute_sql("UPDATE app_accounts SET username=%s, full_name=%s WHERE id=%s", [username, full_name, account_id])
        except psycopg2.Error:
            flash("تعذر تحديث الحساب. تأكد أن اسم المستخدم غير مكرر.", "error")
            return redirect(url_for("edit_account", account_id=account_id))
        if password:
            execute_sql("UPDATE app_accounts SET password_hash=%s WHERE id=%s", [sha256_text(password), account_id])
        execute_sql("DELETE FROM account_permissions WHERE account_id=%s", [account_id])
        for p in perms:
            execute_sql("""
                INSERT INTO account_permissions (account_id, permission_id)
                SELECT %s, id FROM permissions WHERE name=%s
                ON CONFLICT DO NOTHING
            """, [account_id, p])
        if session.get("account_id") == account_id:
            session["username"] = username
            session["full_name"] = full_name
            refresh_session_permissions(account_id)
        log_action("edit_account", "account", account_id, f"تعديل حساب {username}")
        flash("تم تحديث الحساب.", "success")
        return redirect(url_for("accounts_page"))
    content = f"""
    <div class="hero"><h1>تعديل مستخدم</h1><p>تعديل البيانات والصلاحيات وإعادة تعيين كلمة المرور عند الحاجة.</p></div>
    <div class="card">
      <form method="POST">
        <div class="row">
          <div><label>اسم المستخدم</label><input name="username" value="{safe(row['username'])}" required></div>
          <div><label>كلمة المرور الجديدة</label><input type="password" name="password" placeholder="اتركها فارغة بدون تغيير"></div>
          <div><label>الاسم الكامل</label><input name="full_name" value="{safe(row['full_name'])}"></div>
        </div>
        <div style="margin-top:14px">
          <label>الصلاحيات</label>
          {permissions_checkboxes(assigned_names)}
        </div>
        <div class="actions" style="margin-top:14px">
          <button class="btn btn-primary" type="submit">حفظ التعديلات</button>
          <a class="btn btn-outline" href="{url_for('accounts_page')}">رجوع</a>
        </div>
      </form>
    </div>
    """
    return render_page("تعديل مستخدم", content)


@app.route("/accounts/toggle/<int:account_id>", methods=["POST"])
@login_required
@permission_required("manage_accounts")
def toggle_account(account_id):
    execute_sql("UPDATE app_accounts SET is_active = NOT is_active WHERE id=%s", [account_id])
    log_action("toggle_account", "account", account_id, "تفعيل/تعطيل حساب")
    flash("تم تحديث حالة الحساب.", "success")
    return redirect(url_for("accounts_page"))


@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile_page():
    row = query_one("SELECT * FROM app_accounts WHERE id=%s", [session["account_id"]])
    perms = query_all("""
        SELECT p.name FROM account_permissions ap
        JOIN permissions p ON p.id = ap.permission_id
        WHERE ap.account_id=%s
        ORDER BY p.name
    """, [session["account_id"]])
    if request.method == "POST":
        current_password = clean_csv_value(request.form.get("current_password"))
        new_password = clean_csv_value(request.form.get("new_password"))
        full_name = clean_csv_value(request.form.get("full_name"))
        if sha256_text(current_password) != row["password_hash"]:
            flash("كلمة المرور الحالية غير صحيحة.", "error")
            return redirect(url_for("profile_page"))
        if not new_password:
            flash("كلمة المرور الجديدة مطلوبة.", "error")
            return redirect(url_for("profile_page"))
        execute_sql("UPDATE app_accounts SET full_name=%s, password_hash=%s WHERE id=%s", [full_name, sha256_text(new_password), session["account_id"]])
        session["full_name"] = full_name
        log_action("change_password", "account", session["account_id"], "تغيير كلمة المرور من الصفحة الشخصية")
        flash("تم تحديث بياناتك وكلمة المرور.", "success")
        return redirect(url_for("profile_page"))
    perm_html = "".join([f"<span class='badge'>{p['name']}</span> " for p in perms]) or "<span class='small'>لا توجد صلاحيات.</span>"
    content = f"""
    <div class="hero"><h1>صفحتي الشخصية</h1><p>تعديل الاسم الكامل وكلمة المرور.</p></div>
    <div class="card">
      <div class="small" style="margin-bottom:10px">الصلاحيات الحالية: {perm_html}</div>
      <form method="POST">
        <div class="row">
          <div><label>اسم المستخدم</label><input value="{safe(row['username'])}" disabled></div>
          <div><label>الاسم الكامل</label><input name="full_name" value="{safe(row['full_name'])}"></div>
          <div><label>كلمة المرور الحالية</label><input type="password" name="current_password" required></div>
          <div><label>كلمة المرور الجديدة</label><input type="password" name="new_password" required></div>
        </div>
        <div class="actions" style="margin-top:14px"><button class="btn btn-primary" type="submit">حفظ</button></div>
      </form>
    </div>
    """
    return render_page("صفحتي الشخصية", content)




@app.route("/timer")
@login_required
def power_timer_page():
    content = f"""
    <div class="hero">
      <h1>مؤقت الكهرباء</h1>
      <p>مؤقت واحد للدورة الحالية. عند انتهاء الوقت يظهر تنبيه وصوت، ثم يعيد التشغيل تلقائيًا بعد 10 ثوانٍ لنفس المدة المختارة.</p>
    </div>
    <div class="grid-2">
      <div class="card timer-visual-card">
        <div class="actions" style="justify-content:space-between;align-items:flex-start;gap:14px">
          <div>
            <h3 style="margin:0 0 8px 0">الحالة الحالية</h3>
            <div id="timer-page-state" class="timer-state-badge timer-status-stopped">متوقف</div>
          </div>
          <div class="topbar-clock"><i class="fa-regular fa-clock"></i><div><div id="timer-page-live-time">--:--:--</div><small id="timer-page-live-date">--/--/----</small></div></div>
        </div>
        <div class="timer-ring-wrap">
          <div class="timer-ring">
            <svg viewBox="0 0 300 300" aria-hidden="true">
              <defs>
                <linearGradient id="timerRingGradient" x1="0%" y1="0%" x2="100%" y2="100%">
                  <stop offset="0%" stop-color="#35a7e8"></stop>
                  <stop offset="100%" stop-color="#6d4ee8"></stop>
                </linearGradient>
              </defs>
              <circle class="bg" cx="150" cy="150" r="130"></circle>
              <circle id="timer-ring-progress" class="progress" cx="150" cy="150" r="130"></circle>
            </svg>
            <div class="timer-ring-center">
              <div class="timer-ring-label">الوقت المتبقي</div>
              <div id="timer-page-remaining" class="timer-big timer-big-ring">--:--</div>
              <div id="timer-page-phase" class="timer-ring-phase">-</div>
            </div>
          </div>
        </div>
        <div class="timer-progress-strip"><div id="timer-progress-fill" class="timer-progress-fill"></div></div>
        <div class="timer-meta">
          <div class="timer-meta-card"><h4>الدورة الحالية</h4><div class="v" id="timer-page-cycle-minutes">30</div></div>
          <div class="timer-meta-card"><h4>إعادة التشغيل</h4><div class="v"><span id="timer-page-restart">10</span><span style="font-size:18px">ث</span></div></div>
          <div class="timer-meta-card"><h4>النمط</h4><div class="v" style="font-size:20px">تلقائي</div></div>
        </div>
      </div>
      <div class="card">
        <h3 style="margin-top:0">التحكم بالمؤقت</h3>
        <p class="small" style="margin-top:0">واجهة أنيقة للدور الكهربائي مع إعادة تشغيل تلقائي بعد انتهاء كل دورة.</p>
        <div class="row">
          <div><label>مدة الدورة بالدقائق</label><input id="timer-minutes-input" type="number" min="1" step="1" value="30"></div>
          <div><label>إعادة التشغيل التلقائي</label><input value="10 ثوانٍ بعد انتهاء الوقت" disabled></div>
        </div>
        <div class="actions" style="margin-top:14px">
          <button id="timer-start-btn" class="btn btn-primary" type="button" onclick="return startPowerTimer()"><i class="fa-solid fa-play"></i> بدء / إعادة ضبط</button>
          <button id="timer-pause-btn" class="btn btn-accent" type="button" onclick="return pausePowerTimer()" style="display:none"><i class="fa-solid fa-pause"></i> إيقاف مؤقت</button>
          <button id="timer-resume-btn" class="btn btn-secondary" type="button" onclick="return resumePowerTimer()" style="display:none"><i class="fa-solid fa-play"></i> استئناف</button>
          <button id="timer-stop-btn" class="btn btn-danger" type="button" onclick="return stopPowerTimer()" style="display:none"><i class="fa-solid fa-stop"></i> إيقاف نهائي</button>
        </div>
        <div class="info-note" style="margin-top:16px">
          <strong>آلية العمل</strong>
          <div class="small" style="margin-top:6px;line-height:1.8">عند انتهاء الوقت يظهر تنبيه مرئي وصوتي داخل الموقع، ثم يبدأ العد تلقائيًا من جديد بعد 10 ثوانٍ على نفس المدة المختارة.</div>
        </div>
      </div>
    </div>
    """
    return render_page("مؤقت الكهرباء", content)


@app.route("/api/power-timer/status")
@login_required
def power_timer_status_api():
    return jsonify(build_power_timer_status())


@app.route("/api/power-timer/start", methods=["POST"])
@login_required
def power_timer_start_api():
    minutes_raw = clean_csv_value(request.form.get("minutes", "30")) or "30"
    try:
        minutes = max(1, min(24 * 60, int(minutes_raw)))
    except ValueError:
        return jsonify({"ok": False, "message": "قيمة الدقائق غير صحيحة."}), 400
    execute_sql("""
        UPDATE power_timer
        SET duration_minutes=%s, cycle_started_at=CURRENT_TIMESTAMP, paused_remaining_seconds=NULL,
            state='running', auto_restart_delay_seconds=10, updated_by_username=%s, updated_at=CURRENT_TIMESTAMP
        WHERE id=1
    """, [minutes, session.get("username", "")])
    log_action("power_timer_start", "power_timer", 1, f"تشغيل المؤقت لمدة {minutes} دقيقة")
    return jsonify({"ok": True, "message": f"تم تشغيل المؤقت لمدة {minutes} دقيقة."})


@app.route("/api/power-timer/pause", methods=["POST"])
@login_required
def power_timer_pause_api():
    status = build_power_timer_status()
    if status["state"] != "running":
        return jsonify({"ok": False, "message": "المؤقت ليس في حالة تشغيل."}), 400
    remaining = int(status["display_remaining_seconds"] or 0)
    execute_sql("""
        UPDATE power_timer
        SET paused_remaining_seconds=%s, state='paused', updated_by_username=%s, updated_at=CURRENT_TIMESTAMP
        WHERE id=1
    """, [remaining, session.get("username", "")])
    log_action("power_timer_pause", "power_timer", 1, f"إيقاف مؤقت للمؤقت والمتبقي {remaining} ثانية")
    return jsonify({"ok": True, "message": "تم إيقاف المؤقت مؤقتًا."})


@app.route("/api/power-timer/resume", methods=["POST"])
@login_required
def power_timer_resume_api():
    row = get_power_timer_row()
    if (row.get("state") or "") != "paused":
        return jsonify({"ok": False, "message": "المؤقت ليس متوقفًا مؤقتًا."}), 400
    duration_minutes = int(row.get("duration_minutes") or 30)
    duration_seconds = max(60, duration_minutes * 60)
    remaining = int(row.get("paused_remaining_seconds") or duration_seconds)
    elapsed_before_pause = max(0, duration_seconds - remaining)
    execute_sql("""
        UPDATE power_timer
        SET cycle_started_at=(CURRENT_TIMESTAMP - (%s * INTERVAL '1 second')), paused_remaining_seconds=NULL,
            state='running', updated_by_username=%s, updated_at=CURRENT_TIMESTAMP
        WHERE id=1
    """, [elapsed_before_pause, session.get("username", "")])
    log_action("power_timer_resume", "power_timer", 1, "استئناف المؤقت")
    return jsonify({"ok": True, "message": "تم استئناف المؤقت."})


@app.route("/api/power-timer/stop", methods=["POST"])
@login_required
def power_timer_stop_api():
    execute_sql("""
        UPDATE power_timer
        SET state='stopped', cycle_started_at=NULL, paused_remaining_seconds=NULL, updated_by_username=%s, updated_at=CURRENT_TIMESTAMP
        WHERE id=1
    """, [session.get("username", "")])
    log_action("power_timer_stop", "power_timer", 1, "إيقاف نهائي للمؤقت")
    return jsonify({"ok": True, "message": "تم إيقاف المؤقت نهائيًا."})


@app.route("/audit-log")
@login_required
@permission_required("view_audit_log")
def audit_log_page():
    page = max(1, int(request.args.get("page", "1") or "1"))
    per_page = 40
    total = query_one("SELECT COUNT(*) AS c FROM audit_logs")["c"]
    pages = max(1, math.ceil(total / per_page))
    offset = (page - 1) * per_page
    rows = query_all("""
        SELECT * FROM audit_logs
        ORDER BY id DESC
        LIMIT %s OFFSET %s
    """, [per_page, offset])
    html = "<div class='hero'><h1>سجل العمليات</h1><p>كل العمليات الحساسة داخل النظام.</p></div><div class='card'><table><thead><tr><th>ID</th><th>المستخدم</th><th>العملية</th><th>الهدف</th><th>الهدف ID</th><th>التفاصيل</th><th>الوقت</th></tr></thead><tbody>"
    for r in rows:
        html += f"<tr><td>{r['id']}</td><td>{safe(r['username_snapshot'])}</td><td>{action_type_label(r['action_type'])}</td><td>{target_type_label(r['target_type'])}</td><td>{safe(r['target_id'])}</td><td>{safe(r['details'])}</td><td>{format_dt_compact(r['created_at'])}</td></tr>"
    html += "</tbody></table>"
    html += "<div class='pagination'>"
    for p in range(1, pages + 1):
        cls = "active" if p == page else ""
        html += f"<a class='{cls}' href='?page={p}'>{p}</a>"
    html += "</div></div>"
    return render_page("سجل العمليات", html)


# ===== Runtime UI/UX patch =====
# CSS tweaks
BASE_TEMPLATE = BASE_TEMPLATE.replace(
    ".notes-box{min-height:90px}\n.ajax-saving{opacity:.65;pointer-events:none}",
    ".notes-box{min-height:90px}\n.note-preview{display:inline-flex;align-items:center;gap:6px;max-width:260px}\n.note-text{display:inline-block;max-width:190px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}\n.type-tabs{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:14px}\n.type-tab{padding:10px 14px;border-radius:999px;border:1px solid var(--line);background:#fff;color:var(--text);cursor:pointer;font-weight:700}\n.type-tab.active{background:var(--primary);color:#fff;border-color:var(--primary)}\n.ajax-saving{opacity:.65;pointer-events:none}"
)
# JS tweaks
BASE_TEMPLATE = BASE_TEMPLATE.replace(
"""function initBeneficiaryForms(){
  document.querySelectorAll('[data-beneficiary-scope]').forEach(function(scope){
    var input = scope.querySelector('select[name=\"user_type\"], input[name=\"user_type\"]');
    if(input){toggleBeneficiarySections(input, scope.id);}
  });
}
""",
"""function syncTypeTabs(scopeId, value){
  var container = document.getElementById(scopeId);
  if(!container) return;
  container.querySelectorAll('.type-tab').forEach(function(btn){ btn.classList.toggle('active', btn.dataset.value === value); });
}
function setBeneficiaryType(scopeId, value){
  var container = document.getElementById(scopeId);
  if(!container) return false;
  var input = container.querySelector('select[name=\"user_type\"], input[name=\"user_type\"]');
  if(input){ input.value = value; toggleBeneficiarySections(input, scopeId); }
  syncTypeTabs(scopeId, value);
  return false;
}
function initBeneficiaryForms(){
  document.querySelectorAll('[data-beneficiary-scope]').forEach(function(scope){
    var input = scope.querySelector('select[name=\"user_type\"], input[name=\"user_type\"]');
    if(input){toggleBeneficiarySections(input, scope.id); syncTypeTabs(scope.id, input.value || 'tawjihi');}
  });
}
"""
)
BASE_TEMPLATE = BASE_TEMPLATE.replace(
"""async function submitBeneficiaryEdit(form, rowId, modalId){
  form.classList.add('ajax-saving');
  try{
    const data = await ajaxPost(form.action, new FormData(form));
    replaceRowAndModal(data, rowId, modalId);
    window.location.hash = '#!';
  }catch(err){
    showLiveFlash(err.message || 'تعذر حفظ التعديل', 'error');
  }finally{
    form.classList.remove('ajax-saving');
  }
  return false;
}
""",
"""async function submitBeneficiaryEdit(form, rowId, modalId){
  form.classList.add('ajax-saving');
  try{
    const data = await ajaxPost(form.action, new FormData(form));
    replaceRowAndModal(data, rowId, modalId);
    window.location.hash = '#!';
  }catch(err){
    showLiveFlash(err.message || 'تعذر حفظ التعديل', 'error');
  }finally{
    form.classList.remove('ajax-saving');
  }
  return false;
}
function prependRowToTable(rowHtml){
  const tbody = document.querySelector('.table-wrap table tbody');
  if(!tbody || !rowHtml) return;
  const empty = tbody.querySelector('.empty-state');
  if(empty){ empty.parentElement.remove(); }
  tbody.insertAdjacentHTML('afterbegin', rowHtml);
}
function insertModalHtml(modalHtml){ if(modalHtml){ document.body.insertAdjacentHTML('beforeend', modalHtml); } }
function submitBeneficiaryAdd(form){ return guardSingleSubmit(form); }
async function resetWeeklyUsageAjax(url){
  try{
    const data = await ajaxPost(url);
    document.querySelectorAll('tr[id^="beneficiary-row-"]').forEach(function(row){
      if(row.dataset.limited === '1'){
        row.classList.remove('row-complete');
        const cell = row.querySelector('.usage-cell');
        if(cell){ cell.textContent = '0 / 3'; }
      }
    });
    if(data.message){ showLiveFlash(data.message, data.category || 'success'); }
  }catch(err){ showLiveFlash(err.message || 'تعذر تجديد البطاقات', 'error'); }
  return false;
}
"""
)

def truncate_note_words(value, limit=3):
    text = clean_csv_value(value)
    if not text:
        return "-", False, ""
    words = text.split()
    if len(words) <= limit:
        safe_text = safe(text)
        return safe_text, False, safe_text
    return safe(" ".join(words[:limit]) + " ..."), True, safe(text)


def build_add_beneficiary_modal(selected_type='tawjihi'):
    selected_type = clean_csv_value(selected_type) or 'tawjihi'
    modal_form = format_modal_fields(
        {'user_type': selected_type},
        action=url_for('add_beneficiary_page'),
        scope_id='add-beneficiary-scope',
        submit_label='حفظ المستفيد',
        show_type_selector=False,
        fixed_user_type=selected_type,
    )
    modal_form = re.sub(
        r'<div>\s*<label>النوع</label>\s*<input value=".*?" disabled>\s*<input type="hidden" name="user_type" value=".*?">\s*</div>',
        '',
        modal_form,
        count=1,
        flags=re.DOTALL,
    )
    tabs_html = (
        "<div class='type-tabs'>"
        + '<button type=\'button\' class=\'type-tab {cls}\' data-value=\'tawjihi\' onclick="return setBeneficiaryType(\'add-beneficiary-scope\',\'tawjihi\')">توجيهي</button>'.format(cls=('active' if selected_type == 'tawjihi' else ''))
        + '<button type=\'button\' class=\'type-tab {cls}\' data-value=\'university\' onclick="return setBeneficiaryType(\'add-beneficiary-scope\',\'university\')">جامعة</button>'.format(cls=('active' if selected_type == 'university' else ''))
        + '<button type=\'button\' class=\'type-tab {cls}\' data-value=\'freelancer\' onclick="return setBeneficiaryType(\'add-beneficiary-scope\',\'freelancer\')">فري لانسر</button>'.format(cls=('active' if selected_type == 'freelancer' else ''))
        + '</div>'
        + f"<input type='hidden' name='user_type' value='{selected_type}'>"
    )
    modal_form = modal_form.replace(
        '<div id="add-beneficiary-scope" data-beneficiary-scope="1">',
        '<div id="add-beneficiary-scope" data-beneficiary-scope="1">' + tabs_html,
        1,
    )
    modal_form = modal_form.replace('<form method="POST" action="', '<form method="POST" onsubmit="return submitBeneficiaryAdd(this)" action="', 1)
    return f'''
    <div id="add-beneficiary-modal" class="modal">
      <div class="modal-card">
        <a href="#!" class="modal-close">×</a>
        <div class="hero" style="margin-bottom:14px"><h1>إضافة مستفيد</h1><p>اختر النوع من التبويبات أعلى النموذج ثم احفظ بدون إعادة تحميل الصفحة.</p></div>
        {modal_form}
      </div>
    </div>
    '''

def build_beneficiary_row_html(r, selected_type, args_dict, page=1, display_index=None):
    usage_label, limited, count = get_usage_label(r)
    current_qs = build_query_string({**args_dict, "page": page})
    row_class = f"row-type-{safe(r.get('user_type'))}"
    if limited and count >= 3:
        row_class += " row-complete"
    modal_id = f"edit-{r['id']}"
    note_modal_id = f"note-{r['id']}"
    actions = []
    modal_html = ""
    if has_permission("edit"):
        edit_action = f"{url_for('edit_beneficiary_page', beneficiary_id=r['id'])}?current_user_type={selected_type}"
        modal_body = format_modal_fields(
            r,
            action=edit_action,
            scope_id=f"scope-{r['id']}",
            submit_label="حفظ التعديلات",
            show_type_selector=True,
        )
        edit_onsubmit = f"<form method=\"POST\" onsubmit=\"return submitBeneficiaryEdit(this, {r['id']}, '{modal_id}')\" action=\""
        modal_body = modal_body.replace('<form method="POST" action="', edit_onsubmit, 1)
        modal_html = f"""
        <div id="{modal_id}" class="modal">
          <div class="modal-card">
            <a href="#!" class="modal-close">×</a>
            <div class="hero" style="margin-bottom:14px"><h1>تعديل المستفيد #{r['id']}</h1><p>{safe(r.get('full_name'))}</p></div>
            {modal_body}
          </div>
        </div>
        """
        actions.append(f"<a class='btn btn-secondary btn-icon' href='#{modal_id}' title='تعديل'><i class='fa-solid fa-pen'></i></a>")
    if has_permission("delete"):
        delete_url = url_for('delete_beneficiary', beneficiary_id=r['id'])
        actions.append(f"<form class='inline-form' method='POST' action='{delete_url}' onsubmit=\"return confirm('هل أنت متأكد من الحذف؟')\"><button class='btn btn-danger btn-icon' type='submit' title='حذف'><i class='fa-solid fa-trash'></i></button></form>")
    if has_permission("usage_counter") and limited and count < 3:
        usage_url = f"{url_for('add_usage', beneficiary_id=r['id'])}?{current_qs}"
        actions.append(f"<button class='btn btn-accent btn-icon' type='button' onclick=\"return incrementUsageAjax('{usage_url}', {r['id']}, '{modal_id}')\" title='+1 بطاقة'><i class='fa-solid fa-plus'></i></button>")
    type_html = f"<span class='type-badge {get_type_css(r.get('user_type'))}'>{get_type_label(r.get('user_type'))}</span>"
    added_by = safe(r.get('added_by_username')) or '-'
    created_at = format_dt_short(r.get('created_at'))
    note_preview, has_note_modal, note_full = truncate_note_words(r.get('notes'))
    notes_html = note_preview
    if has_note_modal:
        notes_html = f"<span class='note-preview'><span class='note-text'>{note_preview}</span><a class='btn btn-soft btn-icon' href='#{note_modal_id}' title='عرض الملاحظة'><i class='fa-solid fa-eye'></i></a></span>"
        modal_html += f"""
        <div id="{note_modal_id}" class="modal">
          <div class="modal-card" style="width:min(650px,100%)">
            <a href="#!" class="modal-close">×</a>
            <div class="hero" style="margin-bottom:14px"><h1>ملاحظة المستفيد</h1><p>{safe(r.get('full_name'))}</p></div>
            <div class="card"><div class='cell-wrap' style='max-width:none'>{note_full}</div></div>
          </div>
        </div>
        """
    display_no = safe(display_index if display_index is not None else r.get('id'))
    if selected_type == "tawjihi":
        row_cells = [display_no, safe(r['full_name']), safe(r['phone']), safe(r['tawjihi_year']), safe(r['tawjihi_branch']), usage_label, added_by, created_at, notes_html, " ".join(actions) if actions else "-"]
        usage_idx = 5
    elif selected_type == "university":
        row_cells = [display_no, safe(r['full_name']), safe(r['phone']), safe(r['university_name']), safe(r['university_college']), safe(r['university_specialization']), usage_label, added_by, created_at, notes_html, " ".join(actions) if actions else "-"]
        usage_idx = 6
    elif selected_type == "freelancer":
        row_cells = [display_no, safe(r['full_name']), safe(r['phone']), safe(r['freelancer_specialization']), safe(r['freelancer_company']), usage_label, added_by, created_at, notes_html, " ".join(actions) if actions else "-"]
        usage_idx = 5
    else:
        row_cells = [display_no, safe(r['full_name']), safe(r['phone']), type_html, usage_label, added_by, created_at, notes_html, " ".join(actions) if actions else "-"]
        usage_idx = 4
    cells = []
    for i, cell in enumerate(row_cells):
        extra = ' usage-cell' if i == usage_idx else ''
        cells.append(f"<td class='cell-wrap{extra}'>{cell}</td>")
    can_bulk_select = has_permission('delete') or has_permission('export')
    if can_bulk_select:
        cells.append(f"<td class='checkbox-cell'><input class='row-select' type='checkbox' value='{r['id']}' onchange='updateBulkSelectedCount()'></td>")
    row_html = f"<tr id='beneficiary-row-{r['id']}' class='{row_class}' data-limited={'1' if limited else '0'}>" + ''.join(cells) + "</tr>"
    return row_html, modal_html


def _patched_add_beneficiary_page():
    page_title_map = {
        'tawjihi': 'إضافة طالب توجيهي',
        'university': 'إضافة طالب جامعي',
        'freelancer': 'إضافة فري لانسر',
    }
    if request.method == 'POST':
        data = collect_beneficiary_form()
        duplicate = find_duplicate_phone(data.get('phone'))
        if duplicate:
            message = f"رقم الجوال مستخدم مسبقًا للمستفيد: {safe(duplicate.get('full_name'))}."
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'ok': False, 'message': message, 'category': 'error'}), 400
            flash(message, 'error')
            return render_page('إضافة مستفيد', beneficiary_form_html(data, action=url_for('add_beneficiary_page'), title=page_title_map.get(data.get('user_type') or clean_csv_value(request.args.get('user_type', 'tawjihi')) or 'tawjihi', 'إضافة مستفيد')))
        data['added_by_account_id'] = session.get('account_id')
        data['added_by_username'] = session.get('username')
        row = execute_sql("""
            INSERT INTO beneficiaries (
                user_type, first_name, second_name, third_name, fourth_name, full_name, search_name, phone,
                tawjihi_year, tawjihi_branch, freelancer_specialization, freelancer_company,
                freelancer_schedule_type, freelancer_internet_method, freelancer_time_mode,
                freelancer_time_from, freelancer_time_to, university_name, university_college,
                university_specialization, university_days, university_internet_method,
                university_time_mode, university_time_from, university_time_to,
                weekly_usage_count, weekly_usage_week_start, notes, added_by_account_id, added_by_username
            ) VALUES (
                %(user_type)s, %(first_name)s, %(second_name)s, %(third_name)s, %(fourth_name)s, %(full_name)s, %(search_name)s, %(phone)s,
                %(tawjihi_year)s, %(tawjihi_branch)s, %(freelancer_specialization)s, %(freelancer_company)s,
                %(freelancer_schedule_type)s, %(freelancer_internet_method)s, %(freelancer_time_mode)s,
                %(freelancer_time_from)s, %(freelancer_time_to)s, %(university_name)s, %(university_college)s,
                %(university_specialization)s, %(university_days)s, %(university_internet_method)s,
                %(university_time_mode)s, %(university_time_from)s, %(university_time_to)s,
                0, %(weekly_usage_week_start)s, %(notes)s, %(added_by_account_id)s, %(added_by_username)s
            ) RETURNING id
        """, data, fetchone=True)
        new_id = row['id'] if row else None
        log_action('add', 'beneficiary', new_id, f"إضافة مستفيد: {data['full_name']}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            inserted_row = query_one('SELECT * FROM beneficiaries WHERE id=%s', [new_id]) if new_id else None
            current_type = clean_csv_value(request.args.get('current_user_type', ''))
            args_dict = build_request_args_dict()
            row_html, modal_html = ('', '')
            if inserted_row and (not current_type or current_type == inserted_row.get('user_type')):
                row_html, modal_html = build_beneficiary_row_html(inserted_row, current_type, args_dict, page=1)
            return jsonify({'ok': True, 'row_html': row_html, 'modal_html': modal_html, 'message': 'تمت إضافة المستفيد.', 'category': 'success', 'reset_form_modal': build_add_beneficiary_modal(current_type or (inserted_row.get('user_type') if inserted_row else 'tawjihi'))})
        flash('تمت إضافة المستفيد.', 'success')
        return redirect(url_for('beneficiaries_page', user_type=data['user_type']))
    page_user_type = clean_csv_value(request.args.get('user_type', 'tawjihi')) or 'tawjihi'
    return render_page('إضافة مستفيد', beneficiary_form_html(action=url_for('add_beneficiary_page'), title=page_title_map.get(page_user_type, 'إضافة مستفيد')))


def _patched_reset_weekly_usage():
    reset_start = get_week_start()
    execute_sql("UPDATE beneficiaries SET weekly_usage_count = 0, weekly_usage_week_start = %s", [reset_start])
    log_action('reset_weekly_usage', 'beneficiary', None, f'تصفير يدوي لكل البطاقات - week_start={reset_start}')
    message = 'تم تجديد جميع البطاقات يدويًا بنجاح.'
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'ok': True, 'message': message, 'category': 'success'})
    flash(message, 'success')
    return redirect(request.referrer or url_for('beneficiaries_page'))


# monkey-patch flask endpoints
app.view_functions['add_beneficiary_page'] = login_required(permission_required('add')(_patched_add_beneficiary_page))
app.view_functions['reset_weekly_usage'] = login_required(permission_required('reset_weekly_usage')(_patched_reset_weekly_usage))



@app.route("/beneficiaries/bulk-delete", methods=["POST"])
@login_required
@permission_required("delete")
def bulk_delete_beneficiaries():
    ids_raw = clean_csv_value(request.form.get("ids", ""))
    ids = [int(x) for x in ids_raw.split(',') if x.strip().isdigit()]
    if not ids:
        flash("لم يتم تحديد أي مستفيد.", "error")
        return redirect(request.referrer or url_for("beneficiaries_page"))
    execute_sql("DELETE FROM beneficiaries WHERE id = ANY(%s)", [ids])
    log_action("bulk_delete", "beneficiary", None, f"حذف جماعي لعدد {len(ids)} مستفيد")
    flash(f"تم حذف {len(ids)} مستفيد بنجاح.", "success")
    return redirect(request.referrer or url_for("beneficiaries_page"))


@app.route("/beneficiaries/export-selected", methods=["POST"])
@login_required
@permission_required("export")
def export_selected_beneficiaries():
    ids_raw = clean_csv_value(request.form.get("ids", ""))
    ids = [int(x) for x in ids_raw.split(',') if x.strip().isdigit()]
    if not ids:
        flash("لم يتم تحديد أي مستفيد.", "error")
        return redirect(request.referrer or url_for("beneficiaries_page"))
    rows = query_all("SELECT * FROM beneficiaries WHERE id = ANY(%s) ORDER BY id DESC", [ids])
    wb = Workbook()
    ws = wb.active
    ws.title = "Selected Beneficiaries"
    headers = ["ID","النوع","الاسم الكامل","الجوال","سنة التوجيهي","فرع التوجيهي","الجامعة","الكلية","التخصص الجامعي","تخصص الفري لانسر","شركة الفري لانسر","الاستخدام","أضيف بواسطة","التاريخ","ملاحظات"]
    ws.append(headers)
    for r in rows:
        usage_label, _, _ = get_usage_label(r)
        ws.append([r.get("id"), get_type_label(r.get("user_type")), r.get("full_name"), r.get("phone"), r.get("tawjihi_year"), r.get("tawjihi_branch"), r.get("university_name"), r.get("university_college"), r.get("university_specialization"), r.get("freelancer_specialization"), r.get("freelancer_company"), usage_label, r.get("added_by_username"), format_dt_short(r.get("created_at")), r.get("notes")])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    response = Response(out.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = "attachment; filename=selected_beneficiaries.xlsx"
    return response


import os

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
